# -*- coding: utf-8 -*-
"""
Script Mejorado para Extracción de Issues SAP con Mejor Manejo de Scroll
---
Versión 18: Optimizado para rendimiento, mejor manejo de excepciones,
y procesamiento eficiente de datos.
"""

import time
import os
import sys
import logging
import sqlite3
import threading
import re
import json
from datetime import datetime
import webbrowser
import base64
from io import BytesIO

# Importaciones opcionales con gestión de errores
try:
    import pandas as pd
except ImportError:
    pd = None
    print("Pandas no está instalado. La funcionalidad de Excel no estará disponible.")

# Intentar importar PIL, pero no fallar si no está disponible
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Selenium imports
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
    JavascriptException,
    WebDriverException
)

# Tkinter imports
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Configurar logging
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

log_file = os.path.join(
    log_dir, f"extraccion_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(log_file), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# Colores estilo SAP
SAP_COLORS = {
    "primary": "#1870C5",    # Azul SAP
    "secondary": "#354A5F",  # Azul oscuro SAP
    "success": "#107E3E",    # Verde SAP
    "warning": "#E9730C",    # Naranja SAP
    "danger": "#BB0000",     # Rojo SAP
    "light": "#F5F6F7",      # Gris claro SAP
    "dark": "#32363A",       # Gris oscuro SAP
    "white": "#FFFFFF",
    "gray": "#D3D7D9",
    "text": "#000000"        # Texto negro para máximo contraste
}










class DatabaseManager:
    """Clase dedicada al manejo de la base de datos de clientes y proyectos"""
    
    def __init__(self, db_path=None):
        """Inicializa el administrador de base de datos"""
        if db_path is None:
            db_dir = "data"
            if not os.path.exists(db_dir):
                os.makedirs(db_dir)
            db_path = os.path.join(db_dir, "sap_extraction.db")
            
        self.db_path = db_path
        self.setup_database()
    
    def setup_database(self):
        """Configura la estructura de la base de datos"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()


            # Crear tablas si no existen
            cursor.execute('''
            CREATE TABLE IF NOT EXISTS clients (
                erp_number TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                business_partner TEXT,
                last_used TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''')

            cursor.execute('''
            CREATE TABLE IF NOT EXISTS projects (
                project_id TEXT PRIMARY KEY,
                client_erp TEXT NOT NULL,
                name TEXT NOT NULL,
                engagement_case TEXT,
                last_used TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (client_erp) REFERENCES clients(erp_number)
            )
            ''')

            conn.commit()
            logger.debug("Base de datos configurada correctamente")
            return True
        except sqlite3.Error as e:
            logger.error(f"Error al configurar la base de datos: {e}")
            return False
        finally:
            if conn:
                conn.close()

        
    def get_clients(self):
        """Obtiene la lista de clientes ordenados por último uso"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("SELECT erp_number, name FROM clients ORDER BY last_used DESC")
        clients = cursor.fetchall()

        conn.close()

        return [f"{erp} - {name}" for erp, name in clients]
    
    def get_projects(self, client_erp):
        """Obtiene la lista de proyectos para un cliente específico"""
        if not client_erp:
            return []
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT project_id, name 
            FROM projects 
            WHERE client_erp = ? 
            ORDER BY last_used DESC
        """, (client_erp,))

        projects = cursor.fetchall()

        conn.close()

        return [f"{pid} - {name}" for pid, name in projects]
    
    def save_client(self, erp_number, name, business_partner=""):
        """Guarda o actualiza un cliente en la base de datos con mejoras de seguridad"""
        # Validación preliminar de todos los argumentos
        if not self.validate_input(erp_number, "erp"):
            logger.error(f"Número ERP inválido: {erp_number}")
            return False
            
        if not name or not isinstance(name, str):
            logger.error(f"Nombre de cliente inválido: {name}")
            return False
        
        # Limpiar y truncar datos si son demasiado largos
        name = name.strip()[:100]  # Limitar longitud para prevenir ataques
        business_partner = (business_partner or "").strip()[:50]
        
        conn = None
        try:
            conn = sqlite3.connect(self.db_path)
            conn.execute("PRAGMA foreign_keys = ON")  # Asegurar que se verifican las claves foráneas
            cursor = conn.cursor()

            # Verificar si el cliente ya existe
            cursor.execute("SELECT erp_number FROM clients WHERE erp_number = ?", (erp_number,))
            existing = cursor.fetchone()

            if existing:
                # Actualizar cliente existente
                cursor.execute("""
                    UPDATE clients 
                    SET name = ?, business_partner = ?, last_used = CURRENT_TIMESTAMP 
                    WHERE erp_number = ?
                """, (name, business_partner, erp_number))
                logger.info(f"Cliente actualizado: {erp_number} - {name}")
            else:
                # Insertar nuevo cliente
                cursor.execute("""
                    INSERT INTO clients (erp_number, name, business_partner, last_used) 
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                """, (erp_number, name, business_partner))
                logger.info(f"Nuevo cliente creado: {erp_number} - {name}")

            conn.commit()
            return True
        except sqlite3.IntegrityError as ie:
            logger.error(f"Error de integridad de datos al guardar cliente: {ie}")
            if conn:
                conn.rollback()
            return False
        except sqlite3.Error as e:
            logger.error(f"Error SQL al guardar cliente: {e}")
            if conn:
                conn.rollback()
            return False
        except Exception as e:
            logger.error(f"Error general al guardar cliente: {e}")
            if conn:
                conn.rollback()
            return False
        finally:
            if conn:
                conn.close()
                
                
                
                
                
                
                    
    def save_project(self, project_id, client_erp, name, engagement_case=""):
        """Guarda o actualiza un proyecto en la base de datos"""
        if not self.validate_input(project_id, "project") or not self.validate_input(client_erp, "erp"):
            logger.error(f"ID de proyecto o cliente inválido: {project_id}, {client_erp}")
            return False
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            # Verificar si el proyecto ya existe
            cursor.execute("SELECT * FROM projects WHERE project_id = ?", (project_id,))
            existing = cursor.fetchone()

            if existing:
                # Actualizar proyecto existente
                cursor.execute("""
                    UPDATE projects 
                    SET client_erp = ?, name = ?, engagement_case = ?, last_used = CURRENT_TIMESTAMP 
                    WHERE project_id = ?
                """, (client_erp, name, engagement_case, project_id))
            else:
                # Insertar nuevo proyecto
                cursor.execute("""
                    INSERT INTO projects (project_id, client_erp, name, engagement_case, last_used) 
                    VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (project_id, client_erp, name, engagement_case))

            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al guardar proyecto en BD: {e}")
            return False
        finally:
            conn.close()
    
    def update_client_usage(self, erp_number):
        """Actualiza la fecha de último uso de un cliente"""
        if not self.validate_input(erp_number, "erp"):
            return False
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("""
                UPDATE clients 
                SET last_used = CURRENT_TIMESTAMP 
                WHERE erp_number = ?
            """, (erp_number,))
            
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al actualizar uso de cliente: {e}")
            return False
        finally:
            conn.close()
    
    def update_project_usage(self, project_id):
        """Actualiza la fecha de último uso de un proyecto"""
        if not self.validate_input(project_id, "project"):
            return False
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("""
                UPDATE projects 
                SET last_used = CURRENT_TIMESTAMP 
                WHERE project_id = ?
            """, (project_id,))
            
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al actualizar uso de proyecto: {e}")
            return False
        finally:
            conn.close()
    
    @staticmethod
    def validate_input(input_str, input_type="general"):
        """Valida las entradas para prevenir inyecciones SQL"""
        if input_type == "erp":
            # Solo permitir dígitos para ERP
            return bool(re.match(r'^\d+$', str(input_str)))
        elif input_type == "project":
            # Solo permitir dígitos para ID de proyecto
            return bool(re.match(r'^\d+$', str(input_str)))
        elif input_type == "path":
            # Validar ruta de archivo
            return os.path.isabs(input_str) and not any(c in input_str for c in '<>:|?*')
        return True
    
    
    
    
    
    
    
    
    
    
    
    

class ExcelManager:
    """Clase dedicada al manejo de archivos Excel de seguimiento"""
    
    def __init__(self, file_path=None):
        """Inicializa el administrador de Excel"""
        self.file_path = file_path
        
    def select_file(self):
        """Permite al usuario elegir un archivo Excel existente o crear uno nuevo"""
        # Preguntar si desea usar un archivo existente o crear uno nuevo
        choice = messagebox.askquestion(
            "Archivo Excel",
            "¿Desea usar un archivo Excel existente?\n\n"
            + "Seleccione 'Sí' para elegir un archivo existente.\n"
            + "Seleccione 'No' para crear un nuevo archivo.",
            icon='question'
        )

        if choice == "yes":
            # Permitir al usuario seleccionar un archivo Excel existente
            file_path = filedialog.askopenfilename(
                title="Seleccione el archivo Excel de seguimiento",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            )

            if not file_path:  # El usuario canceló la selección
                logger.info("Usuario canceló la selección de archivo. Se creará uno nuevo.")
                self._create_new_file()
        else:
            self._create_new_file()
            
        return self.file_path
    
    def _create_new_file(self):
        """Crea un nuevo archivo Excel para seguimiento"""
        # Crear un nombre de archivo por defecto con fecha y hora
        default_filename = f"Seguimiento_Recomendaciones_EPM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Permitir al usuario guardar con un nombre específico
        file_path = filedialog.asksaveasfilename(
            title="Guardar nuevo archivo Excel",
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Archivos Excel", "*.xlsx")],
        )

        if not file_path:  # Si cancela, usar el nombre por defecto
            file_path = default_filename
            logger.info(f"Se usará el nombre por defecto: {file_path}")

        # Crear un archivo Excel vacío con las columnas necesarias
        self._create_excel_template(file_path)
        logger.info(f"Creado nuevo archivo Excel: {file_path}")
        self.file_path = file_path
        
    def _create_excel_template(self, file_path):
        """Crea la estructura del archivo Excel con las columnas necesarias"""
        try:
            # Crear un DataFrame vacío con las columnas necesarias
            columns = [
                "Title",
                "Type",
                "Priority",
                "Status",
                "Deadline",
                "Due Date",
                "Created By",
                "Created On",
                "Last Updated",
                "Comments",
            ]

            df = pd.DataFrame(columns=columns)

            # Guardar el DataFrame vacío como un archivo Excel
            df.to_excel(file_path, index=False)
            logger.info(f"Archivo Excel creado exitosamente: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error al crear nuevo archivo Excel: {e}")
            return False
            
    def update_with_issues(self, issues_data):
        """Actualiza el archivo Excel con los datos extraídos"""
        if not self.file_path:
            logger.error("No hay ruta de archivo Excel especificada")
            return False
            
        if not issues_data:
            logger.warning("No hay datos para actualizar en Excel")
            return False
            
        try:
            logger.info(f"Actualizando archivo Excel: {self.file_path}")
            
            # Cargar el archivo existente o crear estructura si no existe
            if os.path.exists(self.file_path):
                try:
                    existing_df = pd.read_excel(self.file_path, engine='openpyxl')
                    logger.info(f"Archivo Excel existente cargado con {len(existing_df)} registros")
                except Exception as read_e:
                    logger.warning(f"Error al leer Excel: {read_e}. Creando estructura nueva.")
                    existing_df = pd.DataFrame(
                        columns=[
                            "Title", "Type", "Priority", "Status", "Deadline", "Due Date",
                            "Created By", "Created On", "Last Updated", "Comments"
                        ]
                    )
            else:
                existing_df = pd.DataFrame(
                    columns=[
                        "Title", "Type", "Priority", "Status", "Deadline", "Due Date",
                        "Created By", "Created On", "Last Updated", "Comments"
                    ]
                )
                logger.info("Creando nueva estructura de Excel")

            # Convertir datos de issues a DataFrame
            new_df = pd.DataFrame(issues_data)
            
            # Contadores para estadísticas
            new_items = 0
            updated_items = 0
            
            # Hacer una copia del DataFrame existente para no modificarlo mientras iteramos
            updated_df = existing_df.copy()
            
            # Optimización: crear índice para búsquedas rápidas si hay muchos registros
            title_index = {}
            if len(existing_df) > 0 and "Title" in existing_df.columns:
                for idx, title in enumerate(existing_df["Title"]):
                    if title not in title_index:
                        title_index[title] = idx
            
            # Procesar cada issue nuevo
            for _, new_row in new_df.iterrows():
                title = new_row["Title"]
                title_exists = title in title_index
                
                if not title_exists:
                    # Agregar fecha de última actualización para elementos nuevos
                    new_row_dict = new_row.to_dict()
                    new_row_dict["Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    new_row_df = pd.DataFrame([new_row_dict])
                    updated_df = pd.concat([updated_df, new_row_df], ignore_index=True)
                    new_items += 1
                    logger.info(f"Nuevo issue añadido: '{title}'")
                else:
                    # Obtener índice del elemento existente
                    idx = title_index[title]
                    
                    # Verificar cambios en el estado y otras columnas
                    updated = False
                    
                    for column in ["Status", "Priority", "Type", "Due Date", "Deadline", "Created By", "Created On"]:
                        if column in new_row and column in existing_df.columns:
                            old_value = existing_df.iloc[idx][column] if not pd.isna(existing_df.iloc[idx][column]) else ""
                            new_value = new_row[column] if not pd.isna(new_row[column]) else ""
                            
                            if str(old_value) != str(new_value):
                                mask = updated_df["Title"] == title
                                updated_df.loc[mask, column] = new_value
                                updated_df.loc[mask, "Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                updated = True
                                logger.info(f"Actualizado {column} de '{title}': '{old_value}' → '{new_value}'")
                    
                    if updated:
                        updated_items += 1
            
            # Guardar el DataFrame actualizado
            updated_df.to_excel(self.file_path, index=False, engine='openpyxl')
            
            # Aplicar formato al Excel
            self._apply_excel_formatting()
            
            logger.info(f"Excel actualizado: {new_items} nuevos, {updated_items} actualizados")
            return True, new_items, updated_items
            
        except Exception as e:
            logger.error(f"Error al actualizar el archivo Excel: {e}")
            return False, 0, 0
            
    def _apply_excel_formatting(self):
        """Aplica formato estético al archivo Excel"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Formato para encabezados
            header_fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Aplicar formato a encabezados
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Aplicar formato a celdas de datos
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

                    # Colorear por estado
                    if col == 4:  # Columna Status
                        status = str(cell.value).upper() if cell.value else ""

                        if "DONE" in status:
                            cell.fill = PatternFill(
                                start_color="CCFFCC",
                                end_color="CCFFCC",
                                fill_type="solid",
                            )
                        elif "OPEN" in status:
                            cell.fill = PatternFill(
                                start_color="FFCCCC",
                                end_color="FFCCCC",
                                fill_type="solid",
                            )
                        elif "READY" in status:
                            cell.fill = PatternFill(
                                start_color="FFFFCC",
                                end_color="FFFFCC",
                                fill_type="solid",
                            )
                        elif "IN PROGRESS" in status:
                            cell.fill = PatternFill(
                                start_color="FFE6CC",
                                end_color="FFE6CC",
                                fill_type="solid",
                            )

            # Ajustar ancho de columnas
            for col in range(1, ws.max_column + 1):
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                adjusted_width = max(10, min(50, max_length + 2))
                ws.column_dimensions[
                    ws.cell(row=1, column=col).column_letter
                ].width = adjusted_width

            wb.save(self.file_path)
            logger.info("Formato aplicado al archivo Excel correctamente")
            return True
        except Exception as format_e:
            logger.warning(f"No se pudo aplicar formato al Excel: {format_e}")
            return False
        
        
        
        
        
        
        
        
        
        
        
        
        
        
class SAPBrowser:
    """Clase para la automatización del navegador y extracción de datos de SAP"""
    
    def __init__(self):
        """Inicializa el controlador del navegador"""
        self.driver = None
        self.wait = None
        self.element_cache = {}  # Caché para elementos encontrados frecuentemente
        
    def connect(self):
        """Inicia una sesión de navegador con perfil dedicado"""
        logger.info("Iniciando navegador con perfil guardado...")
        
        try:
            # Ruta al directorio del perfil
            user_data_dir = os.path.join(os.environ['USERPROFILE'], 'AppData', 'Local', 'Google', 'Chrome', 'SAP_Automation')
            
            # Crear directorio si no existe
            if not os.path.exists(user_data_dir):
                os.makedirs(user_data_dir)
                logger.info(f"Creado directorio de perfil: {user_data_dir}")
            
            # Configurar opciones de Chrome
            chrome_options = Options()
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
            chrome_options.add_argument("--profile-directory=Default")
            
            # Opciones para mejorar rendimiento y estabilidad
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-infobars")
            chrome_options.add_argument("--disable-notifications")
            chrome_options.add_argument("--disable-popup-blocking")
            
            # Optimización de memoria
            chrome_options.add_argument("--js-flags=--expose-gc")
            chrome_options.add_argument("--enable-precise-memory-info")
            
            # Agregar opciones para permitir que el usuario use el navegador mientras se ejecuta el script
            chrome_options.add_experimental_option("detach", True)
            
            # Intentar iniciar el navegador
            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 60)  # Timeout de 60 segundos
            
            logger.info("Navegador Chrome iniciado correctamente")
            return True
            
        except WebDriverException as e:
            if "Chrome failed to start" in str(e):
                logger.error(f"Error al iniciar Chrome: {e}. Verificando si hay instancias abiertas...")
                # Intentar recuperar sesión existente
                try:
                    chrome_options = Options()
                    chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
                    self.driver = webdriver.Chrome(options=chrome_options)
                    self.wait = WebDriverWait(self.driver, 60)
                    logger.info("Conexión exitosa a sesión existente de Chrome")
                    return True
                except Exception as debug_e:
                    logger.error(f"No se pudo conectar a sesión existente: {debug_e}")
            
            logger.error(f"Error al iniciar Navegador: {e}")
            return False
    




    def navigate_to_sap(self, erp_number=None, project_id=None):
        """
        Navega a la URL de SAP con parámetros específicos de cliente y proyecto,
        implementando verificación y reintentos para evitar redirecciones.
        
        Args:
            erp_number (str, optional): Número ERP del cliente.
            project_id (str, optional): ID del proyecto.
        
        Returns:
            bool: True si la navegación fue exitosa, False en caso contrario
        """
        if not self.driver:
            logger.error("No hay navegador iniciado")
            return False
            
        try:
            # URL destino exacta
            target_url = f"https://xalm-prod.x.eu20.alm.cloud.sap/launchpad#iam-ui&/?erpNumber={erp_number}&crmProjectId={project_id}&x-app-name=HEP"
            logger.info(f"Intentando navegar a: {target_url}")
            
            # Intentar navegación directa
            self.driver.get(target_url)
            time.sleep(5)  # Esperar carga inicial
            
            # Verificar si fuimos redirigidos
            current_url = self.driver.current_url
            logger.info(f"URL actual después de navegación: {current_url}")
            
            # Si fuimos redirigidos a otra página, intentar navegar directamente por JavaScript
            if "sdwork-center" in current_url or not "iam-ui" in current_url:
                logger.warning("Detectada redirección no deseada, intentando navegación por JavaScript")
                
                # Intentar con JavaScript para evitar redirecciones
                js_navigate_script = f"""
                window.location.href = "{target_url}";
                """
                self.driver.execute_script(js_navigate_script)
                time.sleep(5)  # Esperar a que cargue la página
                
                # Verificar nuevamente
                current_url = self.driver.current_url
                logger.info(f"URL después de navegación por JavaScript: {current_url}")
                
                # Si aún no estamos en la URL correcta, usar hackerParams para forzar
                if "sdwork-center" in current_url or not "iam-ui" in current_url:
                    logger.warning("Redirección persistente, intentando método forzado")
                    
                    # Método más agresivo para forzar la navegación
                    force_script = f"""
                    var hackerParams = new URLSearchParams();
                    hackerParams.append('erpNumber', '{erp_number}');
                    hackerParams.append('crmProjectId', '{project_id}');
                    hackerParams.append('x-app-name', 'HEP');
                    
                    var targetHash = '#iam-ui&/?' + hackerParams.toString();
                    window.location.hash = targetHash;
                    """
                    self.driver.execute_script(force_script)
                    time.sleep(5)
            
            # Intentar aceptar certificados o diálogos si aparecen
            try:
                ok_buttons = self.driver.find_elements(By.XPATH, 
                    "//button[contains(text(), 'OK') or contains(text(), 'Ok') or contains(text(), 'Aceptar')]")
                if ok_buttons:
                    for button in ok_buttons:
                        if button.is_displayed():
                            button.click()
                            logger.info("Se hizo clic en un botón de diálogo")
                            time.sleep(1)
            except Exception as dialog_e:
                logger.debug(f"Error al manejar diálogos: {dialog_e}")
            
            # Verificar URL final después de todos los intentos
            final_url = self.driver.current_url
            logger.info(f"URL final después de todos los intentos: {final_url}")
            
            # Esperar a que la página cargue completamente
            try:
                WebDriverWait(self.driver, 15).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
                logger.info("Página cargada completamente")
            except TimeoutException:
                logger.warning("Tiempo de espera excedido para carga completa de página")
            
            # Considerar éxito si contiene iam-ui o los parámetros específicos
            success = "iam-ui" in final_url or erp_number in final_url
            if success:
                logger.info("Navegación exitosa a la página deseada")
            else:
                logger.warning("No se pudo navegar a la página exacta deseada")
                
            return True  # Continuar con el flujo incluso si no llegamos exactamente a la URL
                
        except Exception as e:
            logger.error(f"Error al navegar a SAP: {e}")
            return False
    
    
    




    def handle_authentication(self):
        """
        Maneja el proceso de autenticación en SAP, detectando si es necesario.
        """
        try:
            logger.info("Verificando si se requiere autenticación...")
            
            # Verificar si hay formulario de login visible
            login_elements = self.driver.find_elements(By.XPATH, "//input[@type='email'] | //input[@type='password']")
            
            if login_elements:
                logger.info("Formulario de login detectado, esperando introducción manual de credenciales")
                
                # Mostrar mensaje al usuario si estamos en interfaz gráfica
                if hasattr(self, 'root') and self.root:
                    messagebox.showinfo(
                        "Autenticación Requerida",
                        "Por favor, introduzca sus credenciales en el navegador.\n\n"
                        "Haga clic en OK cuando haya iniciado sesión."
                    )
                else:
                    print("\n=== AUTENTICACIÓN REQUERIDA ===")
                    print("Por favor, introduzca sus credenciales en el navegador.")
                    input("Presione ENTER cuando haya iniciado sesión...\n")
                
                # Esperar a que desaparezca la pantalla de login
                try:
                    WebDriverWait(self.driver, 60).until_not(
                        EC.presence_of_element_located((By.XPATH, "//input[@type='password']"))
                    )
                    logger.info("Autenticación completada exitosamente")
                    return True
                except TimeoutException:
                    logger.warning("Tiempo de espera excedido para autenticación")
                    return False
            else:
                logger.info("No se requiere autenticación, ya hay una sesión activa")
                return True
                
        except Exception as e:
            logger.error(f"Error durante el proceso de autenticación: {e}")
            return False












    def select_customer_automatically(self, erp_number):
        """
        Selecciona automáticamente un cliente en la pantalla de Project Overview.
        
        Args:
            erp_number (str): Número ERP del cliente a seleccionar
            
        Returns:
            bool: True si la selección fue exitosa, False en caso contrario
        """
        try:
              # Verificar que el ERP no esté vacío
            if not erp_number or erp_number.strip() == "":
                logger.warning("No se puede seleccionar cliente: ERP número vacío")
                return False
            logger.info(f"Intentando seleccionar automáticamente el cliente {erp_number}...")
            
            # Esperar a que la página cargue completamente
            time.sleep(3)
            
            # 1. Localizar el campo de entrada de cliente
            customer_field_selectors = [
                "//input[@placeholder='Enter Customer ID or Name']",
                "//input[contains(@placeholder, 'Customer')]",
                "//input[@id='customer']",
                "//input[contains(@aria-label, 'Customer')]",
                "//div[contains(text(), 'Customer')]/following-sibling::div//input",
                "//label[contains(text(), 'Customer')]/following-sibling::div//input"
            ]
            
            customer_field = None
            for selector in customer_field_selectors:
                try:
                    elements = self.driver.find_elements(By.XPATH, selector)
                    for element in elements:
                        if element.is_displayed():
                            customer_field = element
                            break
                    if customer_field:
                        break
                except:
                    continue
            
            if not customer_field:
                logger.warning("No se pudo encontrar el campo de cliente visible")
                return False
            
            # 2. Hacer clic en el campo para asegurarnos que tiene el foco
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", customer_field)
                time.sleep(0.5)
                self.driver.execute_script("arguments[0].click();", customer_field)
                logger.info("Clic realizado en campo de cliente")
                time.sleep(0.5)
            except:
                logger.warning("No se pudo hacer clic en el campo de cliente")
            
            # 3. Limpiar el campo y escribir el número ERP
            try:
                # Limpiar usando JavaScript y luego las teclas
                self.driver.execute_script("arguments[0].value = '';", customer_field)
                customer_field.clear()
                
                # Escribir el ERP character por character con pequeñas pausas
                for char in erp_number:
                    customer_field.send_keys(char)
                    time.sleep(0.1)
                
                logger.info(f"Número ERP {erp_number} ingresado en campo de cliente")
                time.sleep(1)
            except Exception as e:
                logger.error(f"Error al ingresar número ERP: {e}")
                return False
            
            # 4. Esperar a que aparezcan las sugerencias
            time.sleep(2)
            
            # 5. Seleccionar el cliente de la lista de sugerencias
            suggestion_selectors = [
                f"//div[contains(@class, 'sapMPopover')]//div[contains(text(), '{erp_number}')]",
                f"//div[contains(@class, 'sapMListItems')]//div[contains(text(), '{erp_number}')]",
                f"//ul[contains(@class, 'sapMList')]//li[contains(text(), '{erp_number}')]",
                f"//*[contains(text(), '{erp_number}') and contains(text(), 'Empresas')]"
            ]
            
            suggestion_found = False
            for selector in suggestion_selectors:
                try:
                    suggestions = self.driver.find_elements(By.XPATH, selector)
                    for suggestion in suggestions:
                        if suggestion.is_displayed():
                            # Hacer scroll hasta la sugerencia
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", suggestion)
                            time.sleep(0.5)
                            
                            # Hacer clic en la sugerencia
                            self.driver.execute_script("arguments[0].click();", suggestion)
                            logger.info(f"Clic realizado en sugerencia de cliente: {suggestion.text}")
                            time.sleep(1)
                            suggestion_found = True
                            break
                    if suggestion_found:
                        break
                except:
                    continue
            
            # Si no encontramos sugerencias, intentar presionar Enter
            if not suggestion_found:
                logger.info("No se encontraron sugerencias, presionando Enter")
                customer_field.send_keys(Keys.ENTER)
                time.sleep(2)
                
                # Intentar buscar la primera línea de resultados
                try:
                    first_result = self.driver.find_element(By.XPATH, 
                        "//div[contains(@class, 'sapMListItems')]/div[1] | //tbody/tr[1] | //li[1]")
                    self.driver.execute_script("arguments[0].click();", first_result)
                    logger.info("Clic en primer resultado realizado")
                    time.sleep(1)
                    suggestion_found = True
                except:
                    logger.warning("No se pudo seleccionar el primer resultado")
            
            # 6. Verificar si se seleccionó correctamente
            try:
                # Verificar si hay texto visible con el cliente seleccionado
                visible_elements = self.driver.find_elements(By.XPATH, 
                    f"//*[contains(text(), '{erp_number}') and (contains(text(), 'Empresas') or contains(text(), 'Publicas'))]")
                
                if visible_elements:
                    for element in visible_elements:
                        if element.is_displayed():
                            logger.info(f"Cliente {erp_number} seleccionado con éxito")
                            return True
                
                # También verificar si el campo ahora tiene el valor correcto
                if customer_field.get_attribute("value") and erp_number in customer_field.get_attribute("value"):
                    logger.info(f"Campo de cliente ahora contiene '{erp_number}'")
                    return True
                    
                logger.warning(f"No se pudo confirmar que el cliente {erp_number} fue seleccionado")
                return suggestion_found  # Retornar True si al menos se hizo clic en una sugerencia
            except Exception as verify_e:
                logger.error(f"Error al verificar selección de cliente: {verify_e}")
                return suggestion_found
                
        except Exception as e:
            logger.error(f"Error durante la selección automática de cliente: {e}")
            return False    
    
    
    
    
    def verify_fields_have_expected_values(self, erp_number, project_id):
        """Verifica si los campos ya contienen los valores esperados"""
        try:
            # Verificar el campo de cliente
            customer_fields = self.driver.find_elements(
                By.XPATH,
                "//input[contains(@placeholder, 'Customer') or contains(@placeholder, 'cliente')]"
            )
            
            if customer_fields:
                for field in customer_fields:
                    current_value = field.get_attribute("value") or ""
                    if erp_number in current_value:
                        logger.info(f"Campo de cliente ya contiene '{erp_number}'")
                        return True
            
            # Verificar si hay texto visible con estos valores
            page_text = self.driver.find_element(By.TAG_NAME, "body").text
            if erp_number in page_text and project_id in page_text:
                logger.info(f"La página ya contiene '{erp_number}' y '{project_id}'")
                
                # Verificar si estamos en la página correcta con los datos
                issues_elements = self.driver.find_elements(
                    By.XPATH, 
                    "//div[contains(text(), 'Issues')]"
                )
                if issues_elements:
                    logger.info("En la página correcta con datos cargados")
                    return True
                
            return False
        except Exception as e:
            logger.debug(f"Error al verificar campos: {e}")
            return False
    
    
    
    
    
    
    def _fill_fields_and_extract(self, erp_number, project_id):
        """Rellena los campos y luego ejecuta la extracción"""
        try:
            # Primero verificar si los campos ya tienen los valores correctos
            if self.browser.verify_fields_have_expected_values(erp_number, project_id):
                logger.info("Los campos ya contienen los valores correctos, procediendo con la extracción")
                # Ir directamente a la extracción sin intentar rellenar campos
                self.perform_extraction()
                return
                
            logger.info("Intentando rellenar campos automáticamente")
            
            # Resto del código para rellenar campos...
            # ...
            
        except Exception as e:
            logger.error(f"Error al rellenar campos y extraer: {e}")
            # Manejo de errores...
    
    
    
    
    def select_project_automatically(self, project_id):
        """
        Selecciona automáticamente un proyecto escribiendo directamente su ID.
        """
        try:
            logger.info(f"Seleccionando proyecto {project_id} automáticamente...")
            
            # Buscar campo de proyecto
            project_field_xpath = "//input[contains(@placeholder, 'Project')] | //input[@id='project'] | //input[contains(@aria-label, 'Project')]"
            
            try:
                # Esperar a que el campo sea visible y clickeable
                project_field = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, project_field_xpath))
                )
                
                # Limpiar campo si ya tiene texto
                project_field.clear()
                
                # Escribir el ID del proyecto
                project_field.send_keys(project_id)
                logger.info(f"ID de proyecto {project_id} ingresado en campo")
                time.sleep(1)
                
                # Simular presionar ENTER para buscar
                project_field.send_keys(Keys.ENTER)
                logger.info("Tecla ENTER presionada para buscar proyecto")
                time.sleep(2)
                
                # Hacer clic en el proyecto encontrado
                project_xpath = f"//div[contains(text(), '{project_id}')] | //span[contains(text(), '{project_id}')] | //td[contains(text(), '{project_id}')]"
                
                try:
                    project_element = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, project_xpath))
                    )
                    project_element.click()
                    logger.info(f"Proyecto {project_id} seleccionado exitosamente")
                    time.sleep(2)
                    return True
                except TimeoutException:
                    logger.warning(f"No se encontró el proyecto {project_id} en la lista después de buscar")
                    
                    # Intentar con el primer resultado si existe
                    try:
                        first_result = self.driver.find_element(By.XPATH, 
                            "//div[contains(@class, 'sapMListItems')]/div[1] | //tbody/tr[1]")
                        first_result.click()
                        logger.info("Seleccionado primer resultado de búsqueda de proyecto")
                        time.sleep(2)
                        return True
                    except NoSuchElementException:
                        logger.error("No se encontraron resultados de búsqueda de proyecto")
            except TimeoutException:
                logger.warning("No se encontró campo de búsqueda para proyecto")
                
            # Si llegamos aquí, intentar con el botón de selección como respaldo
            try:
                selector_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[contains(@class, 'project')]//button"))
                )
                selector_button.click()
                logger.info("Clic en botón de selección de proyecto como alternativa")
                time.sleep(2)
                
                # Buscar e ingresar el proyecto en el diálogo
                search_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//input[@type='search']"))
                )
                search_field.clear()
                search_field.send_keys(project_id)
                search_field.send_keys(Keys.ENTER)
                time.sleep(2)
                
                # Seleccionar el proyecto
                project_element = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//div[contains(text(), '{project_id}')]"))
                )
                project_element.click()
                logger.info(f"Proyecto {project_id} seleccionado por método alternativo")
                time.sleep(2)
                return True
            except Exception as backup_e:
                logger.error(f"Falló también el método alternativo: {backup_e}")
                return False
                
        except Exception as e:
            logger.error(f"Error durante la selección automática de proyecto: {e}")
            return False
    










    def click_search_button(self):
        """
        Hace clic en el botón de búsqueda para iniciar la consulta.
        """
        try:
            logger.info("Buscando botón de búsqueda...")
            
            # Diferentes selectores para el botón de búsqueda
            search_button_selectors = [
                "//button[contains(@aria-label, 'Search')]",
                "//button[@title='Search']",
                "//span[contains(text(), 'Search')]/parent::button",
                "//button[contains(@class, 'sapMBarChild') and contains(@class, 'sapMBtn')]",
                "//div[contains(@class, 'sapMBarPH')]//button"
            ]
            
            for selector in search_button_selectors:
                try:
                    search_button = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, selector))
                    )
                    search_button.click()
                    logger.info("Clic en botón de búsqueda exitoso")
                    time.sleep(2)
                    return True
                except:
                    continue
                    
            # Si no se encuentra por selectores específicos, buscar por icono típico
            icon_selectors = [
                "//span[contains(@class, 'sapUiIcon') and contains(@class, 'sapMBtnIcon')]",
                "//span[contains(@data-sap-ui, 'search')]"
            ]
            
            for selector in icon_selectors:
                try:
                    search_icons = self.driver.find_elements(By.XPATH, selector)
                    for icon in search_icons:
                        # Verificar si el icono es visible
                        if icon.is_displayed():
                            self.driver.execute_script("arguments[0].click();", icon)
                            logger.info("Clic en icono de búsqueda exitoso")
                            time.sleep(2)
                            return True
                except:
                    continue
                    
            logger.warning("No se pudo hacer clic en botón de búsqueda")
            return False
        except Exception as e:
            logger.error(f"Error al hacer clic en botón de búsqueda: {e}")
            return False








        
    def get_total_issues_count(self):
        """Obtiene el número total de issues desde el encabezado"""
        try:
            # Estrategia 1: Buscar el texto "Issues (número)"
            try:
                header_text = self.driver.find_element(
                    By.XPATH, "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                ).text
                logger.info(f"Texto del encabezado de issues: {header_text}")
                
                # Extraer el número entre paréntesis
                match = re.search(r'\((\d+)\)', header_text)
                if match:
                    return int(match.group(1))
            except NoSuchElementException:
                logger.warning("No se encontró el encabezado Issues con formato (número)")
            
            # Estrategia 2: Buscar contador específico de SAP UI5
            try:
                counter_element = self.driver.find_element(
                    By.XPATH, "//div[contains(@class, 'sapMITBCount')]"
                )
                if counter_element.text.isdigit():
                    return int(counter_element.text)
            except NoSuchElementException:
                logger.warning("No se encontró contador de issues en formato SAP UI5")
            
            # Estrategia 3: Contar filas visibles y usar como estimación
            rows = self.find_table_rows(highlight=False)
            if rows:
                count = len(rows)
                logger.info(f"Estimando número de issues basado en filas visibles: {count}")
                return max(count, 100)  # Al menos 100 para asegurar cobertura completa
            
            logger.warning("No se pudo determinar el número total de issues, usando valor por defecto")
            return 100  # Valor por defecto
            
        except Exception as e:
            logger.error(f"Error al obtener el total de issues: {e}")
            return 100  # Valor por defecto si hay error
    
    def check_for_pagination(self):
        """Verifica si la tabla tiene paginación y devuelve los controles"""
        try:
            # Selectores para controles de paginación
            pagination_selectors = [
                "//div[contains(@class, 'sapMPaginator')]",
                "//div[contains(@class, 'sapUiTablePaginator')]",
                "//div[contains(@class, 'pagination')]",
                "//button[contains(@class, 'navButton') or contains(@aria-label, 'Next') or contains(@aria-label, 'Siguiente')]",
                "//span[contains(@class, 'sapMPaginatorButton')]",
                "//button[contains(text(), 'Next') or contains(text(), 'Siguiente')]",
                "//a[contains(@class, 'sapMBtn') and contains(@aria-label, 'Next')]"
            ]
            
            for selector in pagination_selectors:
                elements = self.driver.find_elements(By.XPATH, selector)
                if elements:
                    logger.info(f"Se encontraron controles de paginación: {len(elements)} elementos con selector {selector}")
                    return elements
            
            # Buscar elementos "Show More" o "Load More"
            load_more_selectors = [
                "//button[contains(text(), 'More') or contains(text(), 'más') or contains(text(), 'Show')]",
                "//a[contains(text(), 'More') or contains(text(), 'Load')]",
                "//div[contains(@class, 'sapMListShowMoreButton')]",
                "//span[contains(text(), 'Show') and contains(text(), 'More')]/..",
                "//span[contains(@class, 'sapUiTableColShowMoreButton')]"
            ]
            
            for selector in load_more_selectors:
                elements = self.driver.find_elements(By.XPATH, selector)
                if elements:
                    logger.info(f"Se encontraron botones 'Show More': {len(elements)} elementos con selector {selector}")
                    return elements
            
            logger.info("No se encontraron controles de paginación en la tabla")
            return None
            
        except Exception as e:
            logger.error(f"Error al verificar paginación: {e}")
            return None
    









    def click_pagination_next(self, pagination_elements):
        """Hace clic en el botón de siguiente página"""
        if not pagination_elements:
            return False
            
        try:
            # Buscar el botón "Next" entre los elementos de paginación
            next_button = None
            
            for element in pagination_elements:
                try:
                    aria_label = element.get_attribute("aria-label") or ""
                    text = element.text.lower()
                    classes = element.get_attribute("class") or ""
                    
                    # Comprobar si es un botón "Next" o "Siguiente"
                    if ("next" in aria_label.lower() or 
                        "siguiente" in aria_label.lower() or
                        "next" in text or 
                        "siguiente" in text or
                        "show more" in text.lower() or
                        "more" in text.lower()):
                        
                        next_button = element
                        break
                        
                    # Comprobar por clase CSS
                    if ("next" in classes.lower() or 
                        "pagination-next" in classes.lower() or
                        "sapMBtn" in classes and "NavButton" in classes):
                        
                        next_button = element
                        break
                except Exception:
                    continue
            
            # Si se encontró un botón Next, intentar hacer clic
            if next_button:
                # Verificar si el botón está habilitado
                disabled = next_button.get_attribute("disabled") == "true" or next_button.get_attribute("aria-disabled") == "true"
                
                if disabled:
                    logger.info("Botón de siguiente página está deshabilitado")
                    return False
                    
                # Scroll hacia el botón para asegurar que está visible
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
                time.sleep(0.5)
                
                # Intentar clic con distintos métodos, en orden de preferencia
                try:
                    self.driver.execute_script("arguments[0].click();", next_button)
                    logger.info("Clic en botón 'Next' realizado con JavaScript")
                    time.sleep(2)
                    return True
                except JavascriptException:
                    try:
                        next_button.click()
                        logger.info("Clic en botón 'Next' realizado")
                        time.sleep(2)
                        return True
                    except ElementClickInterceptedException:
                        from selenium.webdriver.common.action_chains import ActionChains
                        actions = ActionChains(self.driver)
                        actions.move_to_element(next_button).click().perform()
                        logger.info("Clic en botón 'Next' realizado con ActionChains")
                        time.sleep(2)
                        return True
            
            # Si no se encontró botón específico, intentar con el último elemento
            if pagination_elements and len(pagination_elements) > 0:
                last_element = pagination_elements[-1]
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", last_element)
                time.sleep(0.5)
                self.driver.execute_script("arguments[0].click();", last_element)
                logger.info("Clic en último elemento de paginación realizado")
                time.sleep(2)
                return True
            
            logger.warning("No se pudo identificar o hacer clic en el botón 'Next'")
            return False
            
        except Exception as e:
            logger.error(f"Error al hacer clic en paginación: {e}")
            return False
    




    def scroll_to_load_all_items(self, total_expected=100, max_attempts=100):
        """Estrategia optimizada para cargar todos los elementos mediante scroll con mejor rendimiento"""
        logger.info(f"Iniciando carga de {total_expected} elementos...")
        
        previous_rows_count = 0
        no_change_count = 0
        no_change_threshold = 10
        
        # Verificar tipo de tabla y estrategia de carga
        table_type = self._detect_table_type()
        logger.info(f"Tipo de tabla detectado: {table_type}")
        
        # Verificar si hay paginación
        pagination_elements = self.check_for_pagination()
        has_pagination = pagination_elements is not None and len(pagination_elements) > 0
        
        logger.info(f"¿La tabla tiene paginación? {'Sí' if has_pagination else 'No'}")
        
        # Ejecutar script para optimizar el rendimiento del navegador
        self._optimize_browser_performance()
        
        # Algoritmo principal de scroll
        for attempt in range(max_attempts):
            try:
                # Usar estrategia de scroll adaptada al tipo de tabla detectado
                if table_type == "standard_ui5":
                    self._scroll_standard_ui5_table()
                elif table_type == "responsive_table":
                    self._scroll_responsive_table()
                elif table_type == "grid_table":
                    self._scroll_grid_table()
                else:
                    # Estrategia genérica
                    self._scroll_generic()
                
                # Contar filas actualmente visibles
                rows = self.find_table_rows(highlight=False)
                current_rows_count = len(rows)
                
                # Registrar progreso periódicamente
                if attempt % 10 == 0:
                    logger.info(f"Intento {attempt+1}: {current_rows_count} filas cargadas")
                
                # Verificación de carga completa con lógica mejorada
                if current_rows_count == previous_rows_count:
                    no_change_count += 1
                    
                    # Si hay paginación y no hay cambios, intentar pasar a página siguiente
                    if has_pagination and no_change_count >= 5:
                        logger.info("Intentando pasar a la siguiente página...")
                        pagination_elements = self.check_for_pagination()
                        if pagination_elements and self.click_pagination_next(pagination_elements):
                            logger.info("Se pasó a la siguiente página")
                            no_change_count = 0
                            time.sleep(3)
                            continue
                    
                    # Si no hay cambios, aplicar estrategias adicionales de scroll
                    if no_change_count >= 5:
                        if no_change_count % 5 == 0:  # Alternar estrategias
                            self._apply_alternative_scroll_strategy(no_change_count)
                    
                    # Criterios de finalización adaptados
                    if self._should_finish_scrolling(no_change_count, current_rows_count, total_expected):
                        break
                else:
                    # Reiniciar contador si se encontraron más filas
                    no_change_count = 0
                    
                previous_rows_count = current_rows_count
                
                # Si se alcanzó o superó el número esperado, terminar
                if current_rows_count >= total_expected:
                    logger.info(f"Se han cargado {current_rows_count} filas (>= {total_expected} esperadas)")
                    break
                
                # Tiempo adaptativo de espera basado en el rendimiento
                wait_time = self._calculate_adaptive_wait_time(no_change_count, current_rows_count)
                time.sleep(wait_time)
                    
            except Exception as e:
                logger.warning(f"Error durante el scroll en intento {attempt+1}: {e}")
            
        # Calcular y reportar métricas de éxito
        coverage = (previous_rows_count / total_expected) * 100 if total_expected > 0 else 0
        logger.info(f"Scroll completado. Cobertura: {coverage:.2f}% ({previous_rows_count}/{total_expected})")
        
        return previous_rows_count
    
    
    
    def find_table_rows(self, highlight=False):
        """Encuentra todas las filas de la tabla con múltiples estrategias"""
        all_rows = []
        
        # Usar caché si está disponible
        cache_key = "table_rows"
        if cache_key in self.element_cache:
            cache_time, cached_rows = self.element_cache[cache_key]
            if (datetime.now() - cache_time).total_seconds() < 5:  # Caché válida por 5 segundos
                logger.debug("Usando filas en caché")
                return cached_rows

        # Selectores mejorados para SAP UI5
        selectors = [
            # Selectores de SAP estándar
            "//table[contains(@class, 'sapMListTbl')]/tbody/tr[not(contains(@class, 'sapMListTblHeader'))]",
            "//div[contains(@class, 'sapMList')]//li[contains(@class, 'sapMLIB')]",
            "//table[contains(@class, 'sapMList')]/tbody/tr",
            "//div[@role='row'][not(contains(@class, 'sapMListHeaderSubTitleItems')) and not(contains(@class, 'sapMListTblHeader'))]",
            "//div[contains(@class, 'sapMListItems')]/div[contains(@class, 'sapMListItem')]",
            "//div[contains(@class, 'sapMListItems')]//div[contains(@class, 'sapMObjectIdentifier')]/..",
            "//div[contains(@class, 'sapMListItem')]",
            
            # Selectores de Fiori
            "//div[contains(@class, 'sapMList')]//li[@tabindex]",
            "//div[contains(@class, 'sapUiTable')]//tr[contains(@class, 'sapUiTableRow')]",
            "//div[contains(@class, 'sapUiTableRowHdr')]/..",
            "//table[contains(@class, 'sapMTable')]//tr[not(contains(@class, 'sapMListTblHeaderRow'))]",
            
            # Selectores específicos de SDWork Center
            "//div[contains(@class, 'sdworkItems')]//div[contains(@class, 'sapMLIB')]",
            "//div[contains(@class, 'issueList')]//div[contains(@class, 'sapMLIB')]",
            "//div[contains(@id, 'issue')]//div[contains(@class, 'sapMLIB')]",
            
            # Selectores genéricos más específicos
            "//div[contains(@class, 'sapMLIB-CTX')]",
            "//div[contains(@class, 'sapMObjectListItem')]",
            "//div[contains(@class, 'sapMListModeMultiSelect')]//div[contains(@class, 'sapMLIB')]"
        ]

        for selector in selectors:
            try:
                rows = self.driver.find_elements(By.XPATH, selector)
                if len(rows) > 0:
                    logger.info(f"Se encontraron {len(rows)} filas con selector: {selector}")
                    
                    # Filtrar filas válidas
                    valid_rows = []
                    for row in rows:
                        try:
                            has_content = False
                            text_elements = row.find_elements(By.XPATH, ".//span | .//div | .//a")
                            for element in text_elements:
                                if element.text and element.text.strip():
                                    has_content = True
                                    break
                                    
                            if has_content:
                                # Verificar que no sea un encabezado
                                class_attr = row.get_attribute("class") or ""
                                is_header = "header" in class_attr.lower()
                                if not is_header:
                                    valid_rows.append(row)
                        except:
                            valid_rows.append(row)  # Si hay error, incluir por si acaso
                    
                    if len(valid_rows) > 0:
                        all_rows = valid_rows
                        
                        if len(valid_rows) >= 75:  # Si encontramos muchas filas, probablemente es el selector correcto
                            break
            except Exception as e:
                logger.debug(f"Error con selector {selector}: {e}")
        
        # Si los selectores estándar fallan, usar enfoque alternativo
        if len(all_rows) == 0:
            logger.warning("Usando enfoque alternativo para encontrar filas")
            try:
                any_rows = self.driver.find_elements(
                    By.XPATH,
                    "//div[contains(@class, 'sapM')] | //tr | //li[contains(@class, 'sapM')]"
                )
                
                for element in any_rows:
                    try:
                        # Verificar si parece una fila de datos
                        if element.text and len(element.text.strip()) > 10:
                            children = element.find_elements(By.XPATH, ".//*")
                            if len(children) >= 3:
                                parent_elements = element.find_elements(
                                    By.XPATH, 
                                    "./ancestor::div[contains(@class, 'sapMList') or contains(@class, 'sapMTable')]"
                                )
                                if len(parent_elements) > 0:
                                    all_rows.append(element)
                    except:
                        continue
                        
                logger.info(f"Enfoque alternativo encontró {len(all_rows)} posibles filas")
            except Exception as e:
                logger.error(f"Error en enfoque alternativo: {e}")
        
        # Resaltar filas si se solicita
        if highlight and len(all_rows) > 0:
            try:
                self.driver.execute_script(
                    "arguments[0].scrollIntoView(true); arguments[0].style.border = '2px solid red';",
                    all_rows[0]
                )
            except:
                pass
        
        # Eliminar duplicados manteniendo el orden
        unique_rows = []
        seen_ids = set()
        
        for row in all_rows:
            try:
                row_id = row.id
                if row_id not in seen_ids:
                    seen_ids.add(row_id)
                    unique_rows.append(row)
            except:
                # Si no podemos obtener un ID, usar aproximación con texto y clase
                try:
                    row_text = row.text[:50] if row.text else ""
                    row_class = row.get_attribute("class") or ""
                    row_signature = f"{row_text}|{row_class}"
                    
                    if row_signature not in seen_ids:
                        seen_ids.add(row_signature)
                        unique_rows.append(row)
                except:
                    unique_rows.append(row)
        
        logger.info(f"Total de filas únicas encontradas: {len(unique_rows)}")
        
        # Actualizar caché
        self.element_cache[cache_key] = (datetime.now(), unique_rows)
        
        return unique_rows
    
    
    
    
    
    
    
    
    def _detect_table_headers(self):
        """Detecta y mapea los encabezados de la tabla para mejor extracción"""
        try:
            # Intentar encontrar la fila de encabezados
            header_selectors = [
                "//tr[contains(@class, 'sapMListTblHeader')]",
                "//div[contains(@class, 'sapMListTblHeaderCell')]/..",
                "//div[@role='columnheader']/parent::div[@role='row']",
                "//th[contains(@class, 'sapMListTblHeaderCell')]/.."
            ]
            
            for selector in header_selectors:
                header_rows = self.driver.find_elements(By.XPATH, selector)
                if header_rows:
                    # Tomar la primera fila de encabezados encontrada
                    header_row = header_rows[0]
                    
                    # Extraer las celdas de encabezado
                    header_cells = header_row.find_elements(By.XPATH, 
                        ".//th | .//div[@role='columnheader'] | .//div[contains(@class, 'sapMListTblHeaderCell')]")
                    
                    if header_cells:
                        # Mapear nombres de encabezados a índices
                        header_map = {}
                        for i, cell in enumerate(header_cells):
                            header_text = cell.text.strip()
                            if header_text:
                                header_map[header_text.upper()] = i
                        
                        logger.info(f"Encabezados detectados: {header_map}")
                        return header_map
            
            logger.warning("No se pudieron detectar encabezados de tabla")
            return {}
            
        except Exception as e:
            logger.error(f"Error al detectar encabezados: {e}")
            return {}
    
    
    
    
    
    
    
    def extract_issues_data(self):
        """Extrae datos de issues desde la tabla con procesamiento mejorado"""
        try:
            logger.info("Iniciando extracción de issues...")
            
            # Esperar a que cargue la página inicial
            time.sleep(3)
            
            header_map = self._detect_table_headers()
            
            # Obtener el número total de issues
            total_issues = self.get_total_issues_count()
            logger.info(f"Total de issues a procesar: {total_issues}")
            
            # Hacer scroll para cargar todos los elementos
            loaded_rows_count = self.scroll_to_load_all_items(total_issues)
            
            # Verificar si hay paginación
            pagination_elements = self.check_for_pagination()
            has_pagination = pagination_elements is not None and len(pagination_elements) > 0
            
            # Lista para almacenar todos los datos extraídos
            all_issues_data = []
            seen_titles = set()  # Solo para registrar, ya no para filtrar duplicados
            
            page_num = 1
            max_pages = 20  # Límite de seguridad
            
            while page_num <= max_pages:
                logger.info(f"Procesando página {page_num}...")
                
                # Obtener filas de la página actual
                rows = self.find_table_rows(highlight=False)
                
                if not rows:
                    logger.warning(f"No se encontraron filas en la página {page_num}")
                    break
                
                logger.info(f"Encontradas {len(rows)} filas en la página {page_num}")
                
                # Procesar filas en esta página
                page_issues_data = self._process_table_rows(rows, seen_titles, header_map)
                
                # Validar y corregir los datos extraídos
                corrected_data = []
                for issue in page_issues_data:
                    corrected_issue = self._validate_and_correct_issue_data(issue)
                    corrected_data.append(corrected_issue)
                
                # Agregar los datos corregidos al resultado total
                all_issues_data.extend(corrected_data)
                
                logger.info(f"Extraídos {len(corrected_data)} issues de la página {page_num}")
                logger.info(f"Total de issues extraídos hasta ahora: {len(all_issues_data)}")
                
                # Si no hay paginación o ya procesamos todos los datos, terminar
                if not has_pagination or len(page_issues_data) == 0:
                    break
                
                # Intentar pasar a la siguiente página
                if pagination_elements:
                    if not self.click_pagination_next(pagination_elements):
                        logger.info("No se pudo pasar a la siguiente página, terminando extracción")
                        break
                        
                    # Esperar a que cargue la nueva página
                    time.sleep(3)
                    
                    # Actualizar elementos de paginación (pueden cambiar entre páginas)
                    pagination_elements = self.check_for_pagination()
                    
                    page_num += 1
                else:
                    break
            
            logger.info(f"Extracción completada. Total de issues extraídos: {len(all_issues_data)}")
            
            return all_issues_data
        
        except Exception as e:
            logger.error(f"Error en la extracción de datos: {e}")
            return []    

    
    










    def _detect_table_type(self):
        """
        Detecta el tipo de tabla UI5 presente en la página actual.
        
        Los tipos de tabla incluyen: standard_ui5, responsive_table, grid_table y unknown.
        Cada tipo de tabla requiere una estrategia de scroll diferente.
        
        Returns:
            str: Tipo de tabla detectado ("standard_ui5", "responsive_table", "grid_table" o "unknown")
        """
        try:
            # Buscar patrones característicos de diferentes tipos de tablas
            standard_ui5 = len(self.driver.find_elements(By.XPATH, "//table[contains(@class, 'sapMListTbl')]"))
            responsive_table = len(self.driver.find_elements(By.XPATH, "//div[contains(@class, 'sapMListItems')]"))
            grid_table = len(self.driver.find_elements(By.XPATH, "//div[contains(@class, 'sapUiTable')]"))
            
            # Determinar el tipo de tabla más probable basado en la presencia de elementos
            if standard_ui5 > 0:
                return "standard_ui5"
            elif responsive_table > 0:
                return "responsive_table"
            elif grid_table > 0:
                return "grid_table"
            else:
                return "unknown"
        except Exception as e:
            logger.debug(f"Error al detectar tipo de tabla: {e}")
            return "unknown"  # Si hay error, usar tipo genérico

    def _optimize_browser_performance(self):
        """
        Ejecuta scripts para mejorar el rendimiento del navegador durante el scroll.
        
        Implementa optimizaciones como:
        - Desactivar animaciones CSS
        - Forzar recolección de basura cuando es posible
        - Reducir carga visual
        
        Returns:
            bool: True si la optimización fue exitosa, False en caso contrario
        """
        try:
            # Script para liberar memoria y reducir carga visual
            performance_script = """
            // Desactivar animaciones y transiciones para mejor rendimiento
            try {
                let styleSheet = document.createElement('style');
                styleSheet.textContent = '* { animation-duration: 0.001s !important; transition-duration: 0.001s !important; }';
                document.head.appendChild(styleSheet);
            } catch(e) {}
            
            // Liberar memoria si está disponible el recolector de basura
            if (window.gc) {
                window.gc();
            }
            
            // Optimizar para scroll (desactivar eventos innecesarios)
            try {
                const observer = window.IntersectionObserver;
                if (observer) {
                    // Desconectar observadores de intersección temporalmente
                    const observers = performance.getEntriesByType('resource')
                        .filter(entry => entry.initiatorType === 'observer');
                    
                    for (const obs of observers) {
                        try { obs.disconnect(); } catch(e) {}
                    }
                }
            } catch(e) {}
            """
            
            self.driver.execute_script(performance_script)
            logger.debug("Script de optimización de rendimiento ejecutado")
            return True
        except Exception as e:
            logger.debug(f"Error al optimizar rendimiento del navegador: {e}")
            return False

    def _scroll_standard_ui5_table(self):
        """
        Estrategia de scroll específica para tablas estándar de SAP UI5.
        
        Las tablas estándar de UI5 suelen tener la clase 'sapMListTbl' y
        requieren scroll en sus contenedores específicos.
        
        Returns:
            bool: True si el scroll fue exitoso, False en caso contrario
        """
        try:
            # Identificar contenedores de tablas estándar UI5
            table_containers = self.driver.find_elements(
                By.XPATH, 
                "//div[contains(@class, 'sapMListItems')] | " +
                "//div[contains(@class, 'sapMTableTBody')] | " +
                "//table[contains(@class, 'sapMListTbl')]/parent::div"
            )
            
            # Si se encuentran contenedores específicos, hacer scroll en ellos
            if table_containers:
                for container in table_containers:
                    # Realizar scroll al final del contenedor específico
                    self.driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", container)
                    time.sleep(0.2)  # Breve pausa para permitir carga
            else:
                # Si no se encuentran contenedores específicos, usar scroll general
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                
            return True
        except Exception as e:
            logger.debug(f"Error en scroll de tabla estándar UI5: {e}")
            return False

    def _scroll_responsive_table(self):
        """
        Estrategia de scroll específica para tablas responsivas de SAP UI5.
        
        Las tablas responsivas suelen usar clases como 'sapMList' y 
        requieren técnicas específicas para cargar más elementos.
        
        Returns:
            bool: True si el scroll fue exitoso, False en caso contrario
        """
        try:
            # Script específico para tablas responsivas, con mayor precisión
            scroll_script = """
                // Identificar contenedores de listas y tablas responsivas
                var listContainers = document.querySelectorAll(
                    '.sapMList, .sapMListItems, .sapMListUl, .sapMLIB'
                );
                
                // Si encontramos contenedores específicos, hacer scroll en ellos
                if (listContainers.length > 0) {
                    for (var i = 0; i < listContainers.length; i++) {
                        if (listContainers[i].scrollHeight > listContainers[i].clientHeight) {
                            listContainers[i].scrollTop = listContainers[i].scrollHeight;
                        }
                    }
                    return true;
                } else {
                    // Si no encontramos contenedores específicos, scroll general
                    window.scrollTo(0, document.body.scrollHeight);
                    return false;
                }
            """
            
            result = self.driver.execute_script(scroll_script)
            
            # Si el script no encontró contenedores específicos, intentar con Page Down
            if result is False:
                try:
                    body = self.driver.find_element(By.TAG_NAME, "body")
                    body.send_keys(Keys.PAGE_DOWN)
                except:
                    pass
                    
            return True
        except Exception as e:
            logger.debug(f"Error en scroll de tabla responsiva: {e}")
            return False

    def _scroll_grid_table(self):
        """
        Estrategia de scroll específica para tablas tipo grid de SAP UI5.
        
        Las tablas grid (sapUiTable) tienen una estructura más compleja y
        manejan tanto scroll vertical como horizontal.
        
        Returns:
            bool: True si el scroll fue exitoso, False en caso contrario
        """
        try:
            # Script especializado para tablas grid de UI5
            grid_scroll_script = """
                // Identificar contenedores de scroll en tablas grid
                var gridContainers = document.querySelectorAll(
                    '.sapUiTableCtrlScr, .sapUiTableCtrlCnt, .sapUiTableRowHdr'
                );
                
                var didScroll = false;
                
                // Hacer scroll en cada contenedor relevante
                if (gridContainers.length > 0) {
                    for (var i = 0; i < gridContainers.length; i++) {
                        // Verificar si el contenedor tiene scroll
                        if (gridContainers[i].scrollHeight > gridContainers[i].clientHeight) {
                            // Scroll vertical máximo
                            gridContainers[i].scrollTop = gridContainers[i].scrollHeight;
                            didScroll = true;
                        }
                        
                        // Reset de scroll horizontal para mejor visibilidad
                        if (gridContainers[i].scrollLeft > 0) {
                            gridContainers[i].scrollLeft = 0;
                        }
                    }
                }
                
                // Buscar específicamente botones "More"
                var moreButtons = document.querySelectorAll(
                    'button.sapUiTableMoreBtn, span.sapUiTableColShowMoreBtn'
                );
                
                for (var j = 0; j < moreButtons.length; j++) {
                    if (moreButtons[j] && moreButtons[j].offsetParent !== null) {
                        moreButtons[j].click();
                        didScroll = true;
                        break;  // Solo hacer clic en uno por vez
                    }
                }
                
                return didScroll;
            """
            
            did_specific_scroll = self.driver.execute_script(grid_scroll_script)
            
            # Si no se realizó ningún scroll específico, hacer scroll general
            if not did_specific_scroll:
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                
            return True
        except Exception as e:
            logger.debug(f"Error en scroll de tabla grid: {e}")
            return False

    def _scroll_generic(self):
        """
        Estrategia de scroll genérica para cualquier tipo de tabla o contenido.
        
        Aplica múltiples técnicas de scroll para maximizar la probabilidad
        de cargar más contenido en cualquier interfaz.
        
        Returns:
            bool: True si al menos un método de scroll fue exitoso, False en caso contrario
        """
        try:
            success = False
            
            # 1. Scroll normal al final de la página
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            success = True
            
            # 2. Enviar tecla END para scroll alternativo
            try:
                active_element = self.driver.switch_to.active_element
                active_element.send_keys(Keys.END)
                success = True
            except:
                pass
                
            # 3. Buscar y hacer scroll en cualquier contenedor con capacidad de scroll
            scroll_finder_script = """
                // Identificar todos los elementos con scroll vertical
                var scrollElements = Array.from(document.querySelectorAll('*')).filter(function(el) {
                    var style = window.getComputedStyle(el);
                    return (style.overflowY === 'scroll' || style.overflowY === 'auto') && 
                        el.scrollHeight > el.clientHeight;
                });
                
                var scrolledCount = 0;
                
                // Hacer scroll en cada elemento encontrado
                for (var i = 0; i < scrollElements.length; i++) {
                    var initialScrollTop = scrollElements[i].scrollTop;
                    scrollElements[i].scrollTop = scrollElements[i].scrollHeight;
                    
                    // Verificar si realmente se movió el scroll
                    if (scrollElements[i].scrollTop > initialScrollTop) {
                        scrolledCount++;
                    }
                }
                
                return scrolledCount;
            """
            
            scrolled_count = self.driver.execute_script(scroll_finder_script)
            if scrolled_count > 0:
                success = True
                
            return success
        except Exception as e:
            logger.debug(f"Error en scroll genérico: {e}")
            return False

    def _apply_alternative_scroll_strategy(self, attempt_count):
        """
        Aplica estrategias alternativas de scroll cuando las normales no funcionan.
        
        Alterna entre diferentes técnicas basadas en el número de intentos previos
        para maximizar las posibilidades de cargar más contenido.
        
        Args:
            attempt_count (int): Número de intentos previos sin cambios
            
        Returns:
            bool: True si la estrategia alternativa fue aplicada, False en caso contrario
        """
        try:
            # Rotar entre tres estrategias diferentes basadas en el contador de intentos
            strategy = attempt_count % 15
            
            if strategy < 5:
                # Estrategia 1: Scroll progresivo en incrementos
                logger.debug("Aplicando estrategia de scroll progresivo")
                for pos in range(0, 10000, 500):
                    self.driver.execute_script(f"window.scrollTo(0, {pos});")
                    time.sleep(0.1)
                    
            elif strategy < 10:
                # Estrategia 2: Uso de teclas de navegación
                logger.debug("Aplicando estrategia de teclas de navegación")
                try:
                    body = self.driver.find_element(By.TAG_NAME, "body")
                    # Alternar entre Page Down y End para máxima cobertura
                    for i in range(5):
                        body.send_keys(Keys.PAGE_DOWN)
                        time.sleep(0.1)
                        if i % 2 == 0:
                            body.send_keys(Keys.END)
                            time.sleep(0.1)
                except Exception as key_e:
                    logger.debug(f"Error en estrategia de teclas: {key_e}")
                    
            else:
                # Estrategia 3: Buscar y hacer clic en botones de carga
                logger.debug("Buscando botones de carga adicional")
                load_buttons_script = """
                    // Buscar botones de carga por texto y clase
                    var buttons = [];
                    
                    // Por texto
                    var textPatterns = ['More', 'más', 'Show', 'Ver', 'Load', 'Cargar', 'Next', 'Siguiente'];
                    for (var i = 0; i < textPatterns.length; i++) {
                        var pattern = textPatterns[i];
                        var matches = document.evaluate(
                            "//*[contains(text(), '" + pattern + "')]",
                            document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null
                        );
                        
                        for (var j = 0; j < matches.snapshotLength; j++) {
                            var element = matches.snapshotItem(j);
                            if (element.tagName === 'BUTTON' || element.tagName === 'A' || 
                                element.tagName === 'SPAN' || element.tagName === 'DIV') {
                                buttons.push(element);
                            }
                        }
                    }
                    
                    // Por clase
                    var classPatterns = [
                        'sapMListShowMoreButton', 'sapUiTableMoreBtn', 'sapMPaginatorButton',
                        'loadMore', 'showMore', 'moreButton', 'sapMBtn'
                    ];
                    
                    for (var k = 0; k < classPatterns.length; k++) {
                        var elements = document.getElementsByClassName(classPatterns[k]);
                        for (var l = 0; l < elements.length; l++) {
                            buttons.push(elements[l]);
                        }
                    }
                    
                    return buttons;
                """
                
                load_buttons = self.driver.execute_script(load_buttons_script)
                
                if load_buttons:
                    for btn in load_buttons[:3]:  # Limitar a 3 intentos
                        try:
                            # Hacer scroll hasta el botón
                            self.driver.execute_script(
                                "arguments[0].scrollIntoView({block: 'center', behavior: 'instant'});", 
                                btn
                            )
                            time.sleep(0.2)
                            
                            # Intentar clic
                            self.driver.execute_script("arguments[0].click();", btn)
                            logger.info("Se hizo clic en botón de carga adicional")
                            time.sleep(1.5)  # Esperar a que cargue
                            return True
                        except Exception as btn_e:
                            logger.debug(f"Error al hacer clic en botón: {btn_e}")
                            continue
            
            return True
        except Exception as e:
            logger.debug(f"Error en estrategia alternativa de scroll: {e}")
            return False

    def _should_finish_scrolling(self, no_change_count, current_rows_count, total_expected):
        """
        Determina si se debe finalizar el proceso de scroll basado en criterios adaptativos.
        
        Evalúa múltiples factores para decidir cuándo es el momento óptimo
        para terminar el proceso de carga.
        
        Args:
            no_change_count (int): Número de intentos sin cambios en el conteo de filas
            current_rows_count (int): Número actual de filas detectadas
            total_expected (int): Número total esperado de filas
            
        Returns:
            bool: True si se debe finalizar el scroll, False si se debe continuar
        """
        try:
            # Calcular porcentaje de cobertura
            coverage_percentage = (current_rows_count / total_expected * 100) if total_expected > 0 else 0
            
            # Criterio 1: Muchos intentos sin cambios y buena cobertura (≥90%)
            if no_change_count >= 10 and current_rows_count >= total_expected * 0.9:
                logger.info(f"Finalizando scroll: suficiente cobertura ({coverage_percentage:.1f}%, {current_rows_count}/{total_expected})")
                return True
                
            # Criterio 2: Demasiados intentos sin cambios (indicador de que no hay más contenido)
            if no_change_count >= 20:
                logger.info(f"Finalizando scroll: muchos intentos sin cambios ({no_change_count})")
                return True
                
            # Criterio 3: Se superó el total esperado
            if current_rows_count >= total_expected:
                logger.info(f"Finalizando scroll: se alcanzó o superó el total esperado ({current_rows_count}/{total_expected})")
                return True
            
            # Criterio 4: Cobertura muy alta (≥95%) incluso con pocos intentos sin cambios
            if coverage_percentage >= 95 and no_change_count >= 5:
                logger.info(f"Finalizando scroll: cobertura excelente ({coverage_percentage:.1f}%) con {no_change_count} intentos sin cambios")
                return True
                
            # Continuar con el scroll
            return False
        except Exception as e:
            logger.debug(f"Error al evaluar criterios de finalización: {e}")
            return no_change_count > 15  # Criterio de seguridad en caso de error

    def _calculate_adaptive_wait_time(self, no_change_count, current_rows_count):
        """
        Calcula un tiempo de espera adaptativo basado en el progreso de carga.
        
        Ajusta dinámicamente el tiempo de espera entre operaciones de scroll
        para optimizar la carga de datos.
        
        Args:
            no_change_count (int): Número de intentos sin cambios en el conteo de filas
            current_rows_count (int): Número actual de filas detectadas
            
        Returns:
            float: Tiempo de espera en segundos
        """
        try:
            # Base: tiempo corto para maximizar rendimiento
            base_wait = 0.2
            
            # Factor basado en intentos sin cambios (incrementar gradualmente)
            if no_change_count > 0:
                no_change_factor = min(no_change_count * 0.1, 1.0)
            else:
                no_change_factor = 0
                
            # Factor basado en cantidad de filas (más filas = más tiempo para procesar)
            rows_factor = min(current_rows_count / 500, 0.5)
            
            # Calcular tiempo final, con límite máximo de 1 segundo
            wait_time = min(base_wait + no_change_factor + rows_factor, 1.0)
            
            return wait_time
        except Exception as e:
            logger.debug(f"Error al calcular tiempo adaptativo: {e}")
            return 0.5  # Valor por defecto en caso de error



    def _extract_by_headers(self, row, header_map):
        """Extrae datos directamente basado en el mapa de encabezados con mejor control de errores"""
        try:
            cells = self._get_row_cells(row)
            if not cells:
                logger.debug("No se encontraron celdas en la fila")
                return None
                    
            # Inicializar diccionario de resultados con valores por defecto
            issue_data = {
                'Title': '',
                'Type': '',
                'Priority': '',
                'Status': '',
                'Deadline': '',
                'Due Date': '',
                'Created By': '',
                'Created On': ''
            }
            
            # Mapeo de nombres de encabezados a claves en nuestro diccionario (más flexible)
            header_mappings = {
                'TITLE': 'Title',
                'TYPE': 'Type',
                'PRIORITY': 'Priority',
                'STATUS': 'Status',
                'DEADLINE': 'Deadline',
                'DUE DATE': 'Due Date',
                'CREATED BY': 'Created By',
                'CREATED ON': 'Created On',
                # Añadir mapeos alternativos para diferentes nomenclaturas
                'NAME': 'Title',
                'ISSUE': 'Title',
                'PRIO': 'Priority',
                'STATE': 'Status',
                'DUE': 'Due Date'
            }
            
            # Extraer valores usando el mapa de encabezados
            for header, index in header_map.items():
                if index < len(cells):
                    # Buscar la clave correspondiente
                    header_upper = header.upper()
                    matched = False
                    
                    # Buscar coincidencias exactas o parciales
                    for pattern, key in header_mappings.items():
                        if pattern == header_upper or pattern in header_upper:
                            cell_text = cells[index].text.strip() if cells[index].text else ''
                            issue_data[key] = cell_text
                            matched = True
                            break
                    
                    # Registrar encabezados no reconocidos
                    if not matched and header_upper:
                        logger.debug(f"Encabezado no reconocido: '{header_upper}'")
                
            # Validación mínima - verificar que al menos tengamos un título
            if not issue_data['Title'] and len(cells) > 0:
                issue_data['Title'] = cells[0].text.strip() if cells[0].text else "Issue sin título"
                    
            return issue_data
        except Exception as e:
            logger.debug(f"Error en extracción por encabezados: {e}")
            return None





        
    
    def _process_table_rows(self, rows, seen_titles, header_map=None):
        """Procesa las filas de la tabla y extrae los datos de cada issue"""
        issues_data = []
        processed_count = 0
        batch_size = 10  # Procesar en lotes para actualizar progreso
        
        for index, row in enumerate(rows):
            try:
                
                if header_map and len(header_map) >= 4:
                    try:
                        issue_data = self._extract_by_headers(row, header_map)
                        if issue_data and issue_data['Title']:
                            # Validar y corregir los datos
                            corrected_issue = self._validate_and_correct_issue_data(issue_data)
                            issues_data.append(corrected_issue)
                            processed_count += 1
                            continue  # Pasar a la siguiente fila
                    except Exception as header_e:
                        logger.debug(f"Error en extracción por encabezados: {header_e}, usando método alternativo")
                
                
                
                
                # Extraer todos los datos en un solo paso para análisis conjunto
                title = self._extract_title(row)
                
                if not title:
                    title = f"Issue sin título #{index+1}"
                
                # Extraer resto de datos
                type_text = self._extract_type(row, title)
                priority = self._extract_priority(row)
                status = self._extract_status(row)
                deadline = self._extract_deadline(row)
                due_date = self._extract_due_date(row)
                created_by = self._extract_created_by(row)
                created_on = self._extract_created_on(row)
                
                # Datos del issue completos
                issue_data = {
                    'Title': title,
                    'Type': type_text,
                    'Priority': priority,
                    'Status': status,
                    'Deadline': deadline,
                    'Due Date': due_date,
                    'Created By': created_by,
                    'Created On': created_on
                }
                
                # Validación especial: verificar si los datos están desplazados
                if type_text == title:
                    logger.warning(f"Posible desplazamiento de columnas detectado en issue '{title}' (Type = Title)")
                    
                    # Intentar extraer directamente por orden de columnas
                    cells = self._get_row_cells(row)
                    if cells and len(cells) >= 8:
                        # Extraer datos directamente por posición en la tabla
                        issue_data = {
                            'Title': cells[0].text.strip() if cells[0].text else title,
                            'Type': cells[1].text.strip() if len(cells) > 1 and cells[1].text else "",
                            'Priority': cells[2].text.strip() if len(cells) > 2 and cells[2].text else "",
                            'Status': cells[3].text.strip() if len(cells) > 3 and cells[3].text else "",
                            'Deadline': cells[4].text.strip() if len(cells) > 4 and cells[4].text else "",
                            'Due Date': cells[5].text.strip() if len(cells) > 5 and cells[5].text else "",
                            'Created By': cells[6].text.strip() if len(cells) > 6 and cells[6].text else "",
                            'Created On': cells[7].text.strip() if len(cells) > 7 and cells[7].text else ""
                        }
                        logger.info(f"Extracción directa por celdas realizada para issue '{title}'")
                
                # Validar y corregir los datos
                corrected_issue = self._validate_and_correct_issue_data(issue_data)
                
                # Añadir a la lista de resultados
                issues_data.append(corrected_issue)
                processed_count += 1
                
                if processed_count % batch_size == 0:
                    logger.info(f"Procesados {processed_count} issues hasta ahora")
                
            except Exception as e:
                logger.error(f"Error al procesar la fila {index}: {e}")
        
        logger.info(f"Procesamiento de filas completado. Total procesado: {processed_count} issues")
        return issues_data










    def _extract_title(self, row):
        """Extrae el título de una fila"""
        try:
            # Intentar múltiples métodos para extraer título
            title_extractors = [
                lambda r: r.find_element(By.XPATH, ".//a").text,
                lambda r: r.find_element(By.XPATH, ".//span[contains(@class, 'title')]").text,
                lambda r: r.find_element(By.XPATH, ".//div[contains(@class, 'title')]").text,
                lambda r: r.find_elements(By.XPATH, ".//div[@role='gridcell']")[0].text if r.find_elements(By.XPATH, ".//div[@role='gridcell']") else None,
                lambda r: r.find_elements(By.XPATH, ".//td")[0].text if r.find_elements(By.XPATH, ".//td") else None,
                lambda r: r.find_element(By.XPATH, ".//*[contains(@id, 'title')]").text,
                lambda r: r.find_element(By.XPATH, ".//div[@title]").get_attribute("title"),
                lambda r: r.find_element(By.XPATH, ".//span[@title]").get_attribute("title")
            ]

            for extractor in title_extractors:
                try:
                    title_text = extractor(row)
                    if title_text and title_text.strip():
                        return title_text.strip()
                except:
                    continue
            
            # Si no encontramos un título específico, usar el texto completo
            try:
                full_text = row.text.strip()
                if full_text:
                    lines = full_text.split('\n')
                    if lines:
                        title = lines[0].strip()
                        if len(title) > 100:  # Si es muy largo, recortar
                            title = title[:100] + "..."
                        return title
            except:
                pass
                
            return None
        except Exception as e:
            logger.debug(f"Error al extraer título: {e}")
            return None
    
    
    
    
    
    
    
    


    def _extract_type(self, row, title):
        """Extrae correctamente el tipo de issue"""
        try:
            # Buscar con mayor precisión el tipo
            cells = self._get_row_cells(row)
            
            # Verificar si tenemos suficientes celdas
            if cells and len(cells) >= 2:
                # Extraer de la segunda celda, pero verificar que no sea igual al título
                type_text = cells[1].text.strip()
                
                # Si el tipo es igual al título, algo está mal, buscar en otra parte
                if type_text and type_text != title:
                    return type_text
            
            # Intentos alternativos para obtener el tipo
            # Buscar elementos específicos con clases o atributos que indiquen tipo
            type_elements = row.find_elements(By.XPATH, ".//span[contains(@class, 'type')] | .//div[contains(@class, 'type')]")
            for el in type_elements:
                if el.text and el.text.strip() and el.text.strip() != title:
                    return el.text.strip()
            
            # Intentar con selectores UI5 específicos como último recurso
            ui5_types = self.find_ui5_elements("sap.m.Label", {"text": "Type"})
            for type_label in ui5_types:
                try:
                    value_element = self.driver.find_element(By.XPATH, f"./following-sibling::*[1]")
                    if value_element and value_element.text and value_element.text.strip() != title:
                        return value_element.text.strip()
                except:
                    pass
                    
            # Si no encontramos nada, devolver vacío
            return ""
        except Exception as e:
            logger.debug(f"Error al extraer tipo: {e}")
            return ""
        
    
    
    def _extract_priority(self, row):
        """Extrae la prioridad del issue"""
        try:
            # Buscar específicamente en la tercera columna 
            cells = self._get_row_cells(row)
            if cells and len(cells) >= 3:
                priority_text = cells[2].text.strip()
                if priority_text:
                    return self._normalize_priority(priority_text)
            
            # Intentos alternativos
            priority_indicators = [
                # Por clase de color
                (By.XPATH, ".//span[contains(@class, 'sapMGaugeNegativeColor')]", "Very High"),
                (By.XPATH, ".//span[contains(@class, 'sapMGaugeCriticalColor')]", "High"),
                (By.XPATH, ".//span[contains(@class, 'sapMGaugeNeutralColor')]", "Medium"),
                (By.XPATH, ".//span[contains(@class, 'sapMGaugePositiveColor')]", "Low"),
                
                # Por texto
                (By.XPATH, ".//span[contains(text(), 'Very High')]", "Very High"),
                (By.XPATH, ".//span[contains(text(), 'High') and not(contains(text(), 'Very'))]", "High"),
                (By.XPATH, ".//span[contains(text(), 'Medium')]", "Medium"),
                (By.XPATH, ".//span[contains(text(), 'Low')]", "Low")
            ]
            
            # Buscar indicadores visuales de prioridad
            for locator, indicator_text in priority_indicators:
                elements = row.find_elements(locator)
                if elements:
                    return indicator_text
            
            # Buscar por etiquetas o campos específicos
            priority_labels = row.find_elements(By.XPATH, 
                ".//div[contains(text(), 'Priority')]/following-sibling::*[1] | " +
                ".//span[contains(text(), 'Priority')]/following-sibling::*[1]")
            if priority_labels:
                for label in priority_labels:
                    if label.text:
                        return self._normalize_priority(label.text)
            
            # Verificar por los valores específicos en cualquier lugar de la fila
            for priority_value in ["Very High", "High", "Medium", "Low"]:
                elements = row.find_elements(By.XPATH, f".//*[contains(text(), '{priority_value}')]")
                if elements:
                    return priority_value
                    
            return ""
        except Exception as e:
            logger.debug(f"Error al extraer prioridad: {e}")
            return ""
    
    
    
    
    
    
    
    
    
            
    def _normalize_priority(self, priority_text):
        """Normaliza el texto de prioridad"""
        if not priority_text:
            return ""
            
        priority_lower = priority_text.lower()
        
        if "very high" in priority_lower:
            return "Very High"
        elif "high" in priority_lower:
            return "High"
        elif "medium" in priority_lower:
            return "Medium"
        elif "low" in priority_lower:
            return "Low"
        
        return priority_text







    def _normalize_status(self, status_text):
        """Normaliza el texto de estado"""
        if not status_text:
            return ""
            
        status_upper = status_text.upper()
        
        if "OPEN" in status_upper:
            return "OPEN"
        elif "DONE" in status_upper:
            return "DONE"
        elif "IN PROGRESS" in status_upper:
            return "IN PROGRESS"
        elif "READY" in status_upper:
            return "READY FOR PUBLISHING" if "PUBLISH" in status_upper else "READY"
        elif "ACCEPTED" in status_upper:
            return "ACCEPTED"
        elif "DRAFT" in status_upper:
            return "DRAFT"
        elif "CLOSED" in status_upper:
            return "CLOSED"
        
        return status_text







    def _extract_status(self, row):
        """Extrae el estado del issue"""
        try:
            # Buscar específicamente en la cuarta columna
            cells = self._get_row_cells(row)
            if cells and len(cells) >= 4:
                status_text = cells[3].text.strip()
                if status_text:
                    # Limpiar y extraer solo la primera línea si hay varias
                    status_lines = status_text.split("\n")
                    status = status_lines[0].strip()
                    status = status.replace("Object Status", "").strip()
                    return self._normalize_status(status)
            
            # Buscar por estados conocidos de SAP
            status_patterns = ["OPEN", "DONE", "IN PROGRESS", "READY", "ACCEPTED", "DRAFT"]
            for status in status_patterns:
                elements = row.find_elements(By.XPATH, f".//*[contains(text(), '{status}')]")
                if elements:
                    return status
                    
            # Buscar por etiquetas o campos específicos
            status_labels = row.find_elements(By.XPATH, 
                ".//div[contains(text(), 'Status')]/following-sibling::*[1] | " +
                ".//span[contains(text(), 'Status')]/following-sibling::*[1]")
            if status_labels:
                for label in status_labels:
                    if label.text:
                        return label.text.strip()
                        
            return ""
        except Exception as e:
            logger.debug(f"Error al extraer estado: {e}")
            return ""
        
        
        
        
        
                
    def _extract_deadline(self, row):
        """Extrae la fecha límite del issue"""
        try:
            # Buscar específicamente en la quinta columna
            cells = self._get_row_cells(row)
            if cells and len(cells) >= 5:
                deadline_text = cells[4].text.strip()
                if deadline_text and any(month in deadline_text for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
                    return deadline_text
            
            # Buscar por formato de fecha
            date_elements = row.find_elements(By.XPATH, ".//*[contains(text(), '/') or contains(text(), '-') or contains(text(), 'day,')]")
            for el in date_elements:
                # Verificar si parece una fecha por formato o contenido
                text = el.text.strip()
                if text and any(month in text for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
                    # Verificar que no sea la fecha "Created On"
                    parent_text = ""
                    try:
                        parent = el.find_element(By.XPATH, "./parent::*")
                        parent_text = parent.text
                    except:
                        pass
                    
                    if "created" not in parent_text.lower() and "due" not in parent_text.lower():
                        return text
                        
            # Buscar por etiquetas específicas
            deadline_labels = row.find_elements(By.XPATH, 
                ".//div[contains(text(), 'Deadline')]/following-sibling::*[1] | " +
                ".//span[contains(text(), 'Deadline')]/following-sibling::*[1]")
            if deadline_labels:
                for label in deadline_labels:
                    if label.text:
                        return label.text.strip()
                        
            return ""
        except Exception as e:
            logger.debug(f"Error al extraer deadline: {e}")
            return ""
        
        
        
        
        
        
        
        
        
                
    def _extract_due_date(self, row):
        """Extrae la fecha de vencimiento del issue"""
        try:
            # Buscar específicamente en la sexta columna
            cells = self._get_row_cells(row)
            if cells and len(cells) >= 6:
                due_date_text = cells[5].text.strip()
                if due_date_text and any(month in due_date_text for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
                    return due_date_text
            
            # Buscar por etiquetas específicas
            due_date_labels = row.find_elements(By.XPATH, 
                ".//div[contains(text(), 'Due') or contains(text(), 'Due Date')]/following-sibling::*[1] | " +
                ".//span[contains(text(), 'Due') or contains(text(), 'Due Date')]/following-sibling::*[1]")
            if due_date_labels:
                for label in due_date_labels:
                    if label.text:
                        return label.text.strip()
            
            # Última posibilidad: buscar fechas que no sean deadline ni created on
            date_elements = row.find_elements(By.XPATH, ".//*[contains(text(), '/') or contains(text(), '-') or contains(text(), 'day,')]")
            if len(date_elements) >= 2:  # Si hay al menos 2 fechas, la segunda podría ser due date
                return date_elements[1].text.strip()
                
            return ""
        except Exception as e:
            logger.debug(f"Error al extraer due date: {e}")
            return ""
        
                
    def _extract_created_by(self, row):
        """Extrae quién creó el issue"""
        try:
            # Buscar específicamente en la séptima columna
            cells = self._get_row_cells(row)
            if cells and len(cells) >= 7:
                created_by_text = cells[6].text.strip()
                # Verificar que no sea una fecha (para evitar confusión con otras columnas)
                if created_by_text and not any(month in created_by_text for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
                    return created_by_text
            
            # Buscar por etiquetas específicas
            created_by_labels = row.find_elements(By.XPATH, 
                ".//div[contains(text(), 'Created By')]/following-sibling::*[1] | " +
                ".//span[contains(text(), 'Created By')]/following-sibling::*[1]")
            if created_by_labels:
                for label in created_by_labels:
                    if label.text:
                        return label.text.strip()
                        
            # Buscar elementos que parecen ser usuarios (con formato de ID)
            user_patterns = row.find_elements(By.XPATH, ".//*[contains(text(), 'I') and string-length(text()) <= 8]")
            if user_patterns:
                for user in user_patterns:
                    user_text = user.text.strip()
                    # Si parece un ID de usuario de SAP (como I587465)
                    if user_text.startswith("I") and user_text[1:].isdigit():
                        return user_text
                        
            return ""
        except Exception as e:
            logger.debug(f"Error al extraer creador: {e}")
            return ""
        
        
        
        
        
        
                
    def _extract_created_on(self, row):
        """Extrae la fecha de creación del issue"""
        try:
            # Buscar específicamente en la octava columna
            cells = self._get_row_cells(row)
            if cells and len(cells) >= 8:
                created_on_text = cells[7].text.strip()
                if created_on_text and any(month in created_on_text for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
                    return created_on_text
            
            # Buscar por etiquetas específicas
            created_on_labels = row.find_elements(By.XPATH, 
                ".//div[contains(text(), 'Created On')]/following-sibling::*[1] | " +
                ".//span[contains(text(), 'Created On')]/following-sibling::*[1]")
            if created_on_labels:
                for label in created_on_labels:
                    if label.text:
                        return label.text.strip()
            
            # Buscar la última fecha disponible que podría ser la fecha de creación
            date_elements = row.find_elements(By.XPATH, ".//*[contains(text(), '/') or contains(text(), '-') or contains(text(), 'day,')]")
            if date_elements and len(date_elements) >= 3:  # Si hay al menos 3 fechas, la última podría ser created on
                return date_elements[-1].text.strip()
                
            return ""
        except Exception as e:
            logger.debug(f"Error al extraer fecha de creación: {e}")
            return ""        
        
        
        
        
        
    def _get_row_cells(self, row):
        """Método mejorado para obtener todas las celdas de una fila"""
        cells = []
        
        try:
            # Intentar diferentes métodos para obtener celdas en orden de prioridad
            cell_extractors = [
                # 1. Buscar elementos td directamente (tabla HTML estándar)
                lambda r: r.find_elements(By.XPATH, ".//td"),
                
                # 2. Buscar celdas específicas de SAP UI5
                lambda r: r.find_elements(By.XPATH, ".//div[@role='gridcell']"),
                
                # 3. Buscar elementos con clases que indiquen que son celdas
                lambda r: r.find_elements(By.XPATH, ".//*[contains(@class, 'cell') or contains(@class, 'Cell')]"),
                
                # 4. Buscar divs hijos directos como posible celda
                lambda r: r.find_elements(By.XPATH, "./div[not(contains(@class, 'sapUiNoContentPadding'))]"),
                
                # 5. Buscar spans con información relevante
                lambda r: r.find_elements(By.XPATH, ".//span[contains(@id, 'col')]")
            ]
            
            # Intentar cada método hasta encontrar celdas
            for extractor in cell_extractors:
                try:
                    extracted_cells = extractor(row)
                    if extracted_cells and len(extracted_cells) > 2:  # Necesitamos al menos 3 celdas para ser válidas
                        # Verificar que las celdas tengan texto
                        if all(cell.text.strip() for cell in extracted_cells[:3]):
                            return extracted_cells
                except:
                    continue
                    
            # Intentar método más específico basado en las capturas de pantalla
            try:
                # Localizar por columnas basadas en la estructura de las imágenes
                columns = ["Title", "Type", "Priority", "Status", "Deadline", "Due Date", "Created By", "Created On"]
                column_cells = []
                
                for i, column in enumerate(columns):
                    # Intentar localizar celda específica por su posición o atributos
                    xpath_patterns = [
                        f".//div[contains(@aria-label, '{column}')]",
                        f".//div[contains(@aria-colindex, '{i+1}')]",
                        f".//div[contains(@data-column-index, '{i}')]",
                        f".//div[contains(@class, 'col{i+1}')]"
                    ]
                    
                    for xpath in xpath_patterns:
                        cell_candidates = row.find_elements(By.XPATH, xpath)
                        if cell_candidates:
                            column_cells.append(cell_candidates[0])
                            break
                
                if len(column_cells) >= 3:
                    return column_cells
            except:
                pass
            
            # Estrategia de respaldo: buscar todos los elementos con texto en la fila
            if not cells:
                # Buscar todos los elementos con texto visible
                text_elements = row.find_elements(By.XPATH, ".//*[normalize-space(text())]")
                
                # Filtrar los que parezcan ser encabezados o elementos de UI
                filtered_elements = []
                for el in text_elements:
                    # Excluir elementos que son parte de la UI, no datos
                    classes = el.get_attribute("class") or ""
                    if not any(ui_class in classes.lower() for ui_class in ["icon", "button", "checkbox", "arrow", "header"]):
                        filtered_elements.append(el)
                
                # Devolver si tenemos suficientes elementos
                if len(filtered_elements) >= 4:
                    return filtered_elements
                    
        except Exception as e:
            logger.debug(f"Error al extraer celdas: {e}")
        
        return cells    






    def _validate_and_correct_issue_data(self, issue_data):
        """Valida y corrige los datos del issue antes de guardarlos"""
        # Asegurar que todos los campos esperados estén presentes
        required_fields = ['Title', 'Type', 'Priority', 'Status', 'Deadline', 'Due Date', 'Created By', 'Created On']
        for field in required_fields:
            if field not in issue_data:
                issue_data[field] = ""
        
        # Validación y corrección contextual de los campos
        
        # 1. Verificar que Type no duplique el Title
        if issue_data['Type'] == issue_data['Title']:
            issue_data['Type'] = ""
        
        # 2. Verificar Status - debería contener palabras clave como "OPEN", "DONE", etc.
        if issue_data['Status']:
            status_keywords = ["OPEN", "DONE", "IN PROGRESS", "READY", "ACCEPTED", "DRAFT", "CLOSED"]
            if not any(keyword in issue_data['Status'].upper() for keyword in status_keywords):
                # Si Status no parece un status válido, verificar otros campos
                for field in ['Priority', 'Type', 'Deadline']:
                    if field in issue_data and issue_data[field]:
                        field_value = issue_data[field].upper()
                        if any(keyword in field_value for keyword in status_keywords):
                            # Intercambiar valores
                            temp = issue_data['Status']
                            issue_data['Status'] = issue_data[field]
                            issue_data[field] = temp
                            break
        
        # 3. Verificar prioridad - debería ser "High", "Medium", "Low", etc.
        if issue_data['Priority']:
            priority_keywords = ["HIGH", "MEDIUM", "LOW", "VERY HIGH"]
            if not any(keyword in issue_data['Priority'].upper() for keyword in priority_keywords):
                # Si Priority no parece una prioridad válida, verificar otros campos
                for field in ['Type', 'Status']:
                    if field in issue_data and issue_data[field]:
                        field_value = issue_data[field].upper()
                        if any(keyword in field_value for keyword in priority_keywords):
                            # Intercambiar valores
                            temp = issue_data['Priority']
                            issue_data['Priority'] = issue_data[field]
                            issue_data[field] = temp
                            break
        
        # 4. Verificar que Deadline, Due Date y Created On parezcan fechas
        date_fields = ['Deadline', 'Due Date', 'Created On']
        date_keywords = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
        
        for date_field in date_fields:
            if date_field in issue_data and issue_data[date_field]:
                # Verificar si parece una fecha
                if not any(month in issue_data[date_field].upper() for month in date_keywords):
                    # No parece una fecha, buscar en otros campos que no son fechas
                    for field in ['Status', 'Priority', 'Type']:
                        if field in issue_data and issue_data[field]:
                            field_value = issue_data[field].upper()
                            if any(month in field_value for month in date_keywords):
                                # Intercambiar valores
                                temp = issue_data[date_field]
                                issue_data[date_field] = issue_data[field]
                                issue_data[field] = temp
                                break
        
        # 5. Created By debería parecer un ID de usuario (no una fecha o un status)
        if issue_data['Created By']:
            if any(month in issue_data['Created By'].upper() for month in date_keywords):
                # Parece una fecha, buscar un mejor valor para Created By
                for field in date_fields:
                    if field in issue_data and not issue_data[field]:
                        # Si encontramos un campo de fecha vacío, mover Created By allí
                        issue_data[field] = issue_data['Created By']
                        issue_data['Created By'] = ""
                        break
        
        # 6. Verificar inconsistencia de desplazamiento general
        # Si detectamos un patrón de desplazamiento, corregirlo
        if (issue_data['Type'] == issue_data['Title'] and 
            issue_data['Priority'] == issue_data['Type'] and 
            issue_data['Status'] == issue_data['Priority']):
            
            # Desplazar todos los campos a la izquierda
            issue_data['Type'] = issue_data['Priority']
            issue_data['Priority'] = issue_data['Status']
            issue_data['Status'] = issue_data['Deadline']
            issue_data['Deadline'] = issue_data['Due Date']
            issue_data['Due Date'] = issue_data['Created By']
            issue_data['Created By'] = issue_data['Created On']
            issue_data['Created On'] = ""
        
        return issue_data





    def find_ui5_elements(self, control_type, properties=None):
        """Encuentra elementos UI5 específicos usando JavaScript"""
        script = """
        function findUI5Controls(controlType, properties) {
            if (!window.sap || !window.sap.ui) return [];
            
            var controls = sap.ui.getCore().byFieldGroupId().filter(function(control) {
                return control.getMetadata().getName() === controlType;
            });
            
            if (properties) {
                controls = controls.filter(function(control) {
                    for (var prop in properties) {
                        if (control.getProperty(prop) !== properties[prop]) {
                            return false;
                        }
                    }
                    return true;
                });
            }
            
            return controls.map(function(control) {
                return control.getId();
            });
        }
        return findUI5Controls(arguments[0], arguments[1]);
        """
        
        try:
            control_ids = self.driver.execute_script(script, control_type, properties)
            elements = []
            
            for control_id in control_ids:
                try:
                    element = self.driver.find_element(By.ID, control_id)
                    elements.append(element)
                except:
                    pass
                    
            return elements
        except:
            return []
    
    def close(self):
        """Cierra el navegador"""
        if self.driver:
            try:
                self.driver.quit()
                logger.info("Navegador cerrado correctamente")
                return True
            except Exception as e:
                logger.error(f"Error al cerrar el navegador: {e}")
                return False
        return True





class IssuesExtractor:
    """Clase principal para extraer issues de SAP con interfaz gráfica y base de datos"""

    def __init__(self):
        """Inicializa la clase"""
        self.excel_file_path = None
        self.driver = None
        
        # Variables para la GUI
        self.root = None
        self.status_var = None
        self.client_var = None
        self.project_var = None
        self.project_combo = None
        self.log_text = None
        self.excel_filename_var = None
        self.processing = False
        self.left_panel = None
        self.header_frame = None
        self.client_combo = None
        self.image_cache = {}
        
        # Componentes
        self.db_manager = DatabaseManager()
        self.excel_manager = ExcelManager()
        self.browser = SAPBrowser()
        
    def choose_excel_file(self):
        """Permite al usuario elegir un archivo Excel existente o crear uno nuevo"""
        file_path = self.excel_manager.select_file()
        self.excel_file_path = file_path
        self.excel_manager.file_path = file_path
        
        # Actualizar la interfaz si existe
        if hasattr(self, 'status_var') and self.status_var:
            self.status_var.set(f"Archivo Excel seleccionado: {os.path.basename(file_path)}")
        
        # Actualizar el nombre del archivo en la etiqueta
        if hasattr(self, 'excel_filename_var') and self.excel_filename_var:
            self.excel_filename_var.set(f"Archivo: {os.path.basename(file_path)}")
            
        return file_path
        
    def connect_to_browser(self):
        """Conecta con el navegador"""
        result = self.browser.connect()
        self.driver = self.browser.driver
        return result
        
    def update_excel(self, issues_data):
        """Actualiza el archivo Excel con los datos extraídos"""
        success, new_items, updated_items = self.excel_manager.update_with_issues(issues_data)
        
        # Actualizar la interfaz si existe
        if hasattr(self, 'status_var') and self.status_var:
            if success:
                self.status_var.set(f"Excel actualizado: {new_items} nuevos, {updated_items} actualizados")
            else:
                self.status_var.set("Error al actualizar Excel")
                
        # Mostrar mensaje de éxito
        if success and self.root:
            messagebox.showinfo(
                "Proceso Completado", 
                f"El archivo Excel ha sido actualizado correctamente.\n\n"
                f"Se han agregado {new_items} nuevos issues y actualizado {updated_items} issues existentes."
            )
            
        return success
        










    def run_extraction(self):
        """Ejecuta el proceso completo de extracción"""
        try:
            if not self.connect_to_browser():
                logger.error("Error al conectar con el navegador")
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set("Error al conectar con el navegador")
                    
                return False

            # Obtener valores de cliente y proyecto - APLICAR TRIM PARA ELIMINAR ESPACIOS
            erp_number = self.client_var.get().strip() if hasattr(self, 'client_var') and self.client_var else "1025541"
            project_id = self.project_var.get().strip() if hasattr(self, 'project_var') and self.project_var else "20096444"

            if not erp_number:
                logger.warning("ERP number está vacío, usando valor por defecto")
                erp_number = "1025541"
                
            if not project_id:
                logger.warning("Project ID está vacío, usando valor por defecto")
                project_id = "20096444"
                
            logger.info(f"Iniciando extracción para cliente: {erp_number}, proyecto: {project_id}")
        
            # Navegar a la URL inicial especificada
            logger.info("Navegando a la URL de SAP con parámetros específicos...")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Navegando a SAP...")
                
            if not self.browser.navigate_to_sap(erp_number, project_id):
                logger.error("Error al navegar a la URL de SAP")
                return False
            
            # Manejar autenticación si es necesario
            if not self.browser.handle_authentication():
                logger.error("Error en el proceso de autenticación")
                return False
                
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Seleccionando cliente automáticamente...")
            
            # Seleccionar cliente
            if not self.browser.select_customer_automatically(erp_number):
                logger.warning("No se pudo seleccionar cliente automáticamente")
                # Solicitar selección manual si es necesario
                if self.root:
                    messagebox.showwarning("Selección Manual Requerida", 
                        "No se pudo seleccionar el cliente automáticamente.\n\n"
                        "Por favor, seleccione manualmente el cliente y haga clic en Continuar.")
                    result = messagebox.askokcancel("Confirmación", "¿Ha seleccionado el cliente?")
                    if not result:
                        return False
            
            # Actualizar la interfaz
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Seleccionando proyecto automáticamente...")
            
            # Seleccionar proyecto
            if not self.browser.select_project_automatically(project_id):
                logger.warning("No se pudo seleccionar proyecto automáticamente")
                # Solicitar selección manual si es necesario
                if self.root:
                    messagebox.showwarning("Selección Manual Requerida", 
                        "No se pudo seleccionar el proyecto automáticamente.\n\n"
                        "Por favor, seleccione manualmente el proyecto y haga clic en Continuar.")
                    result = messagebox.askokcancel("Confirmación", "¿Ha seleccionado el proyecto?")
                    if not result:
                        return False
            
            # Hacer clic en botón de búsqueda
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Realizando búsqueda...")
                
            self.browser.click_search_button()
            
            # Continuar con la detección y extracción de issues
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Navegando a la pestaña Issues...")
            
            # Navegar a la pestaña Issues
            if not self.browser.navigate_to_issues_tab():
                logger.warning("No se pudo navegar automáticamente a la pestaña Issues")
                if self.root:
                    messagebox.showwarning("Navegación Manual Requerida", 
                        "Por favor, navegue manualmente a la pestaña 'Issues' y luego haga clic en Continuar.")
                    result = messagebox.askokcancel("Confirmación", "¿Ha navegado a la pestaña Issues?")
                    if not result:
                        return False
            
            # Realizar la extracción
            return self.perform_extraction()
                
        except Exception as e:
            logger.error(f"Error en el proceso de extracción: {e}")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error: {e}")
                
            return False    
    
    
    
    
    
    def navigate_to_issues_tab(self):
        """
        Navega a la pestaña 'Issues' una vez seleccionado el proyecto.
        
        Returns:
            bool: True si la navegación fue exitosa, False en caso contrario
        """
        try:
            logger.info("Intentando navegar a la pestaña Issues...")
            
            # Esperar a que cargue la página del proyecto
            time.sleep(3)
            
            # Buscar la pestaña de Issues por diferentes selectores
            issues_tab_selectors = [
                "//div[contains(text(), 'Issues')] | //span[contains(text(), 'Issues')]",
                "//li[@role='tab']//div[contains(text(), 'Issues')]",
                "//a[contains(text(), 'Issues')]",
                "//div[contains(@class, 'sapMITBItem')]//span[contains(text(), 'Issues')]"
            ]
            
            for selector in issues_tab_selectors:
                try:
                    issues_tabs = self.driver.find_elements(By.XPATH, selector)
                    if issues_tabs:
                        for tab in issues_tabs:
                            try:
                                # Verificar si es visible
                                if tab.is_displayed():
                                    # Hacer scroll hasta el elemento
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                                    time.sleep(0.5)
                                    
                                    # Intentar clic
                                    self.driver.execute_script("arguments[0].click();", tab)
                                    logger.info("Clic en pestaña Issues realizado")
                                    time.sleep(3)  # Esperar a que cargue
                                    return True
                            except:
                                continue
                except:
                    continue
            
            logger.warning("No se encontró la pestaña Issues por selectores directos")
            
            # Intentar buscar por posición relativa (generalmente la tercera pestaña)
            try:
                tabs = self.driver.find_elements(By.XPATH, "//li[@role='tab'] | //div[@role='tab']")
                if len(tabs) >= 3:  # Asumiendo que Issues es la tercera pestaña
                    third_tab = tabs[2]  # Índice 2 para el tercer elemento
                    self.driver.execute_script("arguments[0].click();", third_tab)
                    logger.info("Clic en tercera pestaña realizado")
                    time.sleep(3)
                    return True
            except:
                pass
                
            logger.warning("No se pudo navegar a la pestaña Issues")
            return False
            
        except Exception as e:
            logger.error(f"Error al navegar a la pestaña Issues: {e}")
            return False
    
    
    
               
    def perform_extraction(self):
        """Método principal para ejecutar el proceso de extracción"""
        try:
            # Marcar como procesando
            self.processing = True
            
            logger.info("Comenzando proceso de extracción...")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Comenzando proceso de extracción...")
                if self.root:
                    self.root.update()
            
            # Verificación de si estamos en la página correcta
            in_issues_page = False
            
            # Estrategia 1: Buscar el texto "Issues (número)"
            try:
                issues_title_elements = self.driver.find_elements(
                    By.XPATH, 
                    "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                )
                if issues_title_elements:
                    logger.info(f"Página de Issues detectada por título: {issues_title_elements[0].text}")
                    in_issues_page = True
            except Exception as e:
                logger.debug(f"No se pudo detectar título de Issues: {e}")
            
            # Estrategia 2: Verificar si hay filas de datos visibles
            if not in_issues_page:
                issue_rows = self.browser.find_table_rows(highlight=False)
                if len(issue_rows) > 0:
                    logger.info(f"Se detectaron {len(issue_rows)} filas de datos que parecen issues")
                    in_issues_page = True
            
            # Estrategia 3: Verificar encabezados de columna típicos
            if not in_issues_page:
                try:
                    column_headers = self.driver.find_elements(
                        By.XPATH,
                        "//div[text()='Title'] | //div[text()='Type'] | //div[text()='Priority'] | //div[text()='Status']"
                    )
                    if len(column_headers) >= 3:
                        logger.info(f"Se detectaron encabezados de columna típicos de issues: {len(column_headers)}")
                        in_issues_page = True
                except Exception as e:
                    logger.debug(f"No se pudieron detectar encabezados de columna: {e}")
            
            # Si aún no estamos seguros, intentar hacer clic en la pestaña Issues
            if not in_issues_page:
                logger.warning("No se detectó la página de Issues. Intentando hacer clic en la pestaña...")
                
                tab_selectors = [
                    "//div[@role='tab' and contains(text(), 'Issues')]",
                    "//a[contains(text(), 'Issues')]",
                    "//li[contains(@class, 'tab') and contains(., 'Issues')]",
                    "//div[contains(@class, 'sapMITBItem') and contains(., 'Issues')]",
                    "//div[contains(@class, 'sapMITBItem')]//span[contains(text(), 'Issues')]/..",
                    "//*[contains(text(), 'Issues') and not(contains(text(), '('))]"
                ]
                
                issue_tab_found = False
                for selector in tab_selectors:
                    try:
                        issue_tabs = self.driver.find_elements(By.XPATH, selector)
                        if issue_tabs:
                            for tab in issue_tabs:
                                try:
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                                    time.sleep(1)
                                    
                                    try:
                                        self.driver.execute_script("arguments[0].click();", tab)
                                        logger.info(f"Clic en pestaña Issues realizado con JavaScript: {tab.text}")
                                        time.sleep(3)
                                        issue_tab_found = True
                                        break
                                    except:
                                        tab.click()
                                        logger.info(f"Clic en pestaña Issues realizado: {tab.text}")
                                        time.sleep(3)
                                        issue_tab_found = True
                                        break
                                except Exception as click_e:
                                    logger.debug(f"Error al hacer clic en pestaña: {click_e}")
                                    continue
                            
                            if issue_tab_found:
                                break
                    except Exception as e:
                        logger.debug(f"Error con selector {selector}: {e}")
                
                if issue_tab_found:
                    try:
                        issues_title_elements = self.driver.find_elements(
                            By.XPATH, "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                        )
                        if issues_title_elements:
                            logger.info(f"Página de Issues detectada después de clic: {issues_title_elements[0].text}")
                            in_issues_page = True
                    except:
                        pass
            
            if not in_issues_page:
                logger.warning("No se pudo confirmar que estamos en la página de Issues, pero intentaremos extraer datos de todos modos.")
            
            # Intentar extracción con reintentos
            max_attempts = 3
            for attempt in range(max_attempts):
                try:
                    logger.info(f"Intento de extracción {attempt+1}/{max_attempts}")
                    
                    # Actualizar la interfaz si existe
                    if hasattr(self, 'status_var') and self.status_var:
                        self.status_var.set(f"Intento de extracción {attempt+1}/{max_attempts}...")
                    
                    issues_data = self.browser.extract_issues_data()
                    
                    if issues_data:
                        logger.info(f"Extracción exitosa: {len(issues_data)} issues encontrados")
                        
                        # Actualizar Excel con los datos extraídos
                        if self.excel_file_path:
                            self.update_excel(issues_data)
                        else:
                            logger.warning("No se ha seleccionado archivo Excel para guardar los datos")
                            
                            if hasattr(self, 'status_var') and self.status_var:
                                self.status_var.set("Advertencia: No se ha seleccionado archivo Excel")
                            
                            if self.root:
                                excel_path = self.choose_excel_file()
                                if excel_path:
                                    self.update_excel(issues_data)
                        
                        self.processing = False
                        return True
                    else:
                        logger.warning(f"No se encontraron issues en el intento {attempt+1}")
                        
                        # Si no es el último intento, esperar y reintentar
                        if attempt < max_attempts - 1:
                            logger.info("Esperando antes de reintentar...")
                            time.sleep(5)
                        else:
                            logger.error("Todos los intentos de extracción fallaron")
                            
                            if hasattr(self, 'status_var') and self.status_var:
                                self.status_var.set("Error: No se encontraron issues después de varios intentos")
                            
                            if self.root:
                                messagebox.showerror(
                                    "Error de Extracción", 
                                    "No se pudieron encontrar issues después de varios intentos. Verifique que está en la página correcta y que existen issues para extraer."
                                )
                            
                            self.processing = False
                            return False
                except Exception as e:
                    logger.error(f"Error en el intento {attempt+1}: {e}")
                    
                    # Si no es el último intento, esperar y reintentar
                    if attempt < max_attempts - 1:
                        logger.info("Esperando antes de reintentar...")
                        time.sleep(5)
                    else:
                        logger.error(f"Todos los intentos de extracción fallaron: {e}")
                        
                        if hasattr(self, 'status_var') and self.status_var:
                            self.status_var.set(f"Error de extracción: {e}")
                        
                        if self.root:
                            messagebox.showerror(
                                "Error de Extracción", 
                                f"Se produjo un error durante la extracción: {e}"
                            )
                        
                        self.processing = False
                        return False
            
            # Si llegamos aquí, todos los intentos fallaron
            logger.error("Extracción fallida después de varios intentos")
            
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Error: Extracción fallida")
            
            self.processing = False
            return False
            
        except Exception as e:
            logger.error(f"Error general en el proceso de extracción: {e}")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error: {e}")
            
            self.processing = False
            return False









    def create_gui(self):
        """Crea una interfaz gráfica mejorada para la aplicación"""
        self.root = tk.Tk()
        self.root.title("SAP Recommendations Extractor")
        self.root.geometry("650x800")
        self.root.resizable(True, True)
        self.root.configure(bg=SAP_COLORS["light"])
        
        # Ajustar el tema según el sistema operativo para mejor integración
        self._adjust_theme_for_platform()
        
        # Crear componentes en métodos separados para mejor organización
        self._create_header_panel()
        self._create_client_panel()
        self._create_project_panel()
        self._create_browser_panel()
        self._create_excel_panel()
        self._create_action_panel()
        self._create_log_panel()
        
        # Configurar barra de estado
        self._create_status_bar()
        
        # Configurar eventos de teclado para accesos rápidos
        self._setup_keyboard_shortcuts()
        
        # Configurar logger GUI y carga de configuración al final
        self.setup_gui_logger()
        self.load_config()
        
        # Centrar ventana
        self._center_window()
        
        
        # Configurar icono de la aplicación
        logo_photo = None
        if 'PIL_AVAILABLE' in globals() and PIL_AVAILABLE:
            try:
                # Logo SAP como icono base64
                sap_logo_base64 = """
                iVBORw0KGgoAAAANSUhEUgAAADAAAAAwCAYAAABXAvmHAAAACXBIWXMAAAsTAAALEwEAmpwYAAAB
                g0lEQVR4nO2ZQU7DMBBFv1mcJbAhB4ADcIucgAsAYslpWKQSKnCHZoPEgnuAaFasqESrKE3qicdJ
                +0kjVXXi+T8ex3ZSoCiKotTDEMAUwBxAUrJmXjsCaJR94IeSnbXCVgvAPYAvQ/LWWnjtDUDXdjJr
                A5jvwFTZoWnVBrAImFhZFgBOygxx6VkOytzIYRpZ5jEGOA0sbwy70NTGQHcHkltfN7IxUBTYKPPY
                3kKriMmN8gxMYqLjPY6JdrjEk/SaiGqigYGixGSiYUNI3bQnZZ8npiamCiCDiZaeBbQ/LWBMfAJ4
                5LU7AF9oo0yyiW/eN36QxQvXL4t5YlkxDZwBuAbwxsOI18a89sXrFwDOY05i7J3iCUCnYK6T7x1j
                9A0MPPoGlhLb1jHPdLgmXw7uOU++tg88Lrn+w/UrwvVKcjuxLaEYA5L1VmcQs4UoO9TGLDPJeiuZ
                qDKAxNTaQMwea227kDQT
                """
                
                # Cargar icono desde base64
                logo_data = base64.b64decode(sap_logo_base64)
                logo_image = Image.open(BytesIO(logo_data))
                
                # Convertir a formato para tkinter
                logo_photo = ImageTk.PhotoImage(logo_image)
                
                # Establecer como icono de la ventana
                self.root.iconphoto(True, logo_photo)
            except Exception as e:
                logger.error(f"Error al cargar el icono: {e}")
            
        # Configurar estilo
        style = ttk.Style()
        style.configure('TCombobox', arrowsize=15)
        
        # Estilos para widgets
        style.configure(".", foreground=SAP_COLORS["text"])
        style.configure("TLabel", background=SAP_COLORS["light"], foreground=SAP_COLORS["text"], font=("Arial", 10, "bold"))
        style.configure("Header.TLabel", background=SAP_COLORS["light"], foreground=SAP_COLORS["secondary"], font=("Arial", 16, "bold"))
        style.configure("TLabelframe.Label", background=SAP_COLORS["light"], foreground=SAP_COLORS["text"], font=("Arial", 11, "bold"))
        style.configure("Primary.TButton", background=SAP_COLORS["primary"], foreground=SAP_COLORS["white"], font=("Arial", 10, "bold"))
        style.configure("Success.TButton", background=SAP_COLORS["success"], foreground=SAP_COLORS["white"], font=("Arial", 10, "bold"))
        style.configure("Danger.TButton", background=SAP_COLORS["danger"], foreground=SAP_COLORS["white"], font=("Arial", 10, "bold"))
        style.configure("TCombobox", selectbackground=SAP_COLORS["primary"], selectforeground=SAP_COLORS["white"], 
                        fieldbackground="white", background="white", foreground=SAP_COLORS["text"])
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Cabecera con logo
        self.header_frame = ttk.Frame(main_frame)
        self.header_frame.pack(fill=tk.X, pady=(0, 15))

        try:
            # Usar el logo si ya está cargado y PIL está disponible
            if 'PIL_AVAILABLE' in globals() and PIL_AVAILABLE and logo_photo:
                logo_label = tk.Label(self.header_frame, image=logo_photo, bg=SAP_COLORS["light"])
                logo_label.image = logo_photo  # Mantener referencia
                logo_label.pack(side=tk.LEFT, padx=(0, 10))
        except:
            pass
        
        # Título con fondo de alto contraste
        title_background = "#0A3D6E"  # Azul oscuro
        title_foreground = "#FFFFFF"  # Texto blanco
        
        title_label = tk.Label(
            self.header_frame, 
            text="Extractor de Recomendaciones SAP",
            font=("Arial", 18, "bold"),
            foreground=title_foreground,
            background=title_background,
            padx=8,
            pady=4
        )
        title_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Contenedor principal dividido
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Panel izquierdo para configuración
        self.left_panel = ttk.Frame(content_frame, padding=10, width=435)
        self.left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        self.left_panel.pack_propagate(False)  # Mantener ancho fijo
        
        # Sección de cliente
        client_frame = tk.LabelFrame(self.left_panel, 
                                text="Cliente", 
                                bg=SAP_COLORS["light"],
                                fg="#000000",
                                font=("Arial", 11, "bold"),
                                padx=10, pady=10)
        client_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta ERP
        tk.Label(client_frame, 
            text="ERP Number:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 9)).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Entry ERP
        self.client_var = tk.StringVar(value="1025541")
        client_entry = tk.Entry(client_frame, 
                            textvariable=self.client_var,
                            width=15,
                            font=("Arial", 10),
                            bg="white",
                            fg="black",
                            highlightbackground=SAP_COLORS["primary"],
                            highlightcolor=SAP_COLORS["primary"])
        client_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        # Etiqueta de clientes guardados
        tk.Label(client_frame, 
            text="Clientes guardados:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 9)).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Lista desplegable de clientes guardados
        client_list = self.db_manager.get_clients()
        self.client_combo = ttk.Combobox(client_frame, values=client_list, width=30)
        self.client_combo.config(state='readonly')
        self.client_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W+tk.E)
        self.client_combo.bind("<<ComboboxSelected>>", lambda e: self.select_client(self.client_combo.get()))

        # Sección de proyecto
        project_frame = tk.LabelFrame(self.left_panel, 
                                    text="Proyecto", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
        project_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta ID
        tk.Label(project_frame, 
            text="ID Proyecto:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10)).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        










# Entry Proyecto
        # Para la inicialización de los Entry
        self.client_var = tk.StringVar()
        self.client_var.set("1025541")  # Valor por defecto
        project_entry = tk.Entry(project_frame, 
                            textvariable=self.project_var,
                            width=15,
                            font=("Arial", 10),
                            bg="white",
                            fg="black",
                            highlightbackground=SAP_COLORS["primary"],
                            highlightcolor=SAP_COLORS["primary"])
        project_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        # Etiqueta de proyectos
        tk.Label(project_frame, 
            text="Proyectos:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10)).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Lista desplegable de proyectos guardados
        project_list = self.db_manager.get_projects("1025541")  # Proyectos para el cliente predeterminado
        self.project_combo = ttk.Combobox(project_frame, values=project_list, width=30)
        self.project_combo.config(state='readonly')
        self.project_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W+tk.E)
        self.project_combo.bind("<<ComboboxSelected>>", lambda e: self.select_project(self.project_combo.get()))

        # Sección de navegador
        browser_frame = tk.LabelFrame(self.left_panel, 
                                    text="Navegador", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
        browser_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta de navegador
        browser_label = tk.Label(
            browser_frame, 
            text="Iniciar un navegador con perfil dedicado:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        browser_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón de navegador
        browser_button = tk.Button(
            browser_frame, 
            text="Iniciar Navegador",
            command=self.start_browser,
            bg=SAP_COLORS["primary"],
            fg="#FFFFFF",
            activebackground="#0A3D6E",
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        browser_button.pack(fill=tk.X, pady=5)
        
        # Sección de archivo Excel
        excel_frame = tk.LabelFrame(self.left_panel, 
                                text="Archivo Excel", 
                                bg=SAP_COLORS["light"],
                                fg="#000000",
                                font=("Arial", 11, "bold"),
                                padx=10, pady=10)
        excel_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta de Excel
        excel_label = tk.Label(
            excel_frame, 
            text="Seleccione un archivo existente o cree uno nuevo:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        excel_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón de Excel
        excel_button = tk.Button(
            excel_frame, 
            text="Seleccionar o Crear Excel",
            command=self.choose_excel_file,
            bg=SAP_COLORS["success"],
            fg="#FFFFFF",
            activebackground="#085E2E",
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        excel_button.pack(fill=tk.X, pady=5)
        
        # Mostrar el nombre del archivo seleccionado
        self.excel_filename_var = tk.StringVar(value="No seleccionado")
        excel_file_label = tk.Label(
            excel_frame, 
            textvariable=self.excel_filename_var,
            bg=SAP_COLORS["light"],
            fg="#0A3D6E",
            font=("Arial", 9, "bold"),
            wraplength=200,
            anchor="w",
            justify="left"
        )
        excel_file_label.pack(fill=tk.X, pady=5)
        
        # Sección de acción
        action_frame = tk.LabelFrame(self.left_panel, 
                                text="Acciones", 
                                bg=SAP_COLORS["light"],
                                fg="#000000",
                                font=("Arial", 11, "bold"),
                                padx=10, pady=10)
        action_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta de acción
        action_label = tk.Label(
            action_frame, 
            text="Extraer datos de issues desde SAP:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        action_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón de extracción
        extract_button = tk.Button(
            action_frame, 
            text="Iniciar Extracción de Issues",
            command=self.start_extraction,
            bg=SAP_COLORS["warning"],
            fg="#FFFFFF",
            activebackground="#C25A00",
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        extract_button.pack(fill=tk.X, pady=5)
        
        # Separador visual
        separator = tk.Frame(action_frame, height=2, bg=SAP_COLORS["gray"])
        separator.pack(fill=tk.X, pady=10)
        
        # Botón de salir
        exit_button = tk.Button(
            action_frame, 
            text="Salir de la Aplicación",
            command=self.exit_app,
            bg=SAP_COLORS["danger"],
            fg="#FFFFFF",
            activebackground="#990000",
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        exit_button.pack(fill=tk.X, pady=5)
        
        # Panel derecho para logs
        right_panel = ttk.Frame(content_frame, padding=10, width=300)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        right_panel.pack_propagate(False)
        
        # Log frame
        log_frame = tk.LabelFrame(right_panel, 
                            text="Registro de Actividad", 
                            bg=SAP_COLORS["light"],
                            fg="#000000",
                            font=("Arial", 11, "bold"))
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        # Text widget para logs
        self.log_text = tk.Text(
            log_frame, 
            height=20, 
            wrap=tk.WORD, 
            bg="white",
            fg="black",
            font=("Consolas", 9),
            padx=5,
            pady=5,
            borderwidth=2,
            relief=tk.SUNKEN
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Colores para los logs
        self.log_text.tag_configure("INFO", foreground="black")
        self.log_text.tag_configure("WARNING", foreground="#CC6600")
        self.log_text.tag_configure("ERROR", foreground="#990000")
        self.log_text.tag_configure("DEBUG", foreground="#555555")
        
        # Scrollbar para el log
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Status bar
        self.status_var = tk.StringVar(value="Listo para iniciar")
        status_bar = tk.Label(
            self.root, 
            textvariable=self.status_var,
            fg="#000000",
            bg="#F0F0F0",
            relief=tk.SUNKEN, 
            anchor=tk.W, 
            padx=5,
            pady=2,
            font=("Arial", 10)
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Configurar logger para que también escriba en la GUI
        self.setup_gui_logger()
        
        # Manejar cierre de ventana
        self.root.protocol("WM_DELETE_WINDOW", self.exit_app)
        
        # Centrar la ventana en la pantalla
        self.root.update_idletasks()
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        size = tuple(int(_) for _ in self.root.geometry().split('+')[0].split('x'))
        x = screen_width/2 - size[0]/2
        y = screen_height/2 - size[1]/2
        self.root.geometry("%dx%d+%d+%d" % (size + (x, y)))
        
        # Cargar configuración guardada
        self.load_config()
        












    def _adjust_theme_for_platform(self):
        """
        Ajusta el tema de la interfaz según el sistema operativo para mejor integración.
        
        Detecta el sistema operativo y aplica ajustes específicos para mejorar
        la apariencia visual según la plataforma.
        """
        try:
            import platform
            system = platform.system()
            
            # Configurar tema dependiendo del sistema operativo
            if system == "Windows":
                # Usar colores nativos de Windows para algunos elementos
                self.root.configure(bg="#F0F0F0")
            elif system == "Darwin":  # macOS
                # Ajustes específicos para macOS
                self.root.configure(bg="#ECECEC")
            elif system == "Linux":
                # Ajustes específicos para Linux
                pass
                
            logger.debug(f"Tema ajustado para plataforma: {system}")
        except Exception as e:
            logger.warning(f"No se pudo ajustar el tema para la plataforma: {e}")

    def _create_header_panel(self):
        """
        Crea el panel de cabecera con el logo y título de la aplicación.
        
        Este panel proporciona la identidad visual principal de la aplicación
        con el logo de SAP y un título destacado.
        """
        try:
            # Frame de cabecera
            self.header_frame = ttk.Frame(self.root)
            self.header_frame.pack(fill=tk.X, pady=(10, 15))
            
            # Intentar cargar logo
            logo_photo = None
            if PIL_AVAILABLE:
                try:
                    # Logo SAP como icono base64
                    sap_logo_base64 = """
                    iVBORw0KGgoAAAANSUhEUgAAADAAAAAwCAYAAABXAvmHAAAACXBIWXMAAAsTAAALEwEAmpwYAAAB
                    g0lEQVR4nO2ZQU7DMBBFv1mcJbAhB4ADcIucgAsAYslpWKQSKnCHZoPEgnuAaFasqESrKE3qicdJ
                    +0kjVXXi+T8ex3ZSoCiKotTDEMAUwBxAUrJmXjsCaJR94IeSnbXCVgvAPYAvQ/LWWnjtDUDXdjJr
                    A5jvwFTZoWnVBrAImFhZFgBOygxx6VkOytzIYRpZ5jEGOA0sbwy70NTGQHcHkltfN7IxUBTYKPPY
                    3kKriMmN8gxMYqLjPY6JdrjEk/SaiGqigYGixGSiYUNI3bQnZZ8npiamCiCDiZaeBbQ/LWBMfAJ4
                    5LU7AF9oo0yyiW/eN36QxQvXL4t5YlkxDZwBuAbwxsOI18a89sXrFwDOY05i7J3iCUCnYK6T7x1j
                    9A0MPPoGlhLb1jHPdLgmXw7uOU++tg88Lrn+w/UrwvVKcjuxLaEYA5L1VmcQs4UoO9TGLDPJeiuZ
                    qDKAxNTaQMwea227kDQT
                    """
                    
                    # Cargar icono desde base64
                    logo_data = base64.b64decode(sap_logo_base64)
                    logo_image = Image.open(BytesIO(logo_data))
                    
                    # Convertir a formato para tkinter
                    logo_photo = ImageTk.PhotoImage(logo_image)
                    
                    # Añadir logo a la cabecera
                    logo_label = tk.Label(self.header_frame, image=logo_photo, bg=SAP_COLORS["light"])
                    logo_label.image = logo_photo  # Mantener referencia
                    logo_label.pack(side=tk.LEFT, padx=(0, 10))
                except Exception as e:
                    logger.debug(f"No se pudo cargar el logo: {e}")
            
            # Título con fondo de alto contraste
            title_background = "#0A3D6E"  # Azul oscuro
            title_foreground = "#FFFFFF"  # Texto blanco
            
            title_label = tk.Label(
                self.header_frame, 
                text="Extractor de Recomendaciones SAP",
                font=("Arial", 18, "bold"),
                foreground=title_foreground,
                background=title_background,
                padx=8,
                pady=4
            )
            title_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        except Exception as e:
            logger.error(f"Error al crear panel de cabecera: {e}")

    def _create_client_panel(self):
        """
        Crea el panel de selección de cliente.
        
        Este panel permite al usuario seleccionar o introducir el número ERP
        del cliente para el que se extraerán los datos.
        """
        try:
            # Panel izquierdo para configuración
            if not hasattr(self, 'left_panel'):
                self.left_panel = ttk.Frame(self.root, padding=10, width=435)
                self.left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 10))
                self.left_panel.pack_propagate(False)  # Mantener ancho fijo
                
            # Sección de cliente
            client_frame = tk.LabelFrame(self.left_panel, 
                                    text="Cliente", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
            client_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Etiqueta ERP
            tk.Label(client_frame, 
                text="ERP Number:",
                bg=SAP_COLORS["light"],
                fg="#000000",
                font=("Arial", 9)).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            
            # Entry ERP
            self.client_var = tk.StringVar(value="1025541")
            client_entry = tk.Entry(client_frame, 
                                textvariable=self.client_var,
                                width=15,
                                font=("Arial", 10),
                                bg="white",
                                fg="black",
                                highlightbackground=SAP_COLORS["primary"],
                                highlightcolor=SAP_COLORS["primary"])
            client_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            # Etiqueta de clientes guardados
            tk.Label(client_frame, 
                text="Clientes guardados:",
                bg=SAP_COLORS["light"],
                fg="#000000",
                font=("Arial", 9)).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            
            # Lista desplegable de clientes guardados
            client_list = self.db_manager.get_clients()
            self.client_combo = ttk.Combobox(client_frame, values=client_list, width=30)
            self.client_combo.config(state='readonly')
            self.client_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W+tk.E)
            self.client_combo.bind("<<ComboboxSelected>>", lambda e: self.select_client(self.client_combo.get()))
        except Exception as e:
            logger.error(f"Error al crear panel de cliente: {e}")

    def _create_project_panel(self):
        """
        Crea el panel de selección de proyecto.
        
        Este panel permite al usuario seleccionar o introducir el ID del
        proyecto específico para extracción de datos.
        """
        try:
            # Sección de proyecto
            project_frame = tk.LabelFrame(self.left_panel, 
                                        text="Proyecto", 
                                        bg=SAP_COLORS["light"],
                                        fg="#000000",
                                        font=("Arial", 11, "bold"),
                                        padx=10, pady=10)
            project_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Etiqueta ID
            tk.Label(project_frame, 
                text="ID Proyecto:",
                bg=SAP_COLORS["light"],
                fg="#000000",
                font=("Arial", 10)).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
            
            # Entry Proyecto
            self.project_var = tk.StringVar(value="20096444")
            project_entry = tk.Entry(project_frame, 
                                textvariable=self.project_var,
                                width=15,
                                font=("Arial", 10),
                                bg="white",
                                fg="black",
                                highlightbackground=SAP_COLORS["primary"],
                                highlightcolor=SAP_COLORS["primary"])
            project_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
            
            # Etiqueta de proyectos
            tk.Label(project_frame, 
                text="Proyectos:",
                bg=SAP_COLORS["light"],
                fg="#000000",
                font=("Arial", 10)).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
            
            # Lista desplegable de proyectos guardados
            project_list = self.db_manager.get_projects("1025541")  # Proyectos para el cliente predeterminado
            self.project_combo = ttk.Combobox(project_frame, values=project_list, width=30)
            self.project_combo.config(state='readonly')
            self.project_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W+tk.E)
            self.project_combo.bind("<<ComboboxSelected>>", lambda e: self.select_project(self.project_combo.get()))
        except Exception as e:
            logger.error(f"Error al crear panel de proyecto: {e}")

    def _create_browser_panel(self):
        """
        Crea el panel de control del navegador.
        
        Este panel proporciona opciones para iniciar y controlar la sesión
        de navegador utilizada para la extracción.
        """
        try:
            # Sección de navegador
            browser_frame = tk.LabelFrame(self.left_panel, 
                                        text="Navegador", 
                                        bg=SAP_COLORS["light"],
                                        fg="#000000",
                                        font=("Arial", 11, "bold"),
                                        padx=10, pady=10)
            browser_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Etiqueta de navegador
            browser_label = tk.Label(
                browser_frame, 
                text="Iniciar un navegador con perfil dedicado:",
                bg=SAP_COLORS["light"],
                fg="#000000",
                font=("Arial", 10),
                anchor="w",
                justify="left"
            )
            browser_label.pack(fill=tk.X, pady=(0, 5))
            
            # Botón de navegador
            browser_button = tk.Button(
                browser_frame, 
                text="Iniciar Navegador",
                command=self.start_browser,
                bg=SAP_COLORS["primary"],
                fg="#FFFFFF",
                activebackground="#0A3D6E",
                activeforeground="#FFFFFF",
                font=("Arial", 10, "bold"),
                padx=10, pady=5
            )
            browser_button.pack(fill=tk.X, pady=5)
        except Exception as e:
            logger.error(f"Error al crear panel de navegador: {e}")

    def _create_excel_panel(self):
        """
        Crea el panel de selección y gestión de archivo Excel.
        
        Este panel permite al usuario seleccionar un archivo Excel existente
        o crear uno nuevo para almacenar los datos extraídos.
        """
        try:
            # Sección de archivo Excel
            excel_frame = tk.LabelFrame(self.left_panel, 
                                    text="Archivo Excel", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
            excel_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Etiqueta de Excel
            excel_label = tk.Label(
                excel_frame, 
                text="Seleccione un archivo existente o cree uno nuevo:",
                bg=SAP_COLORS["light"],
                fg="#000000",
                font=("Arial", 10),
                anchor="w",
                justify="left"
            )
            excel_label.pack(fill=tk.X, pady=(0, 5))
            
            # Botón de Excel
            excel_button = tk.Button(
                excel_frame, 
                text="Seleccionar o Crear Excel",
                command=self.choose_excel_file,
                bg=SAP_COLORS["success"],
                fg="#FFFFFF",
                activebackground="#085E2E",
                activeforeground="#FFFFFF",
                font=("Arial", 10, "bold"),
                padx=10, pady=5
            )
            excel_button.pack(fill=tk.X, pady=5)
            
            # Mostrar el nombre del archivo seleccionado
            self.excel_filename_var = tk.StringVar(value="No seleccionado")
            excel_file_label = tk.Label(
                excel_frame, 
                textvariable=self.excel_filename_var,
                bg=SAP_COLORS["light"],
                fg="#0A3D6E",
                font=("Arial", 9, "bold"),
                wraplength=200,
                anchor="w",
                justify="left"
            )
            excel_file_label.pack(fill=tk.X, pady=5)
        except Exception as e:
            logger.error(f"Error al crear panel de Excel: {e}")

    def _create_action_panel(self):
        """
        Crea el panel de acciones principales de la aplicación.
        
        Este panel contiene los botones de acción principal como iniciar
        la extracción y salir de la aplicación.
        """
        try:
            # Sección de acción
            action_frame = tk.LabelFrame(self.left_panel, 
                                    text="Acciones", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
            action_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Etiqueta de acción
            action_label = tk.Label(
                action_frame, 
                text="Extraer datos de issues desde SAP:",
                bg=SAP_COLORS["light"],
                fg="#000000",
                font=("Arial", 10),
                anchor="w",
                justify="left"
            )
            action_label.pack(fill=tk.X, pady=(0, 5))
            
            # Botón de extracción
            extract_button = tk.Button(
                action_frame, 
                text="Iniciar Extracción de Issues",
                command=self.start_extraction,
                bg=SAP_COLORS["warning"],
                fg="#FFFFFF",
                activebackground="#C25A00",
                activeforeground="#FFFFFF",
                font=("Arial", 10, "bold"),
                padx=10, pady=5
            )
            extract_button.pack(fill=tk.X, pady=5)
            
            # Separador visual
            separator = tk.Frame(action_frame, height=2, bg=SAP_COLORS["gray"])
            separator.pack(fill=tk.X, pady=10)
            
            # Botón de salir
            exit_button = tk.Button(
                action_frame, 
                text="Salir de la Aplicación",
                command=self.exit_app,
                bg=SAP_COLORS["danger"],
                fg="#FFFFFF",
                activebackground="#990000",
                activeforeground="#FFFFFF",
                font=("Arial", 10, "bold"),
                padx=10, pady=5
            )
            exit_button.pack(fill=tk.X, pady=5)
        except Exception as e:
            logger.error(f"Error al crear panel de acciones: {e}")

    def _create_log_panel(self):
        """
        Crea el panel de registro de actividad (log).
        
        Este panel muestra un registro detallado de todas las operaciones
        y eventos que ocurren durante la ejecución del programa.
        """
        try:
            # Panel derecho para logs
            right_panel = ttk.Frame(self.root, padding=10, width=300)
            right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
            right_panel.pack_propagate(False)
            
            # Log frame
            log_frame = tk.LabelFrame(right_panel, 
                                text="Registro de Actividad", 
                                bg=SAP_COLORS["light"],
                                fg="#000000",
                                font=("Arial", 11, "bold"))
            log_frame.pack(fill=tk.BOTH, expand=True)
            
            # Text widget para logs
            self.log_text = tk.Text(
                log_frame, 
                height=20, 
                wrap=tk.WORD, 
                bg="white",
                fg="black",
                font=("Consolas", 9),
                padx=5,
                pady=5,
                borderwidth=2,
                relief=tk.SUNKEN
            )
            self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            # Colores para los logs
            self.log_text.tag_configure("INFO", foreground="black")
            self.log_text.tag_configure("WARNING", foreground="#CC6600")
            self.log_text.tag_configure("ERROR", foreground="#990000")
            self.log_text.tag_configure("DEBUG", foreground="#555555")
            
            # Scrollbar para el log
            scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.log_text.config(yscrollcommand=scrollbar.set)
        except Exception as e:
            logger.error(f"Error al crear panel de logs: {e}")

    def _create_status_bar(self):
        """
        Crea la barra de estado en la parte inferior de la ventana.
        
        Esta barra muestra información sobre el estado actual del proceso
        y proporciona retroalimentación visual al usuario.
        """
        try:
            # Status bar
            self.status_var = tk.StringVar(value="Listo para iniciar")
            status_bar = tk.Label(
                self.root, 
                textvariable=self.status_var,
                fg="#000000",
                bg="#F0F0F0",
                relief=tk.SUNKEN, 
                anchor=tk.W, 
                padx=5,
                pady=2,
                font=("Arial", 10)
            )
            status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        except Exception as e:
            logger.error(f"Error al crear barra de estado: {e}")

    def _setup_keyboard_shortcuts(self):
        """
        Configura atajos de teclado para las funciones principales.
        
        Establece combinaciones de teclas para acceder rápidamente a las
        funcionalidades más utilizadas de la aplicación.
        """
        try:
            # Definir atajos de teclado
            self.root.bind("<Control-q>", lambda e: self.exit_app())
            self.root.bind("<Control-b>", lambda e: self.start_browser())
            self.root.bind("<Control-e>", lambda e: self.choose_excel_file())
            self.root.bind("<F5>", lambda e: self.start_extraction())
            
            logger.debug("Atajos de teclado configurados")
        except Exception as e:
            logger.warning(f"No se pudieron configurar los atajos de teclado: {e}")

    def _center_window(self):
        """
        Centra la ventana de la aplicación en la pantalla.
        
        Calcula la posición adecuada para que la ventana aparezca
        centrada en la pantalla del usuario.
        """
        try:
            # Actualizar la información de geometría
            self.root.update_idletasks()
            
            # Obtener dimensiones de la pantalla
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            # Obtener dimensiones de la ventana
            window_width = self.root.winfo_width()
            window_height = self.root.winfo_height()
            
            # Calcular posición para centrar
            x = int((screen_width - window_width) / 2)
            y = int((screen_height - window_height) / 2)
            
            # Establecer geometría
            self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
            
            logger.debug(f"Ventana centrada en: {x},{y}")
        except Exception as e:
            logger.warning(f"No se pudo centrar la ventana: {e}")




    def setup_gui_logger(self):
        """Configura el logger para que también escriba en la GUI"""
        class TextHandler(logging.Handler):
            def __init__(self, text_widget):
                logging.Handler.__init__(self)
                self.text_widget = text_widget
            
            def emit(self, record):
                msg = self.format(record)
                def append():
                    self.text_widget.configure(state='normal')
                    
                    # Agregar marca de tiempo y nivel con color
                    time_str = msg.split(' - ')[0] + ' - '
                    level_str = record.levelname + ' - '
                    msg_content = msg.split(' - ', 2)[2] if len(msg.split(' - ')) > 2 else ""
                    
                    self.text_widget.insert(tk.END, time_str, "INFO")
                    self.text_widget.insert(tk.END, level_str, record.levelname)
                    self.text_widget.insert(tk.END, msg_content + '\n', record.levelname)
                    
                    self.text_widget.configure(state='disabled')
                    self.text_widget.yview(tk.END)
                    
                    # Limitar tamaño del log
                    self.limit_log_length()
                    
                # Llamar a append desde el hilo principal
                self.text_widget.after(0, append)
                
            def limit_log_length(self):
                """Limita la longitud del log para evitar consumo excesivo de memoria"""
                if float(self.text_widget.index('end-1c').split('.')[0]) > 1000:
                    self.text_widget.configure(state='normal')
                    self.text_widget.delete('1.0', '500.0')
                    self.text_widget.configure(state='disabled')
        
        # Crear handler para el widget Text
        text_handler = TextHandler(self.log_text)
        text_handler.setLevel(logging.INFO)
        text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        
        # Añadir el handler al logger
        logger.addHandler(text_handler)
        
        # Deshabilitar el widget
        self.log_text.configure(state='disabled')
    






    def select_client(self, client_string):
        """Maneja la selección de un cliente desde el combobox"""
        try:
            if not client_string:
                return
                
            # Extraer el ERP number del string "1025541 - Nombre del cliente"
            erp_number = client_string.split(" - ")[0].strip()
            
            # Establecer el valor en el Entry
            self.client_var.set(erp_number)
            
            # Actualizar la lista de proyectos para este cliente
            projects = self.db_manager.get_projects(erp_number)
            self.project_combo['values'] = projects
            
            # Si hay proyectos disponibles, seleccionar el primero
            if projects:
                self.project_combo.current(0)
                self.select_project(projects[0])
                    
            # Actualizar el uso de este cliente
            self.db_manager.update_client_usage(erp_number)
            
            logger.info(f"Cliente seleccionado: {erp_number}")
            self.save_config()
        except Exception as e:
            logger.error(f"Error al seleccionar cliente: {e}")    
            
            
            
            
            
            
            
            
            
            
            
            
    def select_project(self, project_string):
        """Maneja la selección de un proyecto desde el combobox"""
        try:
            if not project_string:
                return
                
            # Extraer el ID del proyecto del string "20096444 - Nombre del proyecto"
            project_id = project_string.split(" - ")[0].strip()
            
            # Establecer el valor en el Entry
            self.project_var.set(project_id)
            
            # Actualizar el uso de este proyecto
            self.db_manager.update_project_usage(project_id)
            
            logger.info(f"Proyecto seleccionado: {project_id}")
            self.save_config()
        except Exception as e:
            logger.error(f"Error al seleccionar proyecto: {e}")
    
    def start_browser(self):
        """Inicia el navegador desde la interfaz gráfica"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                messagebox.showwarning("Proceso en curso", "Hay un proceso de extracción en curso.")
                return
                
            # Asegurarse de que no haya un navegador ya abierto
            if self.driver:
                messagebox.showinfo("Navegador ya iniciado", "El navegador ya está iniciado.")
                return
                
            # Actualizar la interfaz para mostrar que se está iniciando el navegador
            self.status_var.set("Iniciando navegador...")
            if self.root:
                self.root.update()
            
            # Iniciar el navegador en un hilo separado
            threading.Thread(target=self._start_browser_thread, daemon=True).start()
            
        except Exception as e:
            logger.error(f"Error al iniciar el navegador: {e}")
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", f"Error al iniciar el navegador: {e}")
        
    def _start_browser_thread(self):
        """Método para ejecutar la inicialización del navegador en un hilo separado"""
        try:
            if self.connect_to_browser():
                logger.info("Navegador iniciado")
                
                # Actualizar la interfaz en el hilo principal
                if self.root:
                    self.root.after(0, lambda: self.status_var.set("Navegador iniciado. Navegando a SAP..."))
                
                # Obtener valores de cliente y proyecto
                erp_number = self.client_var.get() if hasattr(self, 'client_var') and self.client_var else "1025541"
                project_id = self.project_var.get() if hasattr(self, 'project_var') and self.project_var else "20096444"
                
                # Navegar a la URL de SAP con parámetros específicos
                self.browser.navigate_to_sap(erp_number, project_id)
                
                # Mostrar instrucciones en el hilo principal
                if self.root:
                    self.root.after(0, lambda: self.status_var.set("Navegación completada. Inicie la extracción cuando esté listo."))
                    self.root.after(0, self._show_extraction_instructions)
            else:
                if self.root:
                    self.root.after(0, lambda: self.status_var.set("Error al iniciar el navegador"))
                    self.root.after(0, lambda: messagebox.showerror("Error", "No se pudo iniciar el navegador. Revise el log para más detalles."))
        except Exception as e:
            logger.error(f"Error en hilo de navegador: {e}")
            if self.root:
                self.root.after(0, lambda: self.status_var.set(f"Error: {e}"))
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error al iniciar el navegador: {e}"))
    









    def _show_extraction_instructions(self):
        """Muestra instrucciones para la extracción después de la navegación automática"""
        # Obtener valores actuales
        erp_number = self.client_var.get()
        project_id = self.project_var.get()
        
        instructions = f"""
        La aplicación ha navegado automáticamente a la página de SAP con:
        
        Cliente: {erp_number}
        Proyecto: {project_id}
        
        Por favor:
        1. Verifique que ha iniciado sesión correctamente
        2. Compruebe que se muestran los issues del proyecto
        3. Cuando esté listo, haga clic en 'Iniciar Extracción'
        """
        
        messagebox.showinfo("Instrucciones de Extracción", instructions)
    












    def start_extraction(self):
        """Inicia el proceso de extracción desde la interfaz gráfica"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                messagebox.showwarning("Proceso en curso", "Hay un proceso de extracción en curso.")
                return
                
            # Verificar que existe un archivo Excel seleccionado
            if not self.excel_file_path:
                messagebox.showwarning("Archivo Excel no seleccionado", "Debe seleccionar o crear un archivo Excel primero.")
                return
                
            # Verificar que el navegador está abierto
            if not self.driver:
                messagebox.showwarning("Navegador no iniciado", "Debe iniciar el navegador primero.")
                return
                
            # Obtener valores de cliente y proyecto
            erp_number = self.client_var.get()
            project_id = self.project_var.get()
            
            # Iniciar extracción en un hilo separado para no bloquear la GUI
            threading.Thread(
                target=self._fill_fields_and_extract, 
                args=(erp_number, project_id),
                daemon=True
            ).start()
            
        except Exception as e:
            logger.error(f"Error al iniciar extracción: {e}")
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", f"Error al iniciar extracción: {e}")            
    
    
    
    
    
    
    
    def _fill_fields_and_extract(self, erp_number, project_id):
        """Rellena los campos y luego ejecuta la extracción"""
        try:
            # Actualizar la interfaz
            if hasattr(self, 'status_var'):
                self.status_var.set("Seleccionando cliente...")
                if self.root:
                    self.root.update()
            
            # 1. Seleccionar cliente
            if not self.browser.select_customer_automatically(erp_number):
                logger.warning("No se pudo seleccionar cliente automáticamente")
                # Mostrar mensaje al usuario
                if self.root:
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Selección Manual Requerida", 
                        "No se pudo seleccionar el cliente automáticamente.\n\n"
                        "Por favor, seleccione manualmente el cliente."
                    ))
                    time.sleep(3)  # Dar tiempo para selección manual
            else:
                logger.info(f"Cliente {erp_number} seleccionado con éxito")
            
            # Actualizar la interfaz
            if hasattr(self, 'status_var'):
                self.status_var.set("Seleccionando proyecto...")
                if self.root:
                    self.root.update()
            
            # 2. Esperar un momento y luego seleccionar proyecto
            time.sleep(2)
            if not self.browser.select_project_automatically(project_id):
                logger.warning("No se pudo seleccionar proyecto automáticamente")
                # Mostrar mensaje al usuario
                if self.root:
                    self.root.after(0, lambda: messagebox.showwarning(
                        "Selección Manual Requerida", 
                        "No se pudo seleccionar el proyecto automáticamente.\n\n"
                        "Por favor, seleccione manualmente el proyecto."
                    ))
                    time.sleep(3)  # Dar tiempo para selección manual
            else:
                logger.info(f"Proyecto {project_id} seleccionado con éxito")
            
            # 3. Hacer clic en el botón de búsqueda
            if hasattr(self, 'status_var'):
                self.status_var.set("Realizando búsqueda...")
                if self.root:
                    self.root.update()
            
            time.sleep(1)
            self.browser.click_search_button()
            
            # 4. Continuar con la extracción
            logger.info("Iniciando proceso de extracción")
            if hasattr(self, 'status_var'):
                self.status_var.set("Extrayendo datos...")
                if self.root:
                    self.root.update()
            
            self.perform_extraction()
            
        except Exception as e:
            logger.error(f"Error al rellenar campos y extraer: {e}")
            if hasattr(self, 'status_var'):
                self.status_var.set(f"Error: {e}")
            
            if self.root:
                self.root.after(0, lambda: messagebox.showerror(
                    "Error", 
                    f"Error al rellenar campos: {e}"
                ))
    
    
    
    
    
    def exit_app(self):
        """Cierra la aplicación"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                confirm_exit = messagebox.askyesno(
                    "Proceso en curso", 
                    "Hay un proceso de extracción en curso. ¿Realmente desea salir?",
                    icon='warning'
                )
                if not confirm_exit:
                    return
                    
            if self.driver:
                try:
                    close_browser = messagebox.askyesno(
                        "Cerrar navegador", 
                        "¿Desea cerrar también el navegador?",
                        icon='question'
                    )
                    if close_browser:
                        self.browser.close()
                        logger.info("Navegador cerrado correctamente")
                except:
                    logger.warning("No se pudo cerrar el navegador correctamente")
            
            # Guardar configuración antes de salir
            self.save_config()
            
            self.root.destroy()
        except Exception as e:
            logger.error(f"Error al cerrar la aplicación: {e}")
            # En caso de error, forzar cierre
            self.root.destroy()
    
    def save_config(self):
        """Guarda la configuración actual"""
        try:
            config = {
                'client': self.client_var.get(),
                'project': self.project_var.get(),
                'excel_path': self.excel_file_path
            }
            
            config_dir = "config"
            if not os.path.exists(config_dir):
                os.makedirs(config_dir)
                
            config_path = os.path.join(config_dir, 'config.json')
            
            with open(config_path, 'w') as f:
                json.dump(config, f)
        except Exception as e:
            logger.error(f"Error al guardar configuración: {e}")
    
    def load_config(self):
        """Carga la configuración guardada"""
        try:
            config_path = os.path.join('config', 'config.json')
            
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
                        
                        
                    if 'client' in config and config['client']:
                        self.client_var.set(config['client'].strip())
                        
                    if 'project' in config and config['project']:
                        self.project_var.set(config['project'].strip())
                        
        except Exception as e:
            logger.error(f"Error al cargar configuración: {e}")
    
    def main_gui(self):
        """Punto de entrada principal con interfaz gráfica"""
        self.create_gui()
        self.root.mainloop()
        
        
        
        
        
        
        
def check_required_packages():
    """Verifica que estén instalados los paquetes requeridos"""
    required_packages = {
        "selenium": "Para automatización web",
        "pandas": "Para procesamiento de datos",
        "openpyxl": "Para manejo de archivos Excel"
    }
    
    missing = []
    for package, description in required_packages.items():
        try:
            __import__(package)
        except ImportError:
            missing.append(f"{package} ({description})")
    
    return missing

def create_shortcut(target_path, shortcut_path=None, icon_path=None):
    """Crea un acceso directo para la aplicación"""
    try:
        if not shortcut_path:
            desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')
            shortcut_path = os.path.join(desktop_path, "SAP Issues Extractor.lnk")
            
        if os.path.exists(shortcut_path):
            return shortcut_path
            
        import winshell
        from win32com.client import Dispatch
        
        shell = Dispatch('WScript.Shell')
        shortcut = shell.CreateShortCut(shortcut_path)
        shortcut.Targetpath = target_path
        shortcut.WorkingDirectory = os.path.dirname(target_path)
        if icon_path:
            shortcut.IconLocation = icon_path
        shortcut.save()
        
        return shortcut_path
    except Exception as e:
        logger.error(f"Error al crear acceso directo: {e}")
        return None

def main():
    """Función principal que ejecuta la aplicación"""
    extractor = None  # Inicializar la variable para evitar errores
    try:
        # Verificar paquetes requeridos
        missing_packages = check_required_packages()

        if missing_packages:
            print("Faltan las siguientes bibliotecas necesarias:")
            for package in missing_packages:
                print(f"  - {package}")
            print("\nPor favor, instálalas usando:")
            print(f"pip install {' '.join([p.split()[0] for p in missing_packages])}")
            input("\nPresiona ENTER para salir...")
            sys.exit(1)
        
        # Notificar sobre Pillow, pero no detener la ejecución
        try:
            __import__("PIL")
        except ImportError:
            print("Nota: La biblioteca Pillow no está disponible. Algunas características visuales estarán limitadas.")
            print("Si deseas instalarla, ejecuta: pip install Pillow")
        
        # Crear instancia del extractor
        extractor = IssuesExtractor()
        
        # Verificar si se desea interfaz gráfica o consola
        if len(sys.argv) > 1 and sys.argv[1] == "--console":
            # Modo consola
            extractor.choose_excel_file()
            extractor.run_extraction()
        else:
            # Modo interfaz gráfica (predeterminado)
            extractor.main_gui()
            
        logger.info("=== Proceso de extracción finalizado ===")
        
    except Exception as e:
        logger.critical(f"Error crítico en la ejecución: {e}")
        print(f"\n¡ERROR! Se ha producido un error crítico: {e}")
        print(f"Por favor, revisa el archivo de log para más detalles: {log_file}")
    finally:
        if extractor is not None and (not hasattr(extractor, 'root') or not extractor.root):
            # Solo mostrar mensaje final si estamos en modo consola
            input("\nPresiona ENTER para cerrar...")


if __name__ == "__main__":
    main()