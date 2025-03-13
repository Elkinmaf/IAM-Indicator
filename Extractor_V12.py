# -*- coding: utf-8 -*-
"""
Script Mejorado para Extracción de Issues SAP con Mejor Manejo de Scroll
---
Versión 12: Con soluciones optimizadas para extraer todos los issues, mejorando 
el manejo de scroll, detección de elementos, y soporte para paginación.
"""

import time
import pandas as pd
import os.path
import sys
import logging
import sqlite3
import threading
import re
from datetime import datetime
import webbrowser
import base64

# Intentar importar PIL, pero no fallar si no está disponible
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

from io import BytesIO

# Selenium imports
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
    JavascriptException
)

# Tkinter imports
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Configurar logging
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

log_file = os.path.join(
    log_dir, f"extraccion_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(log_file), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# Colores estilo SAP
SAP_COLORS = {
    "primary": "#1870C5",    # Azul SAP
    "secondary": "#354A5F",  # Azul oscuro SAP
    "success": "#107E3E",    # Verde SAP
    "warning": "#E9730C",    # Naranja SAP
    "danger": "#BB0000",     # Rojo SAP
    "light": "#F5F6F7",      # Gris claro SAP
    "dark": "#32363A",       # Gris oscuro SAP
    "white": "#FFFFFF",
    "gray": "#D3D7D9",
    "text": "#000000"        # Texto negro para máximo contraste
}


class IssuesExtractor:
    """Clase para extraer issues de SAP con interfaz gráfica y base de datos"""

    def __init__(self):
        """Inicializa la clase"""
        self.excel_file_path = None
        self.driver = None
        
        # Variables para la GUI
        self.root = None
        self.status_var = None
        self.client_var = None
        self.project_var = None
        self.project_combo = None
        self.log_text = None
        self.excel_filename_var = None  # Nueva variable para mostrar el nombre del archivo
        self.processing = False  # Bandera para indicar si hay un proceso en curso
        self.left_panel = None  # Referencia al panel izquierdo
        self.header_frame = None  # Referencia al marco de encabezado
        self.client_combo = None
        
        # Inicializar base de datos
        self.db_path = self.setup_database()
        
#########################################
    # MÉTODOS PARA LA BASE DE DATOS
    #########################################

    def setup_database(self):
        """Configura una base de datos SQLite para almacenar clientes y proyectos"""
        db_dir = "data"
        if not os.path.exists(db_dir):
            os.makedirs(db_dir)

        db_path = os.path.join(db_dir, "sap_extraction.db")

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Crear tablas si no existen
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS clients (
            erp_number TEXT PRIMARY KEY,
            name TEXT,
            business_partner TEXT,
            last_used TIMESTAMP
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            project_id TEXT PRIMARY KEY,
            client_erp TEXT,
            name TEXT,
            engagement_case TEXT,
            last_used TIMESTAMP,
            FOREIGN KEY (client_erp) REFERENCES clients(erp_number)
        )
        ''')

        # Insertar datos de ejemplo si no existen
        cursor.execute("SELECT COUNT(*) FROM clients")
        if cursor.fetchone()[0] == 0:
            cursor.execute('''
            INSERT INTO clients VALUES 
            ('1025541', 'Empresas Publicas De Medellin E.S.P.', '7976919', CURRENT_TIMESTAMP)
            ''')

            cursor.execute('''
            INSERT INTO projects VALUES 
            ('20096444', '1025541', 'S/4 HANA HEC Implementation', '20082191', CURRENT_TIMESTAMP)
            ''')

        conn.commit()
        conn.close()

        return db_path

    def get_clients_from_db(self):
        """Obtiene la lista de clientes desde la base de datos"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("SELECT erp_number, name FROM clients ORDER BY last_used DESC")
        clients = cursor.fetchall()

        conn.close()

        return [f"{erp} - {name}" for erp, name in clients]

    def get_projects_from_db(self, client_erp):
        """Obtiene la lista de proyectos para un cliente específico"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT project_id, name 
            FROM projects 
            WHERE client_erp = ? 
            ORDER BY last_used DESC
        """, (client_erp,))

        projects = cursor.fetchall()

        conn.close()

        return [f"{pid} - {name}" for pid, name in projects]

    def save_client_to_db(self, erp_number, name, business_partner):
        """Guarda un nuevo cliente en la base de datos"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            # Verificar si el cliente ya existe
            cursor.execute("SELECT * FROM clients WHERE erp_number = ?", (erp_number,))
            existing = cursor.fetchone()

            if existing:
                # Actualizar cliente existente
                cursor.execute("""
                    UPDATE clients 
                    SET name = ?, business_partner = ?, last_used = CURRENT_TIMESTAMP 
                    WHERE erp_number = ?
                """, (name, business_partner, erp_number))
            else:
                # Insertar nuevo cliente
                cursor.execute("""
                    INSERT INTO clients (erp_number, name, business_partner, last_used) 
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                """, (erp_number, name, business_partner))

            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al guardar cliente en BD: {e}")
            return False
        finally:
            conn.close()

    def save_project_to_db(self, project_id, client_erp, name, engagement_case):
        """Guarda un nuevo proyecto en la base de datos"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            # Verificar si el proyecto ya existe
            cursor.execute("SELECT * FROM projects WHERE project_id = ?", (project_id,))
            existing = cursor.fetchone()

            if existing:
                # Actualizar proyecto existente
                cursor.execute("""
                    UPDATE projects 
                    SET client_erp = ?, name = ?, engagement_case = ?, last_used = CURRENT_TIMESTAMP 
                    WHERE project_id = ?
                """, (client_erp, name, engagement_case, project_id))
            else:
                # Insertar nuevo proyecto
                cursor.execute("""
                    INSERT INTO projects (project_id, client_erp, name, engagement_case, last_used) 
                    VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (project_id, client_erp, name, engagement_case))

            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al guardar proyecto en BD: {e}")
            return False
        finally:
            conn.close()
            
    def update_client_usage(self, erp_number):
        """Actualiza la fecha de último uso de un cliente"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("""
                UPDATE clients 
                SET last_used = CURRENT_TIMESTAMP 
                WHERE erp_number = ?
            """, (erp_number,))
            
            conn.commit()
        except Exception as e:
            logger.error(f"Error al actualizar uso de cliente: {e}")
        finally:
            conn.close()
    
    def update_project_usage(self, project_id):
        """Actualiza la fecha de último uso de un proyecto"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("""
                UPDATE projects 
                SET last_used = CURRENT_TIMESTAMP 
                WHERE project_id = ?
            """, (project_id,))
            
            conn.commit()
        except Exception as e:
            logger.error(f"Error al actualizar uso de proyecto: {e}")
        finally:
            conn.close()
            
#########################################
    # MÉTODOS PARA EXCEL
    #########################################

    def choose_excel_file(self):
        """Permite al usuario elegir un archivo Excel existente o crear uno nuevo"""
        logger.info("Solicitando al usuario opciones para el archivo Excel...")

        if self.root:
            # Si la GUI está activa, no crear una ventana temporal
            root = self.root
        else:
            # Crear ventana temporal
            root = tk.Tk()
            root.withdraw()

        # Preguntar si desea usar un archivo existente o crear uno nuevo
        choice = messagebox.askquestion(
            "Archivo Excel",
            "¿Desea usar un archivo Excel existente?\n\n"
            + "Seleccione 'Sí' para elegir un archivo existente.\n"
            + "Seleccione 'No' para crear un nuevo archivo.",
            icon='question'
        )

        if choice == "yes":
            # Permitir al usuario seleccionar un archivo Excel existente
            file_path = filedialog.askopenfilename(
                title="Seleccione el archivo Excel de seguimiento",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            )

            if not file_path:  # El usuario canceló la selección
                logger.info(
                    "Usuario canceló la selección de archivo. Se creará uno nuevo."
                )
                # Crear un nombre de archivo por defecto con fecha y hora
                default_filename = f"Seguimiento_Recomendaciones_EPM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                # Permitir al usuario guardar con un nombre específico
                file_path = filedialog.asksaveasfilename(
                    title="Guardar nuevo archivo Excel",
                    defaultextension=".xlsx",
                    initialfile=default_filename,
                    filetypes=[("Archivos Excel", "*.xlsx")],
                )

                if not file_path:  # Si cancela de nuevo, usar el nombre por defecto
                    file_path = default_filename
                    logger.info(f"Se usará el nombre por defecto: {file_path}")

                # Crear un archivo Excel vacío con las columnas necesarias
                self._create_new_excel(file_path)
                logger.info(f"Creado nuevo archivo Excel: {file_path}")
        else:
            # Crear un nombre de archivo por defecto con fecha y hora
            default_filename = f"Seguimiento_Recomendaciones_EPM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            # Permitir al usuario guardar con un nombre específico
            file_path = filedialog.asksaveasfilename(
                title="Guardar nuevo archivo Excel",
                defaultextension=".xlsx",
                initialfile=default_filename,
                filetypes=[("Archivos Excel", "*.xlsx")],
            )

            if not file_path:  # Si cancela, usar el nombre por defecto
                file_path = default_filename
                logger.info(f"Se usará el nombre por defecto: {file_path}")

            # Crear un archivo Excel vacío con las columnas necesarias
            self._create_new_excel(file_path)
            logger.info(f"Creado nuevo archivo Excel: {file_path}")

        self.excel_file_path = file_path

        # Actualizar la interfaz si existe
        if hasattr(self, 'status_var') and self.status_var:
            self.status_var.set(f"Archivo Excel seleccionado: {os.path.basename(file_path)}")
        
        # Actualizar el nombre del archivo en la etiqueta
        if hasattr(self, 'excel_filename_var') and self.excel_filename_var:
            self.excel_filename_var.set(f"Archivo: {os.path.basename(file_path)}")
            
        return file_path

    def _create_new_excel(self, file_path):
        """Crea un nuevo archivo Excel con las columnas necesarias"""
        try:
            # Crear un DataFrame vacío con las columnas necesarias
            columns = [
                "Title",
                "Type",
                "Priority",
                "Status",
                "Deadline",
                "Due Date",
                "Created By",
                "Created On",
                "Last Updated",
                "Comments",
            ]

            df = pd.DataFrame(columns=columns)

            # Guardar el DataFrame vacío como un archivo Excel
            df.to_excel(file_path, index=False)
            logger.info(f"Archivo Excel creado exitosamente: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error al crear nuevo archivo Excel: {e}")
            return False

    def update_excel(self, issues_data):
        """Actualiza el archivo Excel con los datos extraídos"""
        try:
            logger.info(f"Actualizando archivo Excel: {self.excel_file_path}...")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Actualizando archivo Excel...")

            if not issues_data:
                logger.warning("No hay datos para actualizar en Excel")
                return False

            if os.path.exists(self.excel_file_path):
                existing_df = pd.read_excel(self.excel_file_path)
                logger.info(
                    f"Archivo Excel existente cargado con {len(existing_df)} registros."
                )
            else:
                existing_df = pd.DataFrame(
                    columns=[
                        "Title",
                        "Type",
                        "Priority",
                        "Status",
                        "Deadline",
                        "Due Date",
                        "Created By",
                        "Created On",
                        "Last Updated",
                        "Comments",
                    ]
                )
                logger.info("Creando nuevo archivo Excel.")

            new_df = pd.DataFrame(issues_data)

            new_items = 0
            updated_items = 0

            updated_df = existing_df.copy()

            for _, new_row in new_df.iterrows():
                title_exists = False
                if "Title" in existing_df.columns:
                    existing_row = existing_df[existing_df["Title"] == new_row["Title"]]
                    title_exists = len(existing_row) > 0

                if not title_exists:
                    # Agregar fecha de última actualización para elementos nuevos
                    new_row_dict = new_row.to_dict()
                    new_row_dict["Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    new_row_df = pd.DataFrame([new_row_dict])
                    updated_df = pd.concat([updated_df, new_row_df], ignore_index=True)
                    new_items += 1
                    logger.info(f"Nuevo issue añadido: '{new_row['Title']}'")
                else:
                    # Comprobar si el estado ha cambiado
                    if (
                        "Status" in existing_row.columns
                        and existing_row["Status"].values[0] != new_row["Status"]
                    ):
                        mask = updated_df["Title"] == new_row["Title"]
                        updated_df.loc[mask, "Status"] = new_row["Status"]
                        updated_df.loc[mask, "Last Updated"] = datetime.now().strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        updated_items += 1
                        logger.info(
                            f"Actualizado estado de '{new_row['Title']}': '{existing_row['Status'].values[0]}' → '{new_row['Status']}'"
                        )

                    # Actualizar otras columnas si hay cambios
                    for column in ["Priority", "Type", "Due Date", "Deadline", "Created By", "Created On"]:
                        if column in new_row and column in existing_row:
                            if existing_row[column].values[0] != new_row[column]:
                                mask = updated_df["Title"] == new_row["Title"]
                                updated_df.loc[mask, column] = new_row[column]
                                updated_df.loc[mask, "Last Updated"] = datetime.now().strftime(
                                    "%Y-%m-%d %H:%M:%S"
                                )
                                updated_items += 1
                                logger.info(
                                    f"Actualizado {column} de '{new_row['Title']}': '{existing_row[column].values[0]}' → '{new_row[column]}'"
                                )

            updated_df.to_excel(self.excel_file_path, index=False)
            logger.info(
                f"Archivo Excel actualizado correctamente: {self.excel_file_path}"
            )
            logger.info(
                f"Resumen: {new_items} nuevos items, {updated_items} items actualizados"
            )
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Excel actualizado: {new_items} nuevos, {updated_items} actualizados")

            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = load_workbook(self.excel_file_path)
                ws = wb.active

                header_fill = PatternFill(
                    start_color="1F4E78", end_color="1F4E78", fill_type="solid"
                )
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center")

                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border

                        if col == 4:  # Columna Status
                            status = cell.value.upper() if cell.value else ""

                            if "DONE" in status:
                                cell.fill = PatternFill(
                                    start_color="CCFFCC",
                                    end_color="CCFFCC",
                                    fill_type="solid",
                                )
                            elif "OPEN" in status:
                                cell.fill = PatternFill(
                                    start_color="FFCCCC",
                                    end_color="FFCCCC",
                                    fill_type="solid",
                                )
                            elif "READY" in status:
                                cell.fill = PatternFill(
                                    start_color="FFFFCC",
                                    end_color="FFFFCC",
                                    fill_type="solid",
                                )
                            elif "IN PROGRESS" in status:
                                cell.fill = PatternFill(
                                    start_color="FFE6CC",
                                    end_color="FFE6CC",
                                    fill_type="solid",
                                )

                for col in range(1, ws.max_column + 1):
                    max_length = 0
                    for row in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))

                    adjusted_width = max(10, min(50, max_length + 2))
                    ws.column_dimensions[
                        ws.cell(row=1, column=col).column_letter
                    ].width = adjusted_width

                wb.save(self.excel_file_path)
                logger.info("Formato aplicado al archivo Excel correctamente")
            except Exception as format_e:
                logger.warning(f"No se pudo aplicar formato al Excel: {format_e}")

            # Mostrar un mensaje emergente al finalizar
            if self.root:
                messagebox.showinfo(
                    "Proceso Completado", 
                    f"El archivo Excel ha sido actualizado correctamente.\n\n"
                    f"Se han agregado {new_items} nuevos issues y actualizado {updated_items} issues existentes."
                )

            return True
        except Exception as e:
            logger.error(f"Error al actualizar el archivo Excel: {e}")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error al actualizar Excel: {e}")
                
            return False
        

#########################################
    # MÉTODOS PARA NAVEGACIÓN Y SELENIUM
    #########################################

    def connect_to_browser(self):
        """Conecta con el navegador abierto con perfil guardado"""
        logger.info("Iniciando con perfil guardado...")
        
        try:
            # Ruta al directorio del perfil
            user_data_dir = os.path.join(os.environ['USERPROFILE'], 'AppData', 'Local', 'Google', 'Chrome', 'SAP_Automation')
            
            # Crear directorio si no existe
            if not os.path.exists(user_data_dir):
                os.makedirs(user_data_dir)
            
            # Opciones con perfil guardado
            chrome_options = Options()
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
            chrome_options.add_argument("--profile-directory=Default")
            
            # Opciones adicionales para mejorar rendimiento y estabilidad
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-extensions")
            
            # Agregar opciones para permitir que el usuario use el navegador mientras se ejecuta el script
            chrome_options.add_experimental_option("detach", True)
            
            # Iniciar el navegador
            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 60)  # Aumentar timeout a 60 segundos
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Navegador Chrome iniciado correctamente")
                
            return True
        except Exception as e:
            logger.error(f"Error al iniciar Navegador: {e}")
            
            # Mostrar error en la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error al iniciar el Navegador: {e}")
                
            return False

    def get_total_issues_count(self):
        """Obtiene el número total de issues desde el encabezado de la tabla"""
        try:
            # Intentar varias estrategias para obtener el número total de issues
            # Estrategia 1: Buscar el texto "Issues (96)" o similar
            try:
                header_text = self.driver.find_element(
                    By.XPATH, "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                ).text
                logger.info(f"Texto del encabezado de issues: {header_text}")
                
                # Extraer el número entre paréntesis
                match = re.search(r'\((\d+)\)', header_text)
                if match:
                    return int(match.group(1))
            except NoSuchElementException:
                logger.warning("No se encontró el encabezado Issues con formato (número)")
            
            # Estrategia 2: Buscar contador específico de SAP UI5
            try:
                counter_element = self.driver.find_element(
                    By.XPATH, "//div[contains(@class, 'sapMITBCount')]"
                )
                if counter_element.text.isdigit():
                    return int(counter_element.text)
            except NoSuchElementException:
                logger.warning("No se encontró contador de issues en formato SAP UI5")
            
            # Estrategia 3: Buscar cualquier texto que contenga el número de issues
            patterns = [
                r'(\d+) issues',
                r'issues: (\d+)',
                r'total (\d+)',
                r'count: (\d+)'
            ]
            
            page_text = self.driver.find_element(By.TAG_NAME, "body").text.lower()
            
            for pattern in patterns:
                match = re.search(pattern, page_text)
                if match:
                    return int(match.group(1))
            
            logger.warning("No se pudo determinar el número total de issues, usando valor por defecto")
            return 100  # Valor por defecto
            
        except Exception as e:
            logger.error(f"Error al obtener el total de issues: {e}")
            return 100  # Valor por defecto si hay error

    def check_for_pagination(self):
        """Verifica si la tabla tiene paginación y devuelve los controles de paginación si existen"""
        try:
            # Buscar diferentes tipos de controles de paginación que podrían existir
            pagination_selectors = [
                "//div[contains(@class, 'sapMPaginator')]",
                "//div[contains(@class, 'sapUiTablePaginator')]",
                "//div[contains(@class, 'pagination')]",
                "//button[contains(@class, 'navButton') or contains(@aria-label, 'Next') or contains(@aria-label, 'Siguiente')]",
                "//span[contains(@class, 'sapMPaginatorButton')]",
                "//button[contains(text(), 'Next') or contains(text(), 'Siguiente')]",
                "//a[contains(@class, 'sapMBtn') and contains(@aria-label, 'Next')]"
            ]
            
            for selector in pagination_selectors:
                elements = self.driver.find_elements(By.XPATH, selector)
                if elements:
                    logger.info(f"Se encontraron controles de paginación: {len(elements)} elementos con selector {selector}")
                    return elements
            
            # Buscar elementos específicos para "Show More" o "Load More" que actúan como paginación
            load_more_selectors = [
                "//button[contains(text(), 'More') or contains(text(), 'más') or contains(text(), 'Show')]",
                "//a[contains(text(), 'More') or contains(text(), 'Load')]",
                "//div[contains(@class, 'sapMListShowMoreButton')]",
                "//span[contains(text(), 'Show') and contains(text(), 'More')]/..",
                "//span[contains(@class, 'sapUiTableColShowMoreButton')]"
            ]
            
            for selector in load_more_selectors:
                elements = self.driver.find_elements(By.XPATH, selector)
                if elements:
                    logger.info(f"Se encontraron botones 'Show More': {len(elements)} elementos con selector {selector}")
                    return elements
            
            logger.info("No se encontraron controles de paginación en la tabla")
            return None
            
        except Exception as e:
            logger.error(f"Error al verificar paginación: {e}")
            return None

    def click_pagination_next(self, pagination_elements):
        """Intenta hacer clic en el botón de siguiente página"""
        if not pagination_elements:
            return False
            
        try:
            # Intentar encontrar el botón "Next" o "Siguiente" entre los elementos de paginación
            next_button = None
            
            # Buscar por atributos típicos de botones "Next"
            for element in pagination_elements:
                try:
                    aria_label = element.get_attribute("aria-label") or ""
                    text = element.text.lower()
                    classes = element.get_attribute("class") or ""
                    
                    # Comprobar si es un botón "Next" o "Siguiente"
                    if ("next" in aria_label.lower() or 
                        "siguiente" in aria_label.lower() or
                        "next" in text or 
                        "siguiente" in text or
                        "show more" in text.lower() or
                        "more" in text.lower()):
                        
                        next_button = element
                        break
                        
                    # Comprobar por clase CSS
                    if ("next" in classes.lower() or 
                        "pagination-next" in classes.lower() or
                        "sapMBtn" in classes and "NavButton" in classes):
                        
                        next_button = element
                        break
                except Exception:
                    continue
            
            # Si se encontró un botón Next, intentar hacer clic
            if next_button:
                # Verificar si el botón está habilitado
                disabled = next_button.get_attribute("disabled") or next_button.get_attribute("aria-disabled") == "true"
                
                if disabled:
                    logger.info("Botón de siguiente página está deshabilitado")
                    return False
                    
                # Scroll hacia el botón para asegurar que está visible
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
                time.sleep(0.5)
                
                # Intentar clic primero con JavaScript para evitar problemas de interactividad
                try:
                    self.driver.execute_script("arguments[0].click();", next_button)
                    logger.info("Clic en botón 'Next' realizado con JavaScript")
                    time.sleep(2)  # Esperar a que cargue la siguiente página
                    return True
                except JavascriptException:
                    pass
                
                # Si falló con JavaScript, intentar clic normal
                try:
                    next_button.click()
                    logger.info("Clic en botón 'Next' realizado")
                    time.sleep(2)  # Esperar a que cargue la siguiente página
                    return True
                except ElementClickInterceptedException:
                    # Si hay algún elemento superpuesto, intentar con Actions
                    from selenium.webdriver.common.action_chains import ActionChains
                    actions = ActionChains(self.driver)
                    actions.move_to_element(next_button).click().perform()
                    logger.info("Clic en botón 'Next' realizado con ActionChains")
                    time.sleep(2)
                    return True
                    
            # Si se encontraron elementos de paginación pero no un botón Next claro,
            # intentar con el último elemento (asumiendo que es el de "siguiente")
            if pagination_elements and len(pagination_elements) > 0:
                try:
                    # Intentar con el último elemento
                    last_element = pagination_elements[-1]
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", last_element)
                    time.sleep(0.5)
                    self.driver.execute_script("arguments[0].click();", last_element)
                    logger.info("Clic en último elemento de paginación realizado")
                    time.sleep(2)
                    return True
                except Exception as e:
                    logger.warning(f"Error al hacer clic en último elemento de paginación: {e}")
            
            logger.warning("No se pudo identificar o hacer clic en el botón 'Next'")
            return False
            
        except Exception as e:
            logger.error(f"Error al hacer clic en paginación: {e}")
            return False

    def scroll_to_load_all_items(self, total_expected=100, max_attempts=200):
        """
        Estrategia mejorada de scroll para cargar todos los elementos de la tabla
        con soporte para paginación y detección de carga completa
        """
        logger.info(f"Iniciando proceso de scroll mejorado para cargar {total_expected} elementos...")
        
        # Actualizar la interfaz si existe
        if hasattr(self, 'status_var') and self.status_var:
            self.status_var.set(f"Cargando elementos...")
        
        # Lista para almacenar las filas procesadas y evitar duplicados
        processed_titles = set()
        
        attempt = 0
        previous_rows_count = 0
        no_change_count = 0
        
        # Verificar si hay paginación en la tabla
        pagination_elements = self.check_for_pagination()
        has_pagination = pagination_elements is not None and len(pagination_elements) > 0
        
        logger.info(f"¿La tabla tiene paginación? {'Sí' if has_pagination else 'No'}")
        
        # Estrategia combinada: intentar scroll primero, luego paginación si es necesario
        while attempt < max_attempts:
            try:
                # 1. Hacer scroll invisible para cargar más elementos
                for scroll_method in range(3):
                    # Método 1: Scroll general al final de la página
                    if scroll_method == 0:
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    
                    # Método 2: Scroll en contenedores específicos de SAP
                    elif scroll_method == 1:
                        self.driver.execute_script("""
                            var containers = document.querySelectorAll('.sapMListItems, .sapMTableTBody, .sapUiTableCtrlScr');
                            if (containers.length > 0) {
                                for(var i=0; i<containers.length; i++) {
                                    containers[i].scrollTop = containers[i].scrollHeight;
                                }
                            }
                        """)
                    
                    # Método 3: Scroll por secciones para simular desplazamiento natural
                    elif scroll_method == 2:
                        # Obtener altura del documento
                        doc_height = self.driver.execute_script("return document.body.scrollHeight")
                        # Scroll por secciones
                        for pos in range(0, doc_height, 500):
                            self.driver.execute_script(f"window.scrollTo(0, {pos});")
                            time.sleep(0.1)
                    
                    time.sleep(0.5)
                
                # 2. Buscar y hacer clic en botones "Show More" o "Load More" si existen
                if attempt % 2 == 0:  # Cada 2 intentos para no sobrecargarlo
                    try:
                        load_more_buttons = self.driver.find_elements(
                            By.XPATH,
                            "//button[contains(text(), 'More')] | " +
                            "//button[contains(text(), 'más')] | " +
                            "//a[contains(text(), 'More')] | " +
                            "//div[contains(@class, 'sapMListShowMoreButton')] | " +
                            "//span[contains(text(), 'Show') and contains(text(), 'More')]/.. | " +
                            "//span[contains(@class, 'sapUiTableColShowMoreButton')]"
                        )
                        
                        if load_more_buttons:
                            for btn in load_more_buttons:
                                try:
                                    # Hacer visible el botón
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                                    time.sleep(0.5)
                                    # Hacer clic sin desplazamiento visual
                                    self.driver.execute_script("arguments[0].click();", btn)
                                    logger.info("Se hizo clic en botón 'Show More'")
                                    time.sleep(2)  # Dar más tiempo para cargar
                                except:
                                    continue
                    except Exception as btn_e:
                        logger.debug(f"Error al buscar botón 'Show More': {btn_e}")
                
                # 3. Contar filas visibles sin resaltarlas
                rows = self.find_table_rows(highlight=False)
                current_rows_count = len(rows)
                
                logger.info(f"Intento {attempt+1}: {current_rows_count} filas cargadas")
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set(f"Cargando elementos: {current_rows_count}/{total_expected}")
                    # Forzar actualización de la interfaz
                    if self.root:
                        self.root.update()
                
                # 4. Verificación de carga completa
                if current_rows_count == previous_rows_count:
                    no_change_count += 1
                    
                    # Si hay paginación y no hay cambios por más de 5 intentos, intentar pasar a la siguiente página
                    if has_pagination and no_change_count >= 5:
                        logger.info("Intentando pasar a la siguiente página...")
                        # Actualizar elementos de paginación (podrían haber cambiado)
                        pagination_elements = self.check_for_pagination()
                        if pagination_elements:
                            if self.click_pagination_next(pagination_elements):
                                logger.info("Se pasó a la siguiente página")
                                no_change_count = 0  # Reiniciar contador
                                time.sleep(3)  # Dar tiempo para cargar la nueva página
                                continue
                    
                    # Si no hay cambios por muchos intentos, realizar scroll adicional
                    if no_change_count >= 5:
                        # Scroll por posiciones incrementales
                        for scroll_pos in range(1000, 10000, 1000):
                            self.driver.execute_script(f"window.scrollTo(0, {scroll_pos});")
                            time.sleep(0.3)
                        
                        # Intentar expandir la tabla con teclas de dirección
                        try:
                            body = self.driver.find_element(By.TAG_NAME, "body")
                            body.send_keys(Keys.PAGE_DOWN)
                            time.sleep(0.5)
                            body.send_keys(Keys.PAGE_DOWN)
                            time.sleep(0.5)
                            body.send_keys(Keys.END)
                            time.sleep(0.5)
                        except:
                            pass
                    
                    # Criterios de finalización
                    if no_change_count >= 15 and current_rows_count >= total_expected * 0.9:
                        logger.info(f"Se han cargado {current_rows_count} filas (>= 90% del total esperado). Terminando scroll.")
                        break
                        
                    if no_change_count >= 20:
                        logger.warning(f"No se detectaron más filas después de {no_change_count} intentos sin cambios.")
                        break
                else:
                    # Reiniciar contador si se encontraron más filas
                    no_change_count = 0
                    
                previous_rows_count = current_rows_count
                attempt += 1
                    
                # Si se alcanzó o superó el número esperado, terminar
                if current_rows_count >= total_expected:
                    logger.info(f"Se han cargado {current_rows_count} filas (>= {total_expected} esperadas). Scroll completado.")
                    break
                    
            except Exception as e:
                logger.warning(f"Error durante el scroll en intento {attempt+1}: {e}")
                attempt += 1
            
        # Verificar cobertura
        coverage = (previous_rows_count / total_expected) * 100 if total_expected > 0 else 0
        logger.info(f"Proceso de scroll completado. Cobertura: {coverage:.2f}% ({previous_rows_count}/{total_expected})")
        
        # Actualizar la interfaz si existe
        if hasattr(self, 'status_var') and self.status_var:
            self.status_var.set(f"Elementos cargados: {previous_rows_count}/{total_expected} ({coverage:.1f}%)")
            
        return previous_rows_count

    def find_table_rows(self, highlight=False):
        """
        Encuentra todas las filas de la tabla utilizando múltiples estrategias
        para garantizar la máxima cobertura
        """
        all_rows = []

        # Selectores mejorados para SAP UI5, con mayor cobertura de casos
        selectors = [
            # Selectores estándar de SAP
            "//table[contains(@class, 'sapMListTbl')]/tbody/tr[not(contains(@class, 'sapMListTblHeader'))]",
            "//div[contains(@class, 'sapMList')]//li[contains(@class, 'sapMLIB')]",
            "//table[contains(@class, 'sapMList')]/tbody/tr",
            "//div[@role='row'][not(contains(@class, 'sapMListHeaderSubTitleItems')) and not(contains(@class, 'sapMListTblHeader'))]",
            "//div[contains(@class, 'sapMListItems')]/div[contains(@class, 'sapMListItem')]",
            "//div[contains(@class, 'sapMListItems')]//div[contains(@class, 'sapMObjectIdentifier')]/..",
            "//div[contains(@class, 'sapMListItem')]",
            
            # Selectores adicionales para SAP Fiori
            "//div[contains(@class, 'sapMList')]//li[@tabindex]",
            "//div[contains(@class, 'sapUiTable')]//tr[contains(@class, 'sapUiTableRow')]",
            "//div[contains(@class, 'sapUiTableRowHdr')]/..",
            "//table[contains(@class, 'sapMTable')]//tr[not(contains(@class, 'sapMListTblHeaderRow'))]",
            
            # Selectores específicos para SDWork Center de SAP
            "//div[contains(@class, 'sdworkItems')]//div[contains(@class, 'sapMLIB')]",
            "//div[contains(@class, 'issueList')]//div[contains(@class, 'sapMLIB')]",
            "//div[contains(@id, 'issue')]//div[contains(@class, 'sapMLIB')]",
            
            # Selectores genéricos pero más específicos
            "//div[contains(@class, 'sapMLIB-CTX')]",
            "//div[contains(@class, 'sapMObjectListItem')]",
            "//div[contains(@class, 'sapMListModeMultiSelect')]//div[contains(@class, 'sapMLIB')]"
        ]

        for selector in selectors:
            try:
                rows = self.driver.find_elements(By.XPATH, selector)
                if len(rows) > 0:
                    logger.info(f"Se encontraron {len(rows)} filas con selector: {selector}")

                    valid_rows = []
                    for row in rows:
                        try:
                            has_content = False
                            text_elements = row.find_elements(
                                By.XPATH, ".//span | .//div | .//a"
                            )
                            for element in text_elements:
                                if element.text and element.text.strip():
                                    has_content = True
                                    break

                            if has_content:
                                # Verificar que no sea un encabezado
                                is_header = "header" in row.get_attribute("class").lower() if row.get_attribute("class") else False
                                if not is_header:
                                    valid_rows.append(row)
                        except:
                            valid_rows.append(row)

                    if len(valid_rows) > 0:
                        all_rows = valid_rows
                        logger.info(f"Se encontraron {len(valid_rows)} filas válidas")

                        if len(valid_rows) >= 75:  # Aproximado al 80% de un total normal
                            break
                
                # Si el selector devolvió resultados pero no filas válidas, intentar más a fondo
                elif len(rows) > 0:
                    potential_rows = []
                    for row in rows:
                        try:
                            # Verificar si tiene suficiente contenido para ser una fila de datos
                            if len(row.text.strip()) > 20:  # Al menos tiene algo de texto
                                potential_rows.append(row)
                        except:
                            continue
                    
                    if potential_rows:
                        logger.info(f"Se encontraron {len(potential_rows)} filas potenciales con selector: {selector}")
                        all_rows = potential_rows
            except Exception as e:
                logger.debug(f"Error al buscar filas con selector {selector}: {e}")

        # Si no se encontraron filas con los selectores estándar, usar aproximación alternativa
        if len(all_rows) == 0:
            logger.warning(
                "No se encontraron filas con los selectores estándar. Intentando aproximación alternativa..."
            )
            try:
                # Buscar elementos que parezcan filas por su estructura y contenido
                any_rows = self.driver.find_elements(
                    By.XPATH,
                    "//div[contains(@class, 'sapM')] | //tr | //li[contains(@class, 'sapM')]",
                )

                for element in any_rows:
                    try:
                        # Verificar si tiene texto y varios hijos (probablemente una fila de datos)
                        if element.text and len(element.text.strip()) > 10:
                            children = element.find_elements(By.XPATH, ".//*")
                            if len(children) >= 3:
                                # Verificar que no sea un elemento contenedor principal
                                parent_elements = element.find_elements(By.XPATH, "./ancestor::div[contains(@class, 'sapMList') or contains(@class, 'sapMTable')]")
                                if len(parent_elements) > 0:
                                    all_rows.append(element)
                    except:
                        continue

                logger.info(
                    f"Aproximación alternativa encontró {len(all_rows)} posibles filas"
                )
            except Exception as e:
                logger.error(f"Error en la aproximación alternativa: {e}")

        # Solo resaltar visualmente si se indica explícitamente
        if highlight and len(all_rows) > 0:
            try:
                self.driver.execute_script(
                    """
                    arguments[0].scrollIntoView(true);
                    arguments[0].style.border = '2px solid red';
                """,
                    all_rows[0],
                )
                self.driver.save_screenshot("rows_found.png")
                logger.info(
                    f"Captura de pantalla de filas encontradas guardada como 'rows_found.png'"
                )
            except:
                pass

        # Eliminar duplicados manteniendo el orden
        unique_rows = []
        seen_ids = set()
        
        for row in all_rows:
            try:
                row_id = row.id
                if row_id not in seen_ids:
                    seen_ids.add(row_id)
                    unique_rows.append(row)
            except:
                # Si no podemos obtener un ID, usar una aproximación con texto y clase
                try:
                    row_text = row.text[:50] if row.text else ""
                    row_class = row.get_attribute("class") or ""
                    row_signature = f"{row_text}|{row_class}"
                    
                    if row_signature not in seen_ids:
                        seen_ids.add(row_signature)
                        unique_rows.append(row)
                except:
                    # Si todo falla, simplemente agregar la fila
                    unique_rows.append(row)
        
        logger.info(f"Total de filas únicas encontradas: {len(unique_rows)}")
        return unique_rows
    
    def extract_issues_data(self):
            """
            Extrae datos de issues desde la tabla con procesamiento mejorado
            para garantizar la extracción de todos los issues
            """
            try:
                logger.info("Esperando a que cargue la tabla de issues...")
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set("Esperando a que cargue la tabla de issues...")
                    
                # Esperar a que cargue la página inicial
                time.sleep(3)
                
                # Obtener el número total de issues
                total_issues = self.get_total_issues_count()
                logger.info(f"Total de issues a procesar: {total_issues}")
                
                # Hacer scroll para cargar todos los elementos
                loaded_rows_count = self.scroll_to_load_all_items(total_issues)
                
                # Verificar si hay paginación y procesar múltiples páginas si es necesario
                pagination_elements = self.check_for_pagination()
                has_pagination = pagination_elements is not None and len(pagination_elements) > 0
                
                # Lista para almacenar todos los datos extraídos
                all_issues_data = []
                seen_titles = set()  # Para evitar duplicados
                
                page_num = 1
                max_pages = 20  # Límite de seguridad
                
                while page_num <= max_pages:
                    logger.info(f"Procesando página {page_num}...")
                    
                    # Actualizar la interfaz si existe
                    if hasattr(self, 'status_var') and self.status_var:
                        self.status_var.set(f"Procesando página {page_num}...")
                        if self.root:
                            self.root.update()
                    
                    # Obtener filas de la página actual
                    rows = self.find_table_rows(highlight=False)
                    
                    if not rows:
                        logger.warning(f"No se encontraron filas en la página {page_num}")
                        break
                    
                    logger.info(f"Encontradas {len(rows)} filas en la página {page_num}")
                    
                    # Procesar cada fila
                    page_issues_data = self._process_table_rows(rows, seen_titles)
                    
                    # Agregar los datos de esta página al resultado total
                    all_issues_data.extend(page_issues_data)
                    
                    logger.info(f"Extraídos {len(page_issues_data)} issues de la página {page_num}")
                    logger.info(f"Total de issues extraídos hasta ahora: {len(all_issues_data)}")
                    
                    # Si no hay paginación o ya no se pueden extraer más datos, terminar
                    if not has_pagination or len(page_issues_data) == 0:
                        break
                    
                    # Intentar pasar a la siguiente página
                    if pagination_elements:
                        if not self.click_pagination_next(pagination_elements):
                            logger.info("No se pudo pasar a la siguiente página, terminando extracción")
                            break
                            
                        # Esperar a que cargue la nueva página
                        time.sleep(3)
                        
                        # Actualizar elementos de paginación (pueden cambiar entre páginas)
                        pagination_elements = self.check_for_pagination()
                        
                        page_num += 1
                    else:
                        break
                
                logger.info(f"Extracción completada. Total de issues extraídos: {len(all_issues_data)}")
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set(f"Extracción completada. Total: {len(all_issues_data)} issues")
                
                # Guardar los datos crudos para depuración
                with open("issues_data_raw.txt", "w", encoding="utf-8") as f:
                    for item in all_issues_data:
                        f.write(str(item) + "\n")
                
                return all_issues_data
            
            except Exception as e:
                logger.error(f"Error en la extracción de datos: {e}")
                
                return []

    def _process_table_rows(self, rows, seen_titles):
        """
        Procesa las filas de la tabla y extrae los datos de cada issue
        Evita duplicados usando el conjunto seen_titles
        """
        issues_data = []
        processed_count = 0
        
        # Actualizar la interfaz si existe
        if hasattr(self, 'status_var') and self.status_var:
            self.status_var.set(f"Procesando {len(rows)} filas...")
            
        for index, row in enumerate(rows):
            try:
                # Actualizar la interfaz si existe, pero sin resaltar filas
                if index % 20 == 0:  # Cada 20 filas
                    if hasattr(self, 'status_var') and self.status_var:
                        self.status_var.set(f"Procesando fila {index+1} de {len(rows)}...")
                        if self.root:
                            self.root.update()

                # Intentar extraer el título primero
                title = None
                
                # Intentar múltiples formas de extraer el título
                title_extractors = [
                    lambda r: r.find_element(By.XPATH, ".//a").text,
                    lambda r: r.find_element(By.XPATH, ".//span[contains(@class, 'title')]").text,
                    lambda r: r.find_element(By.XPATH, ".//div[contains(@class, 'title')]").text,
                    lambda r: r.find_elements(By.XPATH, ".//div[@role='gridcell']")[0].text if r.find_elements(By.XPATH, ".//div[@role='gridcell']") else None,
                    lambda r: r.find_elements(By.XPATH, ".//td")[0].text if r.find_elements(By.XPATH, ".//td") else None,
                    lambda r: r.find_element(By.XPATH, ".//*[contains(@id, 'title')]").text,
                    lambda r: r.find_element(By.XPATH, ".//div[@title]").get_attribute("title"),
                    lambda r: r.find_element(By.XPATH, ".//span[@title]").get_attribute("title")
                ]

                for extractor in title_extractors:
                    try:
                        title_text = extractor(row)
                        if title_text and title_text.strip():
                            title = title_text.strip()
                            break
                    except:
                        continue
                
                # Si no encontramos un título específico, intentar con el texto completo de la fila
                if not title:
                    try:
                        full_text = row.text.strip()
                        if full_text:
                            # Tomar la primera línea como título o los primeros 50 caracteres
                            lines = full_text.split('\n')
                            if lines:
                                title = lines[0].strip()
                                if len(title) > 100:  # Si es muy largo, recortar
                                    title = title[:100] + "..."
                    except:
                        pass
                
                # Si aún no hay título, generar uno genérico
                if not title:
                    title = f"Issue sin título #{index+1}"
                    logger.warning(f"No se pudo extraer título para la fila {index}, usando '{title}'")
                
                # Si ya procesamos este título, saltarlo
                title_lower = title.lower()
                if title_lower in seen_titles:
                    logger.debug(f"Título duplicado: '{title}', saltando...")
                    continue
                
                seen_titles.add(title_lower)
                
                # Inicializar todas las columnas con valores por defecto
                type_text = "Issue"  # Valor por defecto mejorado para Type
                priority = "N/A"
                status = "N/A"
                deadline = ""  # Deadline suele estar vacío
                due_date = "N/A"
                created_by = "N/A"
                created_on = "N/A"
                
                # Buscar el tipo de issue de forma más específica
                try:
                    # Primero buscar en la segunda columna
                    type_elements = row.find_elements(By.XPATH, 
                        ".//div[@role='gridcell'][2]//span | "
                        ".//td[2]//span | "
                        ".//div[contains(@class, 'type')] | "
                        ".//span[contains(@class, 'type')]")
                    
                    if type_elements:
                        type_text = type_elements[0].text.strip()
                        
                    # Si no se encontró, intentar otra aproximación
                    if not type_text or type_text == "Issue":
                        # Si el título contiene indicaciones de tipo, usarlo
                        if title:
                            potential_types = ["Recommendation", "Implementation", "Question", 
                                              "Problem", "Incident", "Request", "Task"]
                            for potential_type in potential_types:
                                if potential_type.lower() in title.lower():
                                    type_text = potential_type
                                    break
                except Exception as type_e:
                    logger.debug(f"Error al extraer tipo de issue: {type_e}")
                
                # Extraer prioridad desde los iconos o indicadores visuales
                try:
                    # Buscar diferentes indicadores de prioridad
                    priority_indicators = [
                        (By.XPATH, ".//span[contains(@class, 'sapMGaugeNegativeColor')]", "Very High"),
                        (By.XPATH, ".//span[contains(@class, 'sapMGaugeCriticalColor')]", "High"),
                        (By.XPATH, ".//span[contains(@class, 'sapMGaugeNeutralColor')]", "Medium"),
                        (By.XPATH, ".//span[contains(@class, 'sapMGaugePositiveColor')]", "Low"),
                        (By.XPATH, ".//span[contains(text(), 'Very High')]", "Very High"),
                        (By.XPATH, ".//span[contains(text(), 'High')]", "High"),
                        (By.XPATH, ".//span[contains(text(), 'Medium')]", "Medium"),
                        (By.XPATH, ".//span[contains(text(), 'Low')]", "Low"),
                        (By.XPATH, ".//i[contains(@class, 'High')]", "High"),
                        (By.XPATH, ".//div[@role='gridcell'][3]//span", ""),
                    ]
                    
                    for locator, indicator_text in priority_indicators:
                        elements = row.find_elements(locator)
                        if elements:
                            if indicator_text:
                                priority = indicator_text
                            elif elements[0].text:
                                cell_text = elements[0].text.lower()
                                if "very high" in cell_text:
                                    priority = "Very High"
                                elif "high" in cell_text:
                                    priority = "High"
                                elif "medium" in cell_text:
                                    priority = "Medium"
                                elif "low" in cell_text:
                                    priority = "Low"
                            break
                    
                    # Si aún no se ha encontrado la prioridad, buscar por clases de color
                    if priority == "N/A":
                        color_classes = [
                            (By.XPATH, ".//*[contains(@class, 'red') or contains(@style, 'red')]", "High"),
                            (By.XPATH, ".//*[contains(@class, 'yellow') or contains(@style, 'yellow')]", "Medium"),
                            (By.XPATH, ".//*[contains(@class, 'green') or contains(@style, 'green')]", "Low"),
                            (By.XPATH, ".//*[contains(@class, 'orange') or contains(@style, 'orange')]", "High")
                        ]
                        
                        for locator, indicator_text in color_classes:
                            elements = row.find_elements(locator)
                            if elements:
                                priority = indicator_text
                                break
                except Exception as e:
                    logger.debug(f"Error al extraer prioridad: {e}")
                
                # Obtener todas las celdas de la fila
                try:
                    # Intentar diferentes métodos para obtener celdas
                    cells = []
                    cell_extractors = [
                        lambda r: r.find_elements(By.XPATH, ".//div[@role='gridcell']"),
                        lambda r: r.find_elements(By.XPATH, ".//td"),
                        lambda r: r.find_elements(By.XPATH, "./div")
                    ]
                    
                    for extractor in cell_extractors:
                        try:
                            extracted_cells = extractor(row)
                            if extracted_cells and len(extracted_cells) > 1:
                                cells = extracted_cells
                                break
                        except:
                            continue

                    # Capturar celdas individuales según columna y su posición correcta
                    if cells:
                        # Type - Columna 2
                        if len(cells) >= 2:
                            cell_text = cells[1].text.strip()
                            if cell_text:
                                type_text = cell_text
                        
                        # Priority - Columna 3
                        if len(cells) >= 3 and priority == "N/A":
                            cell_text = cells[2].text.strip()
                            if cell_text:
                                if "high" in cell_text.lower():
                                    if "very" in cell_text.lower():
                                        priority = "Very High"
                                    else:
                                        priority = "High"
                                elif "medium" in cell_text.lower():
                                    priority = "Medium"
                                elif "low" in cell_text.lower():
                                    priority = "Low"
                        
                        # Status - Columna 4
                        if len(cells) >= 4:
                            status_text = cells[3].text.strip()
                            if status_text:
                                # Limpiar el texto para extraer solo el estado
                                status_lines = status_text.split("\n")
                                if status_lines:
                                    # Tomar la primera línea o buscar palabras clave como OPEN, DONE, etc.
                                    for line in status_lines:
                                        if any(keyword in line.upper() for keyword in ["OPEN", "DONE", "READY", "DRAFT", "IN PROGRESS"]):
                                            status = line.strip()
                                            break
                                    else:
                                        status = status_lines[0].strip()
                            
                            # Si no encontró texto directo, buscar por clase específica
                            if status == "N/A":
                                status_elements = cells[3].find_elements(By.XPATH, 
                                    ".//div[contains(@class, 'status')] | " +
                                    ".//span[contains(@class, 'status')]")
                                if status_elements:
                                    status = status_elements[0].text.strip()
                        
                        # Deadline - Columna 5
                        if len(cells) >= 5:
                            deadline_text = cells[4].text.strip()
                            if deadline_text:
                                # Buscar palabras clave de estado
                                deadline_lines = deadline_text.split("\n")
                                for line in deadline_lines:
                                    if any(keyword in line.upper() for keyword in ["OPEN", "DONE", "READY", "DRAFT"]):
                                        deadline = line.strip()
                                        break
                                else:
                                    deadline = deadline_text
                        
                        # Due Date - Columna 6
                        if len(cells) >= 6:
                            due_date_text = cells[5].text.strip()
                            if due_date_text:
                                due_date = due_date_text
                        
                        # Created By - Columna 7
                        if len(cells) >= 7:
                            created_by_text = cells[6].text.strip()
                            if created_by_text:
                                created_by = created_by_text
                                
                                # Limpiar el texto si contiene múltiples líneas
                                if "\n" in created_by:
                                    created_by_lines = created_by.split("\n")
                                    for line in created_by_lines:
                                        if line.startswith("I") and len(line) <= 10:
                                            created_by = line  # Probablemente un ID de usuario
                                            break
                        
                        # Created On - Columna 8
                        if len(cells) >= 8:
                            created_on_text = cells[7].text.strip()
                            if created_on_text:
                                created_on = created_on_text
                                
                                # Limpiar la fecha si tiene formato largo
                                if "," in created_on:
                                    date_parts = created_on.split(",")
                                    if len(date_parts) > 1:
                                        created_on = ",".join(date_parts[-2:]).strip()
                except Exception as cell_e:
                    logger.debug(f"Error al extraer celdas: {cell_e}")
                
                # Si no se pudo extraer el status de las celdas, intentar un enfoque más genérico
                if status == "N/A":
                    try:
                        # Buscar elementos con clases o textos específicos de estado
                        status_texts = [
                            "OPEN", "DONE", "READY FOR PUBLISHING", "IN PROGRESS", "CLOSED", "ACCEPTED", "DRAFT"
                        ]
                        
                        # 1. Buscar texto exacto
                        for status_text in status_texts:
                            status_elements = row.find_elements(
                                By.XPATH, 
                                f".//div[contains(text(), '{status_text}')] | .//span[contains(text(), '{status_text}')]"
                            )
                            
                            if status_elements:
                                status = status_text
                                break
                        
                        # 2. Buscar por clases relacionadas con estado
                        if status == "N/A":
                            status_class_elements = row.find_elements(
                                By.XPATH,
                                ".//*[contains(@class, 'status') or contains(@class, 'state')]"
                            )
                            
                            if status_class_elements:
                                for element in status_class_elements:
                                    if element.text.strip():
                                        status = element.text.strip()
                                        break
                    except Exception as status_e:
                        logger.debug(f"Error al buscar status específico: {status_e}")
                
                # Limpiar los datos si tienen formato inconsistente
                
                # Limpiar el estado
                if status != "N/A" and "\n" in status:
                    status_lines = status.split("\n")
                    # Buscar palabras clave de estado en las líneas
                    for line in status_lines:
                        if any(keyword in line.upper() for keyword in ["OPEN", "DONE", "READY", "DRAFT", "IN PROGRESS", "ACCEPTED"]):
                            status = line.strip()
                            break
                    else:
                        status = status_lines[0].strip()
                
                # Limpiar la prioridad
                if priority != "N/A" and "\n" in priority:
                    priority_lines = priority.split("\n")
                    # Buscar palabras clave de prioridad en las líneas
                    for line in priority_lines:
                        if any(keyword in line.lower() for keyword in ["high", "medium", "low"]):
                            priority = line.strip()
                            break
                    else:
                        priority = priority_lines[0].strip()
                
                # Datos del issue completos
                issue_data = {
                    'Title': title,
                    'Type': type_text,
                    'Priority': priority,
                    'Status': status,
                    'Deadline': deadline,
                    'Due Date': due_date,
                    'Created By': created_by,
                    'Created On': created_on
                }
                
                issues_data.append(issue_data)
                
                processed_count += 1
                
                if processed_count % 10 == 0:
                    logger.info(f"Procesados {processed_count} issues hasta ahora")
                    
                    # Actualizar la interfaz si existe
                    if hasattr(self, 'status_var') and self.status_var:
                        self.status_var.set(f"Procesados {processed_count} de {len(rows)} issues...")
            except Exception as e:
                logger.error(f"Error al procesar la fila {index}: {e}")
        
        logger.info(f"Procesamiento de filas completado. Total procesado: {processed_count} issues")
        return issues_data
    
    
    
    
#########################################
    # MÉTODOS DE PROCESO PRINCIPAL
    #########################################

    def run_extraction(self):
        """Ejecuta el proceso completo de extracción"""
        try:
            if not self.connect_to_browser():
                logger.error("Error al conectar con el navegador")
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set("Error al conectar con el navegador")
                    
                return False
                
            # Navegar automáticamente a la URL
            logger.info("Navegando automáticamente a la URL de SAP...")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Navegando a la URL de SAP...")
                
            self.driver.get("https://xalm-prod.x.eu20.alm.cloud.sap/launchpad#sdwork-center&/")
            
            # Esperar a que se cargue la página
            time.sleep(5)
            
            # Intentar aceptar certificados si aparece el diálogo
            try:
                ok_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'OK') or contains(text(), 'Ok')]"))
                )
                ok_button.click()
                logger.info("Se hizo clic en el botón OK del certificado")
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set("Certificado aceptado. Siga las instrucciones")
            except:
                logger.info("No se encontró diálogo de certificado o ya fue aceptado")
                
            # Si estamos en la GUI, el usuario debe seguir las instrucciones en pantalla
            if self.root:
                instructions = """
                Por favor, realice los siguientes pasos:
                
                1. Inicie sesión si es necesario
                2. Haga clic en 'Project Overview'
                3. Seleccione el cliente con ERP Number: {}
                4. Seleccione el proyecto con ID: {}
                5. Navegue a la pestaña 'Issues'
                
                Una vez en la tabla de issues, haga clic en 'Iniciar Extracción'
                """.format(self.client_var.get() if self.client_var else "1025541", 
                        self.project_var.get() if self.project_var else "20096444")
                
                messagebox.showinfo("Instrucciones de Navegación", instructions)
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set("Esperando a que el usuario navegue a la tabla de issues...")
                    
                # En la GUI, no continuamos automáticamente - esperar a que el usuario haga clic en un botón
                return True
            else:
                # En modo consola, mostrar instrucciones y esperar ENTER
                print("\n=== INSTRUCCIONES PARA NAVEGACIÓN ASISTIDA ===")
                print("Se ha navegado automáticamente a la URL de SAP.")
                print("Por favor, siga estos pasos:")
                print("   1. Inicie sesión si es necesario")
                print("   2. Haga clic en 'Project Overview'")
                print("   3. Seleccione el cliente con ERP Number: 1025541")
                print("   4. Seleccione el proyecto con ID: 20096444")
                print("   5. Navegue a la pestaña 'Issues'")
                
                print("\nUna vez que esté viendo la lista de issues,")
                input("presione ENTER para comenzar la extracción automática...\n")
                
                # Continuar con la extracción
                return self.perform_extraction()
                
        except Exception as e:
            logger.error(f"Error en el proceso de extracción: {e}")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error: {e}")
                
            return False
    def perform_extraction(self):
        """
        Método principal mejorado para ejecutar el proceso de extracción
        Incluye manejo de errores y reintentos con detección mejorada
        """
        try:
            # Marcar como procesando
            self.processing = True
            
            logger.info("Comenzando proceso de extracción...")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Comenzando proceso de extracción...")
                if self.root:
                    self.root.update()
            
            # MEJORA: Verificación más flexible y robusta de si estamos en la página correcta
            in_issues_page = False
            
            # Estrategia 1: Buscar el texto "Issues (número)" que indica que estamos en la página correcta
            try:
                issues_title_elements = self.driver.find_elements(
                    By.XPATH, 
                    "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                )
                if issues_title_elements:
                    logger.info(f"Página de Issues detectada por título: {issues_title_elements[0].text}")
                    in_issues_page = True
            except Exception as e:
                logger.debug(f"No se pudo detectar título de Issues: {e}")
            
            # Estrategia 2: Verificar si hay filas de datos visibles que parecen issues
            if not in_issues_page:
                issue_rows = self.find_table_rows(highlight=False)
                if len(issue_rows) > 0:
                    logger.info(f"Se detectaron {len(issue_rows)} filas de datos que parecen issues")
                    in_issues_page = True
            
            # Estrategia 3: Verificar si hay encabezados de columna típicos de la tabla de issues
            if not in_issues_page:
                try:
                    column_headers = self.driver.find_elements(
                        By.XPATH,
                        "//div[text()='Title'] | //div[text()='Type'] | //div[text()='Priority'] | //div[text()='Status']"
                    )
                    if len(column_headers) >= 3:  # Si encontramos al menos 3 de los encabezados típicos
                        logger.info(f"Se detectaron encabezados de columna típicos de issues: {len(column_headers)}")
                        in_issues_page = True
                except Exception as e:
                    logger.debug(f"No se pudieron detectar encabezados de columna: {e}")
            
            # Si aún no estamos seguros, intentamos hacer clic en la pestaña Issues
            if not in_issues_page:
                logger.warning("No se detectó la página de Issues. Intentando hacer clic en la pestaña...")
                
                # Lista de selectores para intentar encontrar la pestaña Issues
                tab_selectors = [
                    "//div[@role='tab' and contains(text(), 'Issues')]",
                    "//a[contains(text(), 'Issues')]",
                    "//li[contains(@class, 'tab') and contains(., 'Issues')]",
                    "//div[contains(@class, 'sapMITBItem') and contains(., 'Issues')]",
                    # Selector específico para la estructura que veo en tu captura
                    "//div[contains(@class, 'sapMITBItem')]//span[contains(text(), 'Issues')]/..",
                    # Intento por texto directo
                    "//*[contains(text(), 'Issues') and not(contains(text(), '('))]"
                ]
                
                issue_tab_found = False
                for selector in tab_selectors:
                    try:
                        issue_tabs = self.driver.find_elements(By.XPATH, selector)
                        if issue_tabs:
                            # Intentar hacer clic con diferentes métodos
                            for tab in issue_tabs:
                                try:
                                    # Intentar scroll hacia el elemento
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                                    time.sleep(1)
                                    
                                    # Intentar clic con JavaScript
                                    try:
                                        self.driver.execute_script("arguments[0].click();", tab)
                                        logger.info(f"Clic en pestaña Issues realizado con JavaScript: {tab.text}")
                                        time.sleep(3)  # Esperar a que cargue
                                        issue_tab_found = True
                                        break
                                    except:
                                        # Intentar clic normal
                                        tab.click()
                                        logger.info(f"Clic en pestaña Issues realizado: {tab.text}")
                                        time.sleep(3)  # Esperar a que cargue
                                        issue_tab_found = True
                                        break
                                except Exception as click_e:
                                    logger.debug(f"Error al hacer clic en pestaña: {click_e}")
                                    continue
                            
                            if issue_tab_found:
                                break
                    except Exception as e:
                        logger.debug(f"Error con selector {selector}: {e}")
                
                # Verificar nuevamente si estamos en la página de Issues después de intentar clic
                if issue_tab_found:
                    # Verificar nuevamente con las estrategias anteriores
                    try:
                        issues_title_elements = self.driver.find_elements(
                            By.XPATH, "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                        )
                        if issues_title_elements:
                            logger.info(f"Página de Issues detectada después de clic: {issues_title_elements[0].text}")
                            in_issues_page = True
                    except:
                        pass
            
            # ALTERNATIVA: Si todo falla, continuar de todos modos y confiar en que estamos en la página correcta
            if not in_issues_page:
                logger.warning("No se pudo confirmar que estamos en la página de Issues, pero intentaremos extraer datos de todos modos.")
            
            # Realizar la extracción de datos
            max_attempts = 3
            for attempt in range(max_attempts):
                try:
                    logger.info(f"Intento de extracción {attempt+1}/{max_attempts}")
                    
                    # Actualizar la interfaz si existe
                    if hasattr(self, 'status_var') and self.status_var:
                        self.status_var.set(f"Intento de extracción {attempt+1}/{max_attempts}...")
                    
                    issues_data = self.extract_issues_data()
                    
                    if issues_data:
                        logger.info(f"Extracción exitosa: {len(issues_data)} issues encontrados")
                        
                        # Actualizar Excel con los datos extraídos
                        if self.excel_file_path:
                            self.update_excel(issues_data)
                        else:
                            logger.warning("No se ha seleccionado archivo Excel para guardar los datos")
                            
                            if hasattr(self, 'status_var') and self.status_var:
                                self.status_var.set("Advertencia: No se ha seleccionado archivo Excel")
                            
                            if self.root:
                                excel_path = self.choose_excel_file()
                                if excel_path:
                                    self.update_excel(issues_data)
                        
                        self.processing = False
                        return True
                    else:
                        logger.warning(f"No se encontraron issues en el intento {attempt+1}")
                        
                        # Si no es el último intento, esperar y reintentar
                        if attempt < max_attempts - 1:
                            logger.info("Esperando antes de reintentar...")
                            time.sleep(5)
                        else:
                            logger.error("Todos los intentos de extracción fallaron")
                            
                            if hasattr(self, 'status_var') and self.status_var:
                                self.status_var.set("Error: No se encontraron issues después de varios intentos")
                            
                            if self.root:
                                messagebox.showerror(
                                    "Error de Extracción", 
                                    "No se pudieron encontrar issues después de varios intentos. Verifique que está en la página correcta y que existen issues para extraer."
                                )
                            
                            self.processing = False
                            return False
                except Exception as e:
                    logger.error(f"Error en el intento {attempt+1}: {e}")
                    
                    # Si no es el último intento, esperar y reintentar
                    if attempt < max_attempts - 1:
                        logger.info("Esperando antes de reintentar...")
                        time.sleep(5)
                    else:
                        logger.error(f"Todos los intentos de extracción fallaron: {e}")
                        
                        if hasattr(self, 'status_var') and self.status_var:
                            self.status_var.set(f"Error de extracción: {e}")
                        
                        if self.root:
                            messagebox.showerror(
                                "Error de Extracción", 
                                f"Se produjo un error durante la extracción: {e}"
                            )
                        
                        self.processing = False
                        return False
            
            # Si llegamos aquí, todos los intentos fallaron
            logger.error("Extracción fallida después de varios intentos")
            
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Error: Extracción fallida")
            
            self.processing = False
            return False
            
        except Exception as e:
            logger.error(f"Error general en el proceso de extracción: {e}")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error: {e}")
            
            self.processing = False
            return False
            
        except Exception as e:
            logger.error(f"Error general en el proceso de extracción: {e}")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error: {e}")
            
            self.processing = False
            return False


#########################################
    # MÉTODOS PARA LA INTERFAZ GRÁFICA
    #########################################

    def create_gui(self):
        """Crea una interfaz gráfica mejorada para la aplicación"""
        self.root = tk.Tk()
        self.root.title("SAP Recommendations Extractor")
        self.root.geometry("650x800")
        self.root.resizable(True, True)
        self.root.configure(bg=SAP_COLORS["light"])
        
        # Configurar icono de la aplicación
        logo_photo = None
        if 'PIL_AVAILABLE' in globals() and PIL_AVAILABLE:
            try:
                # Logo SAP como icono base64
                sap_logo_base64 = """
                iVBORw0KGgoAAAANSUhEUgAAADAAAAAwCAYAAABXAvmHAAAACXBIWXMAAAsTAAALEwEAmpwYAAAB
                g0lEQVR4nO2ZQU7DMBBFv1mcJbAhB4ADcIucgAsAYslpWKQSKnCHZoPEgnuAaFasqESrKE3qicdJ
                +0kjVXXi+T8ex3ZSoCiKotTDEMAUwBxAUrJmXjsCaJR94IeSnbXCVgvAPYAvQ/LWWnjtDUDXdjJr
                A5jvwFTZoWnVBrAImFhZFgBOygxx6VkOytzIYRpZ5jEGOA0sbwy70NTGQHcHkltfN7IxUBTYKPPY
                3kKriMmN8gxMYqLjPY6JdrjEk/SaiGqigYGixGSiYUNI3bQnZZ8npiamCiCDiZaeBbQ/LWBMfAJ4
                5LU7AF9oo0yyiW/eN36QxQvXL4t5YlkxDZwBuAbwxsOI18a89sXrFwDOY05i7J3iCUCnYK6T7x1j
                9A0MPPoGlhLb1jHPdLgmXw7uOU++tg88Lrn+w/UrwvVKcjuxLaEYA5L1VmcQs4UoO9TGLDPJeiuZ
                qDKAxNTaQMwea227kDQT
                """
                
                # Cargar icono desde base64
                logo_data = base64.b64decode(sap_logo_base64)
                logo_image = Image.open(BytesIO(logo_data))
                
                # Convertir a formato para tkinter
                logo_photo = ImageTk.PhotoImage(logo_image)
                
                # Establecer como icono de la ventana
                self.root.iconphoto(True, logo_photo)
            except Exception as e:
                logger.error(f"Error al cargar el icono: {e}")
            
        # Configurar estilo
        style = ttk.Style()
        style.configure('TCombobox', arrowsize=15)  # Aumentar tamaño de flecha
        
        # Estilo general para todos los widgets con texto negro
        style.configure(".", 
                        foreground=SAP_COLORS["text"])  # Texto negro para todo por defecto
        
        # Estilo para etiquetas con texto negro sobre fondo claro
        style.configure("TLabel", 
                        background=SAP_COLORS["light"],
                        foreground=SAP_COLORS["text"],  # NEGRO para mejor contraste
                        font=("Arial", 10, "bold"))     # Texto en negrita para mejor visibilidad
        
        # Estilo para etiquetas de título con texto azul oscuro visible
        style.configure("Header.TLabel", 
                        background=SAP_COLORS["light"],
                        foreground=SAP_COLORS["secondary"],  # Azul oscuro para títulos
                        font=("Arial", 16, "bold"))
        
        # Estilo para frames con texto negro
        style.configure("TLabelframe.Label", 
                        background=SAP_COLORS["light"],
                        foreground=SAP_COLORS["text"],  # Texto negro
                        font=("Arial", 11, "bold"))
                        
        # Estilo para botones de navegador
        style.configure("Browser.TButton", 
                        background=SAP_COLORS["primary"],
                        foreground="white",  # Color de texto blanco 
                        font=("Arial", 10, "bold"))

        # Estilo para botones de Excel
        style.configure("Excel.TButton", 
                        background=SAP_COLORS["success"],
                        foreground="white",  # Color de texto blanco
                        font=("Arial", 10, "bold"))

        # Estilo para botones de acción  
        style.configure("Action.TButton", 
                        background=SAP_COLORS["warning"],
                        foreground="white",  # Color de texto blanco
                        font=("Arial", 10, "bold"))
        
        style.configure("Primary.TButton", 
                        background=SAP_COLORS["primary"], 
                        foreground=SAP_COLORS["white"],
                        font=("Arial", 10, "bold"))
        
        style.configure("Success.TButton", 
                        background=SAP_COLORS["success"],
                        foreground=SAP_COLORS["white"],
                        font=("Arial", 10, "bold"))
        
        style.configure("Danger.TButton", 
                        background=SAP_COLORS["danger"],
                        foreground=SAP_COLORS["white"],
                        font=("Arial", 10, "bold"))
        
        # Estilo para combos con texto negro
        style.configure("TCombobox", 
                        selectbackground=SAP_COLORS["primary"],
                        selectforeground=SAP_COLORS["white"],
                        fieldbackground="white",
                        background="white",
                        foreground=SAP_COLORS["text"])  # Texto negro en combos
        
        # Estilo para entradas de texto con texto negro
        style.configure("TEntry", 
                        fieldbackground="white",
                        foreground=SAP_COLORS["text"],  # Texto negro
                        font=("Arial", 10))
        
        # Para garantizar que el texto sea visible en entradas configuramos un estilo específico
        style.configure(
            "Custom.TEntry",
            fieldbackground="white",
            foreground="black",
            insertcolor="black",     # Color del cursor
            bordercolor=SAP_COLORS["primary"],
            lightcolor=SAP_COLORS["primary"],
            darkcolor=SAP_COLORS["primary"]
        )
        
        # Estilo personalizado para los combobox para asegurar visibilidad del texto
        style.configure(
            "Custom.TCombobox",
            fieldbackground="white",  # Fondo blanco
            background="white",       # Fondo blanco
            foreground="black",       # Texto negro
            arrowcolor=SAP_COLORS["primary"],  # Flecha azul
            bordercolor=SAP_COLORS["primary"]  # Borde azul
        )
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Cabecera con logo
        self.header_frame = ttk.Frame(main_frame)
        self.header_frame.pack(fill=tk.X, pady=(0, 15))

        try:
            # Usar el logo si ya está cargado y PIL está disponible
            if 'PIL_AVAILABLE' in globals() and PIL_AVAILABLE and logo_photo:
                logo_label = tk.Label(self.header_frame, image=logo_photo, bg=SAP_COLORS["light"])
                logo_label.image = logo_photo  # Mantener referencia
                logo_label.pack(side=tk.LEFT, padx=(0, 10))
        except:
            pass
        
        # Título con fondo de alto contraste
        title_background = "#0A3D6E"  # Azul oscuro
        title_foreground = "#FFFFFF"  # Texto blanco
        
        title_label = tk.Label(
            self.header_frame, 
            text="Extractor de Recomendaciones SAP",
            font=("Arial", 18, "bold"),
            foreground=title_foreground,
            background=title_background,
            padx=8,
            pady=4
        )
        title_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Contenedor principal dividido
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Panel izquierdo para configuración
        self.left_panel = ttk.Frame(content_frame, padding=10, width=435)
        self.left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        self.left_panel.pack_propagate(False)  # Mantener ancho fijo
        
        # Sección de cliente
        client_frame = tk.LabelFrame(self.left_panel, 
                                  text="Cliente", 
                                  bg=SAP_COLORS["light"],
                                  fg="#000000",  # TEXTO NEGRO para el título
                                  font=("Arial", 11, "bold"),
                                  padx=10, pady=10)
        client_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta ERP con texto negro explícito
        tk.Label(client_frame, 
               text="ERP Number:",
               bg=SAP_COLORS["light"],
               fg="#000000",  # NEGRO explícito
               font=("Arial", 9)).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Entry con fondo blanco y texto negro
        self.client_var = tk.StringVar(value="1025541")
        client_entry = tk.Entry(client_frame, 
                              textvariable=self.client_var,
                              width=15,
                              font=("Arial", 10),
                              bg="white",
                              fg="black",  # TEXTO NEGRO para la entrada
                              highlightbackground=SAP_COLORS["primary"],  # Borde cuando no tiene foco
                              highlightcolor=SAP_COLORS["primary"])       # Borde cuando tiene foco
        client_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        # Etiqueta de clientes guardados con texto negro
        tk.Label(client_frame, 
               text="Clientes guardados:",
               bg=SAP_COLORS["light"],
               fg="#000000",  # NEGRO explícito
               font=("Arial", 9)).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Lista desplegable de clientes guardados
        client_list = self.get_clients_from_db()
        self.client_combo = ttk.Combobox(client_frame, values=client_list, width=30)
        self.client_combo.config(state='readonly')
        self.client_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W+tk.E)
        self.client_combo.bind("<<ComboboxSelected>>", lambda e: self.select_client(self.client_combo.get()))

        # Sección de proyecto
        project_frame = tk.LabelFrame(self.left_panel, 
                                    text="Proyecto", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",  # TEXTO NEGRO para el título
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
        project_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta ID con texto negro explícito
        tk.Label(project_frame, 
               text="ID Proyecto:",
               bg=SAP_COLORS["light"],
               fg="#000000",  # NEGRO explícito
               font=("Arial", 10)).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Entry con fondo blanco y texto negro
        self.project_var = tk.StringVar(value="20096444")
        project_entry = tk.Entry(project_frame, 
                               textvariable=self.project_var,
                               width=15,
                               font=("Arial", 10),
                               bg="white",
                               fg="black",  # TEXTO NEGRO para la entrada
                               highlightbackground=SAP_COLORS["primary"],  # Borde cuando no tiene foco
                               highlightcolor=SAP_COLORS["primary"])       # Borde cuando tiene foco
        project_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        # Etiqueta de proyectos con texto negro
        tk.Label(project_frame, 
               text="Proyectos:",
               bg=SAP_COLORS["light"],
               fg="#000000",  # NEGRO explícito
               font=("Arial", 10)).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Lista desplegable de proyectos guardados
        project_list = self.get_projects_from_db("1025541")  # Proyectos para el cliente predeterminado
        self.project_combo = ttk.Combobox(project_frame, values=project_list, width=30)
        self.project_combo.config(state='readonly')
        self.project_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W+tk.E)
        self.project_combo.bind("<<ComboboxSelected>>", lambda e: self.select_project(self.project_combo.get()))

        # Sección de navegador
        browser_frame = tk.LabelFrame(self.left_panel, 
                                    text="Navegador", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",  # TEXTO NEGRO para el título
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
        browser_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta con texto negro explícito
        browser_label = tk.Label(
            browser_frame, 
            text="Iniciar un navegador con perfil dedicado:",
            bg=SAP_COLORS["light"],
            fg="#000000",  # NEGRO explícito
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        browser_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón con colores de alto contraste
        browser_button = tk.Button(
            browser_frame, 
            text="Iniciar Navegador",
            command=self.start_browser,
            bg=SAP_COLORS["primary"],
            fg="#FFFFFF",  # TEXTO BLANCO
            activebackground="#0A3D6E",  # Azul más oscuro al hacer clic
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        browser_button.pack(fill=tk.X, pady=5)
        
        # Sección de archivo Excel
        excel_frame = tk.LabelFrame(self.left_panel, 
                                  text="Archivo Excel", 
                                  bg=SAP_COLORS["light"],
                                  fg="#000000",  # TEXTO NEGRO para el título
                                  font=("Arial", 11, "bold"),
                                  padx=10, pady=10)
        excel_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta con texto negro explícito
        excel_label = tk.Label(
            excel_frame, 
            text="Seleccione un archivo existente o cree uno nuevo:",
            bg=SAP_COLORS["light"],
            fg="#000000",  # NEGRO explícito
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        excel_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón con colores de alto contraste
        excel_button = tk.Button(
            excel_frame, 
            text="Seleccionar o Crear Excel",
            command=self.choose_excel_file,
            bg=SAP_COLORS["success"],
            fg="#FFFFFF",  # TEXTO BLANCO
            activebackground="#085E2E",  # Verde más oscuro al hacer clic
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        excel_button.pack(fill=tk.X, pady=5)
        
        # Mostrar el nombre del archivo seleccionado
        self.excel_filename_var = tk.StringVar(value="No seleccionado")
        excel_file_label = tk.Label(
            excel_frame, 
            textvariable=self.excel_filename_var,
            bg=SAP_COLORS["light"],
            fg="#0A3D6E",  # Azul oscuro para destacar
            font=("Arial", 9, "bold"),
            wraplength=200,
            anchor="w",
            justify="left"
        )
        excel_file_label.pack(fill=tk.X, pady=5)
        
        # Sección de acción
        action_frame = tk.LabelFrame(self.left_panel, 
                                   text="Acciones", 
                                   bg=SAP_COLORS["light"],
                                   fg="#000000",  # TEXTO NEGRO para el título
                                   font=("Arial", 11, "bold"),
                                   padx=10, pady=10)
        action_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta con texto negro explícito
        action_label = tk.Label(
            action_frame, 
            text="Extraer datos de issues desde SAP:",
            bg=SAP_COLORS["light"],
            fg="#000000",  # NEGRO explícito
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        action_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón de extracción con colores de alto contraste
        extract_button = tk.Button(
            action_frame, 
            text="Iniciar Extracción de Issues",
            command=self.start_extraction,
            bg=SAP_COLORS["warning"],
            fg="#FFFFFF",  # TEXTO BLANCO
            activebackground="#C25A00",  # Naranja más oscuro al hacer clic
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        extract_button.pack(fill=tk.X, pady=5)
        
        # Separador visual
        separator = tk.Frame(action_frame, height=2, bg=SAP_COLORS["gray"])
        separator.pack(fill=tk.X, pady=10)
        
        # Botón de salir con colores de alto contraste
        exit_button = tk.Button(
            action_frame, 
            text="Salir de la Aplicación",
            command=self.exit_app,
            bg=SAP_COLORS["danger"],
            fg="#FFFFFF",  # TEXTO BLANCO
            activebackground="#990000",  # Rojo más oscuro al hacer clic
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        exit_button.pack(fill=tk.X, pady=5)
        
        # Panel derecho para logs
        right_panel = ttk.Frame(content_frame, padding=10, width=300)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        right_panel.pack_propagate(False)  # Agregar esta línea para mantener ancho fijo
        
        # Log frame
        log_frame = tk.LabelFrame(right_panel, 
                               text="Registro de Actividad", 
                               bg=SAP_COLORS["light"],
                               fg="#000000",  # TEXTO NEGRO para el título
                               font=("Arial", 11, "bold"))
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        # Text widget para logs con colores de sintaxis
        self.log_text = tk.Text(
            log_frame, 
            height=20, 
            wrap=tk.WORD, 
            bg="white",             # Fondo blanco
            fg="black",             # Texto negro
            font=("Consolas", 9),
            padx=5,
            pady=5,
            borderwidth=2,
            relief=tk.SUNKEN
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Colores para los logs
        self.log_text.tag_configure("INFO", foreground="black")
        self.log_text.tag_configure("WARNING", foreground="#CC6600")  # Naranja más oscuro
        self.log_text.tag_configure("ERROR", foreground="#990000")    # Rojo más oscuro
        self.log_text.tag_configure("DEBUG", foreground="#555555")    # Gris oscuro
        
        # Scrollbar para el log
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Status bar con texto negro
        self.status_var = tk.StringVar(value="Listo para iniciar")
        status_bar = tk.Label(
            self.root, 
            textvariable=self.status_var,
            fg="#000000",           # Negro
            bg="#F0F0F0",           # Gris muy claro
            relief=tk.SUNKEN, 
            anchor=tk.W, 
            padx=5,
            pady=2,
            font=("Arial", 10)
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Configurar logger para que también escriba en la GUI
        self.setup_gui_logger()
        
        self.root.protocol("WM_DELETE_WINDOW", self.exit_app)
        
        # Centrar la ventana en la pantalla
        self.root.update_idletasks()  # Actualiza info de geometría
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        size = tuple(int(_) for _ in self.root.geometry().split('+')[0].split('x'))
        x = screen_width/2 - size[0]/2
        y = screen_height/2 - size[1]/2
        self.root.geometry("%dx%d+%d+%d" % (size + (x, y)))

    def setup_gui_logger(self):
        """Configura el logger para que también escriba en la GUI"""
        class TextHandler(logging.Handler):
            def __init__(self, text_widget):
                logging.Handler.__init__(self)
                self.text_widget = text_widget
            
            def emit(self, record):
                msg = self.format(record)
                def append():
                    self.text_widget.configure(state='normal')
                    
                    # Agregar marca de tiempo y nivel con color
                    time_str = msg.split(' - ')[0] + ' - '
                    level_str = record.levelname + ' - '
                    msg_content = msg.split(' - ', 2)[2] if len(msg.split(' - ')) > 2 else ""
                    
                    self.text_widget.insert(tk.END, time_str, "INFO")
                    self.text_widget.insert(tk.END, level_str, record.levelname)
                    self.text_widget.insert(tk.END, msg_content + '\n', record.levelname)
                    
                    self.text_widget.configure(state='disabled')
                    self.text_widget.yview(tk.END)
                    
                # Llamar a append desde el hilo principal
                self.text_widget.after(0, append)
        
        # Crear un handler que escriba en el widget Text
        text_handler = TextHandler(self.log_text)
        text_handler.setLevel(logging.INFO)
        text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        
        # Añadir el handler al logger
        logger.addHandler(text_handler)
        
        # Deshabilitar el widget para que el usuario no pueda editar el texto
        self.log_text.configure(state='disabled')
        
    def select_client(self, client_string):
            """Maneja la selección de un cliente desde el combobox"""
            try:
                # Extraer el ERP number del string "1025541 - Nombre del cliente"
                erp_number = client_string.split(" - ")[0]
                self.client_var.set(erp_number)
                
                # Actualizar la lista de proyectos para este cliente
                projects = self.get_projects_from_db(erp_number)
                self.project_combo['values'] = projects
                
                if projects:
                    self.project_combo.current(0)
                    self.select_project(projects[0])
                    
                # Actualizar el uso de este cliente
                self.update_client_usage(erp_number)
                
                logger.info(f"Cliente seleccionado: {client_string}")
            except Exception as e:
                logger.error(f"Error al seleccionar cliente: {e}")

    def select_project(self, project_string):
        """Maneja la selección de un proyecto desde el combobox"""
        try:
            # Extraer el ID del proyecto del string "20096444 - Nombre del proyecto"
            project_id = project_string.split(" - ")[0]
            self.project_var.set(project_id)
            
            # Actualizar el uso de este proyecto
            self.update_project_usage(project_id)
            
            logger.info(f"Proyecto seleccionado: {project_string}")
        except Exception as e:
            logger.error(f"Error al seleccionar proyecto: {e}")

    def start_browser(self):
        """Inicia el navegador desde la interfaz gráfica"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                messagebox.showwarning("Proceso en curso", "Hay un proceso de extracción en curso.")
                return
                
            # Asegurarse de que no haya un navegador ya abierto
            if self.driver:
                messagebox.showinfo("Navegador ya iniciado", "El navegador ya está iniciado.")
                return
                
            # Actualizar la interfaz para mostrar que se está iniciando el navegador
            self.status_var.set("Iniciando navegador...")
            if self.root:
                self.root.update()
                
            # Iniciar el navegador
            if self.connect_to_browser():
                logger.info("Navegador iniciado")
                self.status_var.set("Navegador iniciado. Inicie la extracción cuando esté listo.")
                
                # Navegar a la URL de SAP
                self.driver.get("https://xalm-prod.x.eu20.alm.cloud.sap/launchpad#sdwork-center&/projects")
                
                # Mostrar instrucciones
                instructions = """
                Por favor, realice los siguientes pasos:
                
                1. Inicie sesión si es necesario
                2. Haga clic en 'Project Overview'
                3. Seleccione el cliente con ERP Number: {}
                4. Seleccione el proyecto con ID: {}
                5. Navegue a la pestaña 'Issues'
                
                Una vez en la tabla de issues, haga clic en 'Iniciar Extracción'
                """.format(self.client_var.get(), self.project_var.get())
                
                messagebox.showinfo("Instrucciones de Navegación", instructions)
            else:
                self.status_var.set("Error al iniciar el navegador")
                messagebox.showerror("Error", "No se pudo iniciar el navegador. Revise el log para más detalles.")
        except Exception as e:
            logger.error(f"Error al iniciar el navegador: {e}")
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", f"Error al iniciar el navegador: {e}")

    def start_extraction(self):
        """Inicia el proceso de extracción desde la interfaz gráfica"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                messagebox.showwarning("Proceso en curso", "Hay un proceso de extracción en curso.")
                return
                
            # Verificar que existe un archivo Excel seleccionado
            if not self.excel_file_path:
                messagebox.showwarning("Archivo Excel no seleccionado", "Debe seleccionar o crear un archivo Excel primero.")
                return
                
            # Verificar que el navegador está abierto
            if not self.driver:
                messagebox.showwarning("Navegador no iniciado", "Debe iniciar el navegador primero.")
                return
                
            # Iniciar extracción en un hilo separado para no bloquear la GUI
            threading.Thread(target=self.perform_extraction, daemon=True).start()
            
        except Exception as e:
            logger.error(f"Error al iniciar extracción: {e}")
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", f"Error al iniciar extracción: {e}")

    def exit_app(self):
        """Cierra la aplicación cerrando también el navegador si está abierto"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                confirm_exit = messagebox.askyesno(
                    "Proceso en curso", 
                    "Hay un proceso de extracción en curso. ¿Realmente desea salir?",
                    icon='warning'
                )
                if not confirm_exit:
                    return
                    
            if self.driver:
                try:
                    close_browser = messagebox.askyesno(
                        "Cerrar navegador", 
                        "¿Desea cerrar también el navegador?",
                        icon='question'
                    )
                    if close_browser:
                        self.driver.quit()
                        logger.info("Navegador cerrado correctamente")
                except:
                    logger.warning("No se pudo cerrar el navegador correctamente")
            
            self.root.destroy()
        except Exception as e:
            logger.error(f"Error al cerrar la aplicación: {e}")
            # En caso de error, forzar cierre
            self.root.destroy()

    def main_gui(self):
        """Punto de entrada principal con interfaz gráfica"""
        self.create_gui()
        self.root.mainloop()

    def create_simple_icon(self):
        """Crea un ícono simple usando Tkinter nativo"""
        icon = tk.Canvas(self.root, width=32, height=32, bg=SAP_COLORS["primary"], highlightthickness=0)
        icon.create_text(16, 16, text="SAP", fill="white", font=("Arial", 9, "bold"))
        
        # Guardar referencia para evitar que el recolector de basura lo elimine
        self.icon_canvas = icon
        return icon

    @staticmethod
    def create_shortcut(target_path, shortcut_path=None, icon_path=None):
        """Crea un acceso directo para la aplicación"""
        try:
            if not shortcut_path:
                desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')
                shortcut_path = os.path.join(desktop_path, "SAP Issues Extractor.lnk")
                
            if os.path.exists(shortcut_path):
                return shortcut_path
                
            import winshell
            from win32com.client import Dispatch
            
            shell = Dispatch('WScript.Shell')
            shortcut = shell.CreateShortCut(shortcut_path)
            shortcut.Targetpath = target_path
            shortcut.WorkingDirectory = os.path.dirname(target_path)
            if icon_path:
                shortcut.IconLocation = icon_path
            shortcut.save()
            
            return shortcut_path
        except Exception as e:
            logger.error(f"Error al crear acceso directo: {e}")
            return None


def main():
    """Función principal que ejecuta la aplicación"""
    extractor = None  # Inicializar la variable para evitar el UnboundLocalError
    try:
        # Verificar paquetes requeridos (quitar Pillow de los requeridos)
        required_packages = ["selenium", "pandas", "openpyxl"]
        missing_packages = []
        
        for package in required_packages:
            try:
                __import__(package)
            except ImportError:
                missing_packages.append(package)

        if missing_packages:
            print("Faltan las siguientes bibliotecas necesarias:")
            for package in missing_packages:
                print(f"  - {package}")
            print("\nPor favor, instálalas usando:")
            print(f"pip install {' '.join(missing_packages)}")
            input("\nPresiona ENTER para salir...")
            sys.exit(1)
        
        # Notificar sobre Pillow, pero no detener la ejecución
        try:
            __import__("PIL")
        except ImportError:
            print("Nota: La biblioteca Pillow no está disponible. Algunas características visuales estarán limitadas.")
            print("Si deseas instalarla, ejecuta: pip install Pillow")
        
        # Crear instancia del extractor
        extractor = IssuesExtractor()
        
        # Verificar si se desea interfaz gráfica o consola
        if len(sys.argv) > 1 and sys.argv[1] == "--console":
            # Modo consola
            extractor.choose_excel_file()
            extractor.run_extraction()
        else:
            # Modo interfaz gráfica (predeterminado)
            extractor.main_gui()
            
        logger.info("=== Proceso de extracción finalizado ===")
        
    except Exception as e:
        logger.critical(f"Error crítico en la ejecución: {e}")
        print(f"\n¡ERROR! Se ha producido un error crítico: {e}")
        print(f"Por favor, revisa el archivo de log para más detalles: {log_file}")
    finally:
        if extractor is not None and (not hasattr(extractor, 'root') or not extractor.root):
            # Solo mostrar mensaje final si estamos en modo consola
            input("\nPresiona ENTER para cerrar...")


if __name__ == "__main__":
    main()