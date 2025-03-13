# -*- coding: utf-8 -*-
"""
Script Mejorado para Extracción de Issues SAP con Mejor Manejo de Scroll
---
Versión 16: Optimizado para rendimiento, mejor manejo de excepciones,
y procesamiento eficiente de datos.
"""

import time
import pandas as pd
import os.path
import sys
import logging
import sqlite3
import threading
import re
import json
from datetime import datetime
import webbrowser
import base64
from io import BytesIO

# Intentar importar PIL, pero no fallar si no está disponible
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Selenium imports
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
    JavascriptException,
    WebDriverException
)

# Tkinter imports
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Configurar logging
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

log_file = os.path.join(
    log_dir, f"extraccion_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(log_file), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# Colores estilo SAP
SAP_COLORS = {
    "primary": "#1870C5",    # Azul SAP
    "secondary": "#354A5F",  # Azul oscuro SAP
    "success": "#107E3E",    # Verde SAP
    "warning": "#E9730C",    # Naranja SAP
    "danger": "#BB0000",     # Rojo SAP
    "light": "#F5F6F7",      # Gris claro SAP
    "dark": "#32363A",       # Gris oscuro SAP
    "white": "#FFFFFF",
    "gray": "#D3D7D9",
    "text": "#000000"        # Texto negro para máximo contraste
}










class DatabaseManager:
    """Clase dedicada al manejo de la base de datos de clientes y proyectos"""
    
    def __init__(self, db_path=None):
        """Inicializa el administrador de base de datos"""
        if db_path is None:
            db_dir = "data"
            if not os.path.exists(db_dir):
                os.makedirs(db_dir)
            db_path = os.path.join(db_dir, "sap_extraction.db")
            
        self.db_path = db_path
        self.setup_database()
    
    def setup_database(self):
        """Configura la estructura de la base de datos"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Crear tablas si no existen
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS clients (
            erp_number TEXT PRIMARY KEY,
            name TEXT,
            business_partner TEXT,
            last_used TIMESTAMP
        )
        ''')

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            project_id TEXT PRIMARY KEY,
            client_erp TEXT,
            name TEXT,
            engagement_case TEXT,
            last_used TIMESTAMP,
            FOREIGN KEY (client_erp) REFERENCES clients(erp_number)
        )
        ''')
        
        conn.commit()
        conn.close()
        
    def get_clients(self):
        """Obtiene la lista de clientes ordenados por último uso"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("SELECT erp_number, name FROM clients ORDER BY last_used DESC")
        clients = cursor.fetchall()

        conn.close()

        return [f"{erp} - {name}" for erp, name in clients]
    
    def get_projects(self, client_erp):
        """Obtiene la lista de proyectos para un cliente específico"""
        if not client_erp:
            return []
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT project_id, name 
            FROM projects 
            WHERE client_erp = ? 
            ORDER BY last_used DESC
        """, (client_erp,))

        projects = cursor.fetchall()

        conn.close()

        return [f"{pid} - {name}" for pid, name in projects]
    
    def save_client(self, erp_number, name, business_partner=""):
        """Guarda o actualiza un cliente en la base de datos"""
        if not self.validate_input(erp_number, "erp"):
            logger.error(f"Número ERP inválido: {erp_number}")
            return False
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            # Verificar si el cliente ya existe
            cursor.execute("SELECT * FROM clients WHERE erp_number = ?", (erp_number,))
            existing = cursor.fetchone()

            if existing:
                # Actualizar cliente existente
                cursor.execute("""
                    UPDATE clients 
                    SET name = ?, business_partner = ?, last_used = CURRENT_TIMESTAMP 
                    WHERE erp_number = ?
                """, (name, business_partner, erp_number))
            else:
                # Insertar nuevo cliente
                cursor.execute("""
                    INSERT INTO clients (erp_number, name, business_partner, last_used) 
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                """, (erp_number, name, business_partner))

            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al guardar cliente en BD: {e}")
            return False
        finally:
            conn.close()
    
    def save_project(self, project_id, client_erp, name, engagement_case=""):
        """Guarda o actualiza un proyecto en la base de datos"""
        if not self.validate_input(project_id, "project") or not self.validate_input(client_erp, "erp"):
            logger.error(f"ID de proyecto o cliente inválido: {project_id}, {client_erp}")
            return False
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            # Verificar si el proyecto ya existe
            cursor.execute("SELECT * FROM projects WHERE project_id = ?", (project_id,))
            existing = cursor.fetchone()

            if existing:
                # Actualizar proyecto existente
                cursor.execute("""
                    UPDATE projects 
                    SET client_erp = ?, name = ?, engagement_case = ?, last_used = CURRENT_TIMESTAMP 
                    WHERE project_id = ?
                """, (client_erp, name, engagement_case, project_id))
            else:
                # Insertar nuevo proyecto
                cursor.execute("""
                    INSERT INTO projects (project_id, client_erp, name, engagement_case, last_used) 
                    VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                """, (project_id, client_erp, name, engagement_case))

            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al guardar proyecto en BD: {e}")
            return False
        finally:
            conn.close()
    
    def update_client_usage(self, erp_number):
        """Actualiza la fecha de último uso de un cliente"""
        if not self.validate_input(erp_number, "erp"):
            return False
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("""
                UPDATE clients 
                SET last_used = CURRENT_TIMESTAMP 
                WHERE erp_number = ?
            """, (erp_number,))
            
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al actualizar uso de cliente: {e}")
            return False
        finally:
            conn.close()
    
    def update_project_usage(self, project_id):
        """Actualiza la fecha de último uso de un proyecto"""
        if not self.validate_input(project_id, "project"):
            return False
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("""
                UPDATE projects 
                SET last_used = CURRENT_TIMESTAMP 
                WHERE project_id = ?
            """, (project_id,))
            
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error al actualizar uso de proyecto: {e}")
            return False
        finally:
            conn.close()
    
    @staticmethod
    def validate_input(input_str, input_type="general"):
        """Valida las entradas para prevenir inyecciones SQL"""
        if input_type == "erp":
            # Solo permitir dígitos para ERP
            return bool(re.match(r'^\d+$', str(input_str)))
        elif input_type == "project":
            # Solo permitir dígitos para ID de proyecto
            return bool(re.match(r'^\d+$', str(input_str)))
        elif input_type == "path":
            # Validar ruta de archivo
            return os.path.isabs(input_str) and not any(c in input_str for c in '<>:|?*')
        return True
    
    
    
    
    
    
    
    
    
    
    
    

class ExcelManager:
    """Clase dedicada al manejo de archivos Excel de seguimiento"""
    
    def __init__(self, file_path=None):
        """Inicializa el administrador de Excel"""
        self.file_path = file_path
        
    def select_file(self):
        """Permite al usuario elegir un archivo Excel existente o crear uno nuevo"""
        # Preguntar si desea usar un archivo existente o crear uno nuevo
        choice = messagebox.askquestion(
            "Archivo Excel",
            "¿Desea usar un archivo Excel existente?\n\n"
            + "Seleccione 'Sí' para elegir un archivo existente.\n"
            + "Seleccione 'No' para crear un nuevo archivo.",
            icon='question'
        )

        if choice == "yes":
            # Permitir al usuario seleccionar un archivo Excel existente
            file_path = filedialog.askopenfilename(
                title="Seleccione el archivo Excel de seguimiento",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            )

            if not file_path:  # El usuario canceló la selección
                logger.info("Usuario canceló la selección de archivo. Se creará uno nuevo.")
                self._create_new_file()
        else:
            self._create_new_file()
            
        return self.file_path
    
    def _create_new_file(self):
        """Crea un nuevo archivo Excel para seguimiento"""
        # Crear un nombre de archivo por defecto con fecha y hora
        default_filename = f"Seguimiento_Recomendaciones_EPM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Permitir al usuario guardar con un nombre específico
        file_path = filedialog.asksaveasfilename(
            title="Guardar nuevo archivo Excel",
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Archivos Excel", "*.xlsx")],
        )

        if not file_path:  # Si cancela, usar el nombre por defecto
            file_path = default_filename
            logger.info(f"Se usará el nombre por defecto: {file_path}")

        # Crear un archivo Excel vacío con las columnas necesarias
        self._create_excel_template(file_path)
        logger.info(f"Creado nuevo archivo Excel: {file_path}")
        self.file_path = file_path
        
    def _create_excel_template(self, file_path):
        """Crea la estructura del archivo Excel con las columnas necesarias"""
        try:
            # Crear un DataFrame vacío con las columnas necesarias
            columns = [
                "Title",
                "Type",
                "Priority",
                "Status",
                "Deadline",
                "Due Date",
                "Created By",
                "Created On",
                "Last Updated",
                "Comments",
            ]

            df = pd.DataFrame(columns=columns)

            # Guardar el DataFrame vacío como un archivo Excel
            df.to_excel(file_path, index=False)
            logger.info(f"Archivo Excel creado exitosamente: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error al crear nuevo archivo Excel: {e}")
            return False
            
    def update_with_issues(self, issues_data):
        """Actualiza el archivo Excel con los datos extraídos"""
        if not self.file_path:
            logger.error("No hay ruta de archivo Excel especificada")
            return False
            
        if not issues_data:
            logger.warning("No hay datos para actualizar en Excel")
            return False
            
        try:
            logger.info(f"Actualizando archivo Excel: {self.file_path}")
            
            # Cargar el archivo existente o crear estructura si no existe
            if os.path.exists(self.file_path):
                try:
                    existing_df = pd.read_excel(self.file_path, engine='openpyxl')
                    logger.info(f"Archivo Excel existente cargado con {len(existing_df)} registros")
                except Exception as read_e:
                    logger.warning(f"Error al leer Excel: {read_e}. Creando estructura nueva.")
                    existing_df = pd.DataFrame(
                        columns=[
                            "Title", "Type", "Priority", "Status", "Deadline", "Due Date",
                            "Created By", "Created On", "Last Updated", "Comments"
                        ]
                    )
            else:
                existing_df = pd.DataFrame(
                    columns=[
                        "Title", "Type", "Priority", "Status", "Deadline", "Due Date",
                        "Created By", "Created On", "Last Updated", "Comments"
                    ]
                )
                logger.info("Creando nueva estructura de Excel")

            # Convertir datos de issues a DataFrame
            new_df = pd.DataFrame(issues_data)
            
            # Contadores para estadísticas
            new_items = 0
            updated_items = 0
            
            # Hacer una copia del DataFrame existente para no modificarlo mientras iteramos
            updated_df = existing_df.copy()
            
            # Optimización: crear índice para búsquedas rápidas si hay muchos registros
            title_index = {}
            if len(existing_df) > 0 and "Title" in existing_df.columns:
                for idx, title in enumerate(existing_df["Title"]):
                    if title not in title_index:
                        title_index[title] = idx
            
            # Procesar cada issue nuevo
            for _, new_row in new_df.iterrows():
                title = new_row["Title"]
                title_exists = title in title_index
                
                if not title_exists:
                    # Agregar fecha de última actualización para elementos nuevos
                    new_row_dict = new_row.to_dict()
                    new_row_dict["Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    new_row_df = pd.DataFrame([new_row_dict])
                    updated_df = pd.concat([updated_df, new_row_df], ignore_index=True)
                    new_items += 1
                    logger.info(f"Nuevo issue añadido: '{title}'")
                else:
                    # Obtener índice del elemento existente
                    idx = title_index[title]
                    
                    # Verificar cambios en el estado y otras columnas
                    updated = False
                    
                    for column in ["Status", "Priority", "Type", "Due Date", "Deadline", "Created By", "Created On"]:
                        if column in new_row and column in existing_df.columns:
                            old_value = existing_df.iloc[idx][column] if not pd.isna(existing_df.iloc[idx][column]) else ""
                            new_value = new_row[column] if not pd.isna(new_row[column]) else ""
                            
                            if str(old_value) != str(new_value):
                                mask = updated_df["Title"] == title
                                updated_df.loc[mask, column] = new_value
                                updated_df.loc[mask, "Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                updated = True
                                logger.info(f"Actualizado {column} de '{title}': '{old_value}' → '{new_value}'")
                    
                    if updated:
                        updated_items += 1
            
            # Guardar el DataFrame actualizado
            updated_df.to_excel(self.file_path, index=False, engine='openpyxl')
            
            # Aplicar formato al Excel
            self._apply_excel_formatting()
            
            logger.info(f"Excel actualizado: {new_items} nuevos, {updated_items} actualizados")
            return True, new_items, updated_items
            
        except Exception as e:
            logger.error(f"Error al actualizar el archivo Excel: {e}")
            return False, 0, 0
            
    def _apply_excel_formatting(self):
        """Aplica formato estético al archivo Excel"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = load_workbook(self.file_path)
            ws = wb.active
            
            # Formato para encabezados
            header_fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Aplicar formato a encabezados
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Aplicar formato a celdas de datos
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

                    # Colorear por estado
                    if col == 4:  # Columna Status
                        status = str(cell.value).upper() if cell.value else ""

                        if "DONE" in status:
                            cell.fill = PatternFill(
                                start_color="CCFFCC",
                                end_color="CCFFCC",
                                fill_type="solid",
                            )
                        elif "OPEN" in status:
                            cell.fill = PatternFill(
                                start_color="FFCCCC",
                                end_color="FFCCCC",
                                fill_type="solid",
                            )
                        elif "READY" in status:
                            cell.fill = PatternFill(
                                start_color="FFFFCC",
                                end_color="FFFFCC",
                                fill_type="solid",
                            )
                        elif "IN PROGRESS" in status:
                            cell.fill = PatternFill(
                                start_color="FFE6CC",
                                end_color="FFE6CC",
                                fill_type="solid",
                            )

            # Ajustar ancho de columnas
            for col in range(1, ws.max_column + 1):
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                adjusted_width = max(10, min(50, max_length + 2))
                ws.column_dimensions[
                    ws.cell(row=1, column=col).column_letter
                ].width = adjusted_width

            wb.save(self.file_path)
            logger.info("Formato aplicado al archivo Excel correctamente")
            return True
        except Exception as format_e:
            logger.warning(f"No se pudo aplicar formato al Excel: {format_e}")
            return False
        
        
        
        
        
        
        
        
        
        
        
        
        
        
class SAPBrowser:
    """Clase para la automatización del navegador y extracción de datos de SAP"""
    
    def __init__(self):
        """Inicializa el controlador del navegador"""
        self.driver = None
        self.wait = None
        self.element_cache = {}  # Caché para elementos encontrados frecuentemente
        
    def connect(self):
        """Inicia una sesión de navegador con perfil dedicado"""
        logger.info("Iniciando navegador con perfil guardado...")
        
        try:
            # Ruta al directorio del perfil
            user_data_dir = os.path.join(os.environ['USERPROFILE'], 'AppData', 'Local', 'Google', 'Chrome', 'SAP_Automation')
            
            # Crear directorio si no existe
            if not os.path.exists(user_data_dir):
                os.makedirs(user_data_dir)
                logger.info(f"Creado directorio de perfil: {user_data_dir}")
            
            # Configurar opciones de Chrome
            chrome_options = Options()
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
            chrome_options.add_argument("--profile-directory=Default")
            
            # Opciones para mejorar rendimiento y estabilidad
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-infobars")
            chrome_options.add_argument("--disable-notifications")
            chrome_options.add_argument("--disable-popup-blocking")
            
            # Optimización de memoria
            chrome_options.add_argument("--js-flags=--expose-gc")
            chrome_options.add_argument("--enable-precise-memory-info")
            
            # Agregar opciones para permitir que el usuario use el navegador mientras se ejecuta el script
            chrome_options.add_experimental_option("detach", True)
            
            # Intentar iniciar el navegador
            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 60)  # Timeout de 60 segundos
            
            logger.info("Navegador Chrome iniciado correctamente")
            return True
            
        except WebDriverException as e:
            if "Chrome failed to start" in str(e):
                logger.error(f"Error al iniciar Chrome: {e}. Verificando si hay instancias abiertas...")
                # Intentar recuperar sesión existente
                try:
                    chrome_options = Options()
                    chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
                    self.driver = webdriver.Chrome(options=chrome_options)
                    self.wait = WebDriverWait(self.driver, 60)
                    logger.info("Conexión exitosa a sesión existente de Chrome")
                    return True
                except Exception as debug_e:
                    logger.error(f"No se pudo conectar a sesión existente: {debug_e}")
            
            logger.error(f"Error al iniciar Navegador: {e}")
            return False
    
    def navigate_to_sap(self):
        """Navega a la URL de SAP y maneja posibles diálogos de certificados"""
        if not self.driver:
            logger.error("No hay navegador iniciado")
            return False
            
        try:
            # Navegar a la URL de SAP
            self.driver.get("https://xalm-prod.x.eu20.alm.cloud.sap/launchpad#sdwork-center&/")
            logger.info("Navegando a URL de SAP...")
            
            # Esperar a que cargue la página
            time.sleep(5)
            
            # Intentar aceptar certificados si aparece el diálogo
            try:
                ok_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'OK') or contains(text(), 'Ok')]"))
                )
                ok_button.click()
                logger.info("Se hizo clic en el botón OK del certificado")
            except TimeoutException:
                logger.info("No se encontró diálogo de certificado o ya fue aceptado")
                
            return True
            
        except Exception as e:
            logger.error(f"Error al navegar a SAP: {e}")
            return False
    
    def get_total_issues_count(self):
        """Obtiene el número total de issues desde el encabezado"""
        try:
            # Estrategia 1: Buscar el texto "Issues (número)"
            try:
                header_text = self.driver.find_element(
                    By.XPATH, "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                ).text
                logger.info(f"Texto del encabezado de issues: {header_text}")
                
                # Extraer el número entre paréntesis
                match = re.search(r'\((\d+)\)', header_text)
                if match:
                    return int(match.group(1))
            except NoSuchElementException:
                logger.warning("No se encontró el encabezado Issues con formato (número)")
            
            # Estrategia 2: Buscar contador específico de SAP UI5
            try:
                counter_element = self.driver.find_element(
                    By.XPATH, "//div[contains(@class, 'sapMITBCount')]"
                )
                if counter_element.text.isdigit():
                    return int(counter_element.text)
            except NoSuchElementException:
                logger.warning("No se encontró contador de issues en formato SAP UI5")
            
            # Estrategia 3: Contar filas visibles y usar como estimación
            rows = self.find_table_rows(highlight=False)
            if rows:
                count = len(rows)
                logger.info(f"Estimando número de issues basado en filas visibles: {count}")
                return max(count, 100)  # Al menos 100 para asegurar cobertura completa
            
            logger.warning("No se pudo determinar el número total de issues, usando valor por defecto")
            return 100  # Valor por defecto
            
        except Exception as e:
            logger.error(f"Error al obtener el total de issues: {e}")
            return 100  # Valor por defecto si hay error
    
    def check_for_pagination(self):
        """Verifica si la tabla tiene paginación y devuelve los controles"""
        try:
            # Selectores para controles de paginación
            pagination_selectors = [
                "//div[contains(@class, 'sapMPaginator')]",
                "//div[contains(@class, 'sapUiTablePaginator')]",
                "//div[contains(@class, 'pagination')]",
                "//button[contains(@class, 'navButton') or contains(@aria-label, 'Next') or contains(@aria-label, 'Siguiente')]",
                "//span[contains(@class, 'sapMPaginatorButton')]",
                "//button[contains(text(), 'Next') or contains(text(), 'Siguiente')]",
                "//a[contains(@class, 'sapMBtn') and contains(@aria-label, 'Next')]"
            ]
            
            for selector in pagination_selectors:
                elements = self.driver.find_elements(By.XPATH, selector)
                if elements:
                    logger.info(f"Se encontraron controles de paginación: {len(elements)} elementos con selector {selector}")
                    return elements
            
            # Buscar elementos "Show More" o "Load More"
            load_more_selectors = [
                "//button[contains(text(), 'More') or contains(text(), 'más') or contains(text(), 'Show')]",
                "//a[contains(text(), 'More') or contains(text(), 'Load')]",
                "//div[contains(@class, 'sapMListShowMoreButton')]",
                "//span[contains(text(), 'Show') and contains(text(), 'More')]/..",
                "//span[contains(@class, 'sapUiTableColShowMoreButton')]"
            ]
            
            for selector in load_more_selectors:
                elements = self.driver.find_elements(By.XPATH, selector)
                if elements:
                    logger.info(f"Se encontraron botones 'Show More': {len(elements)} elementos con selector {selector}")
                    return elements
            
            logger.info("No se encontraron controles de paginación en la tabla")
            return None
            
        except Exception as e:
            logger.error(f"Error al verificar paginación: {e}")
            return None
    









    def click_pagination_next(self, pagination_elements):
        """Hace clic en el botón de siguiente página"""
        if not pagination_elements:
            return False
            
        try:
            # Buscar el botón "Next" entre los elementos de paginación
            next_button = None
            
            for element in pagination_elements:
                try:
                    aria_label = element.get_attribute("aria-label") or ""
                    text = element.text.lower()
                    classes = element.get_attribute("class") or ""
                    
                    # Comprobar si es un botón "Next" o "Siguiente"
                    if ("next" in aria_label.lower() or 
                        "siguiente" in aria_label.lower() or
                        "next" in text or 
                        "siguiente" in text or
                        "show more" in text.lower() or
                        "more" in text.lower()):
                        
                        next_button = element
                        break
                        
                    # Comprobar por clase CSS
                    if ("next" in classes.lower() or 
                        "pagination-next" in classes.lower() or
                        "sapMBtn" in classes and "NavButton" in classes):
                        
                        next_button = element
                        break
                except Exception:
                    continue
            
            # Si se encontró un botón Next, intentar hacer clic
            if next_button:
                # Verificar si el botón está habilitado
                disabled = next_button.get_attribute("disabled") == "true" or next_button.get_attribute("aria-disabled") == "true"
                
                if disabled:
                    logger.info("Botón de siguiente página está deshabilitado")
                    return False
                    
                # Scroll hacia el botón para asegurar que está visible
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
                time.sleep(0.5)
                
                # Intentar clic con distintos métodos, en orden de preferencia
                try:
                    self.driver.execute_script("arguments[0].click();", next_button)
                    logger.info("Clic en botón 'Next' realizado con JavaScript")
                    time.sleep(2)
                    return True
                except JavascriptException:
                    try:
                        next_button.click()
                        logger.info("Clic en botón 'Next' realizado")
                        time.sleep(2)
                        return True
                    except ElementClickInterceptedException:
                        from selenium.webdriver.common.action_chains import ActionChains
                        actions = ActionChains(self.driver)
                        actions.move_to_element(next_button).click().perform()
                        logger.info("Clic en botón 'Next' realizado con ActionChains")
                        time.sleep(2)
                        return True
            
            # Si no se encontró botón específico, intentar con el último elemento
            if pagination_elements and len(pagination_elements) > 0:
                last_element = pagination_elements[-1]
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", last_element)
                time.sleep(0.5)
                self.driver.execute_script("arguments[0].click();", last_element)
                logger.info("Clic en último elemento de paginación realizado")
                time.sleep(2)
                return True
            
            logger.warning("No se pudo identificar o hacer clic en el botón 'Next'")
            return False
            
        except Exception as e:
            logger.error(f"Error al hacer clic en paginación: {e}")
            return False
    
    def scroll_to_load_all_items(self, total_expected=100, max_attempts=100):
        """Estrategia optimizada para cargar todos los elementos mediante scroll"""
        logger.info(f"Iniciando carga de {total_expected} elementos...")
        
        previous_rows_count = 0
        no_change_count = 0
        no_change_threshold = 10
        
        # Verificar si hay paginación
        pagination_elements = self.check_for_pagination()
        has_pagination = pagination_elements is not None and len(pagination_elements) > 0
        
        logger.info(f"¿La tabla tiene paginación? {'Sí' if has_pagination else 'No'}")
        
        for attempt in range(max_attempts):
            try:
                # Estrategia 1: Scroll adaptativo con diferentes métodos
                for scroll_method in range(3):
                    if scroll_method == 0:
                        # Scroll normal al final de la página
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    elif scroll_method == 1:
                        # Scroll en contenedores específicos de SAP
                        self.driver.execute_script("""
                            var containers = document.querySelectorAll('.sapMListItems, .sapMTableTBody, .sapUiTableCtrlScr');
                            if (containers.length > 0) {
                                for(var i=0; i<containers.length; i++) {
                                    containers[i].scrollTop = containers[i].scrollHeight;
                                }
                            }
                        """)
                    elif scroll_method == 2 and attempt % 3 == 0:
                        # Scroll progresivo cada 3 intentos
                        doc_height = self.driver.execute_script("return document.body.scrollHeight")
                        for pos in range(0, doc_height, 300):
                            self.driver.execute_script(f"window.scrollTo(0, {pos});")
                            time.sleep(0.05)
                
                # Estrategia 2: Hacer clic en botones "Show More" cada 2 intentos
                if attempt % 2 == 0:
                    try:
                        load_more_buttons = self.driver.find_elements(
                            By.XPATH,
                            "//button[contains(text(), 'More')] | " +
                            "//button[contains(text(), 'más')] | " +
                            "//a[contains(text(), 'More')] | " +
                            "//div[contains(@class, 'sapMListShowMoreButton')]"
                        )
                        
                        if load_more_buttons:
                            for btn in load_more_buttons:
                                try:
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                                    time.sleep(0.3)
                                    self.driver.execute_script("arguments[0].click();", btn)
                                    logger.info("Clic en botón 'Show More'")
                                    time.sleep(1.5)
                                    break  # Solo hacer clic en un botón por intento
                                except:
                                    continue
                    except Exception as btn_e:
                        logger.debug(f"Error al buscar botón 'Show More': {btn_e}")
                
                # Contar filas visibles
                rows = self.find_table_rows(highlight=False)
                current_rows_count = len(rows)
                
                if attempt % 10 == 0:
                    logger.info(f"Intento {attempt+1}: {current_rows_count} filas cargadas")
                
                # Verificación de carga completa
                if current_rows_count == previous_rows_count:
                    no_change_count += 1
                    
                    # Si hay paginación y no hay cambios, intentar pasar a la siguiente página
                    if has_pagination and no_change_count >= 5:
                        logger.info("Intentando pasar a la siguiente página...")
                        pagination_elements = self.check_for_pagination()
                        if pagination_elements and self.click_pagination_next(pagination_elements):
                            logger.info("Se pasó a la siguiente página")
                            no_change_count = 0
                            time.sleep(3)
                            continue
                    
                    # Si no hay cambios por muchos intentos, hacer scroll adicional
                    if no_change_count >= 5:
                        # Scroll por posiciones incrementales
                        for scroll_pos in range(1000, 10000, 1000):
                            self.driver.execute_script(f"window.scrollTo(0, {scroll_pos});")
                            time.sleep(0.3)
                        
                        # Intentar con teclas de dirección
                        try:
                            body = self.driver.find_element(By.TAG_NAME, "body")
                            body.send_keys(Keys.PAGE_DOWN)
                            time.sleep(0.3)
                            body.send_keys(Keys.END)
                            time.sleep(0.3)
                        except:
                            pass
                    
                    # Criterios de finalización
                    if no_change_count >= no_change_threshold and current_rows_count >= total_expected * 0.9:
                        logger.info(f"Se han cargado {current_rows_count} filas (>= 90% del total esperado)")
                        break
                        
                    if no_change_count >= no_change_threshold * 2:
                        logger.warning(f"No se detectaron más filas después de {no_change_count} intentos")
                        break
                else:
                    # Reiniciar contador si se encontraron más filas
                    no_change_count = 0
                    
                previous_rows_count = current_rows_count
                
                # Si se alcanzó o superó el número esperado, terminar
                if current_rows_count >= total_expected:
                    logger.info(f"Se han cargado {current_rows_count} filas (>= {total_expected} esperadas)")
                    break
                
                # Tiempo adaptativo de espera
                wait_time = 0.2 + (no_change_count * 0.1)
                time.sleep(min(wait_time, 1.0))
                    
            except Exception as e:
                logger.warning(f"Error durante el scroll en intento {attempt+1}: {e}")
            
        # Calcular cobertura
        coverage = (previous_rows_count / total_expected) * 100 if total_expected > 0 else 0
        logger.info(f"Scroll completado. Cobertura: {coverage:.2f}% ({previous_rows_count}/{total_expected})")
        
        return previous_rows_count
    
    def find_table_rows(self, highlight=False):
        """Encuentra todas las filas de la tabla con múltiples estrategias"""
        all_rows = []
        
        # Usar caché si está disponible
        cache_key = "table_rows"
        if cache_key in self.element_cache:
            cache_time, cached_rows = self.element_cache[cache_key]
            if (datetime.now() - cache_time).total_seconds() < 5:  # Caché válida por 5 segundos
                logger.debug("Usando filas en caché")
                return cached_rows

        # Selectores mejorados para SAP UI5
        selectors = [
            # Selectores de SAP estándar
            "//table[contains(@class, 'sapMListTbl')]/tbody/tr[not(contains(@class, 'sapMListTblHeader'))]",
            "//div[contains(@class, 'sapMList')]//li[contains(@class, 'sapMLIB')]",
            "//table[contains(@class, 'sapMList')]/tbody/tr",
            "//div[@role='row'][not(contains(@class, 'sapMListHeaderSubTitleItems')) and not(contains(@class, 'sapMListTblHeader'))]",
            "//div[contains(@class, 'sapMListItems')]/div[contains(@class, 'sapMListItem')]",
            "//div[contains(@class, 'sapMListItems')]//div[contains(@class, 'sapMObjectIdentifier')]/..",
            "//div[contains(@class, 'sapMListItem')]",
            
            # Selectores de Fiori
            "//div[contains(@class, 'sapMList')]//li[@tabindex]",
            "//div[contains(@class, 'sapUiTable')]//tr[contains(@class, 'sapUiTableRow')]",
            "//div[contains(@class, 'sapUiTableRowHdr')]/..",
            "//table[contains(@class, 'sapMTable')]//tr[not(contains(@class, 'sapMListTblHeaderRow'))]",
            
            # Selectores específicos de SDWork Center
            "//div[contains(@class, 'sdworkItems')]//div[contains(@class, 'sapMLIB')]",
            "//div[contains(@class, 'issueList')]//div[contains(@class, 'sapMLIB')]",
            "//div[contains(@id, 'issue')]//div[contains(@class, 'sapMLIB')]",
            
            # Selectores genéricos más específicos
            "//div[contains(@class, 'sapMLIB-CTX')]",
            "//div[contains(@class, 'sapMObjectListItem')]",
            "//div[contains(@class, 'sapMListModeMultiSelect')]//div[contains(@class, 'sapMLIB')]"
        ]

        for selector in selectors:
            try:
                rows = self.driver.find_elements(By.XPATH, selector)
                if len(rows) > 0:
                    logger.info(f"Se encontraron {len(rows)} filas con selector: {selector}")
                    
                    # Filtrar filas válidas
                    valid_rows = []
                    for row in rows:
                        try:
                            has_content = False
                            text_elements = row.find_elements(By.XPATH, ".//span | .//div | .//a")
                            for element in text_elements:
                                if element.text and element.text.strip():
                                    has_content = True
                                    break
                                    
                            if has_content:
                                # Verificar que no sea un encabezado
                                class_attr = row.get_attribute("class") or ""
                                is_header = "header" in class_attr.lower()
                                if not is_header:
                                    valid_rows.append(row)
                        except:
                            valid_rows.append(row)  # Si hay error, incluir por si acaso
                    
                    if len(valid_rows) > 0:
                        all_rows = valid_rows
                        
                        if len(valid_rows) >= 75:  # Si encontramos muchas filas, probablemente es el selector correcto
                            break
            except Exception as e:
                logger.debug(f"Error con selector {selector}: {e}")
        
        # Si los selectores estándar fallan, usar enfoque alternativo
        if len(all_rows) == 0:
            logger.warning("Usando enfoque alternativo para encontrar filas")
            try:
                any_rows = self.driver.find_elements(
                    By.XPATH,
                    "//div[contains(@class, 'sapM')] | //tr | //li[contains(@class, 'sapM')]"
                )
                
                for element in any_rows:
                    try:
                        # Verificar si parece una fila de datos
                        if element.text and len(element.text.strip()) > 10:
                            children = element.find_elements(By.XPATH, ".//*")
                            if len(children) >= 3:
                                parent_elements = element.find_elements(
                                    By.XPATH, 
                                    "./ancestor::div[contains(@class, 'sapMList') or contains(@class, 'sapMTable')]"
                                )
                                if len(parent_elements) > 0:
                                    all_rows.append(element)
                    except:
                        continue
                        
                logger.info(f"Enfoque alternativo encontró {len(all_rows)} posibles filas")
            except Exception as e:
                logger.error(f"Error en enfoque alternativo: {e}")
        
        # Resaltar filas si se solicita
        if highlight and len(all_rows) > 0:
            try:
                self.driver.execute_script(
                    "arguments[0].scrollIntoView(true); arguments[0].style.border = '2px solid red';",
                    all_rows[0]
                )
            except:
                pass
        
        # Eliminar duplicados manteniendo el orden
        unique_rows = []
        seen_ids = set()
        
        for row in all_rows:
            try:
                row_id = row.id
                if row_id not in seen_ids:
                    seen_ids.add(row_id)
                    unique_rows.append(row)
            except:
                # Si no podemos obtener un ID, usar aproximación con texto y clase
                try:
                    row_text = row.text[:50] if row.text else ""
                    row_class = row.get_attribute("class") or ""
                    row_signature = f"{row_text}|{row_class}"
                    
                    if row_signature not in seen_ids:
                        seen_ids.add(row_signature)
                        unique_rows.append(row)
                except:
                    unique_rows.append(row)
        
        logger.info(f"Total de filas únicas encontradas: {len(unique_rows)}")
        
        # Actualizar caché
        self.element_cache[cache_key] = (datetime.now(), unique_rows)
        
        return unique_rows
    
    def extract_issues_data(self):
        """Extrae datos de issues desde la tabla con procesamiento mejorado"""
        try:
            logger.info("Iniciando extracción de issues...")
            
            # Esperar a que cargue la página inicial
            time.sleep(3)
            
            # Obtener el número total de issues
            total_issues = self.get_total_issues_count()
            logger.info(f"Total de issues a procesar: {total_issues}")
            
            # Hacer scroll para cargar todos los elementos
            loaded_rows_count = self.scroll_to_load_all_items(total_issues)
            
            # Verificar si hay paginación
            pagination_elements = self.check_for_pagination()
            has_pagination = pagination_elements is not None and len(pagination_elements) > 0
            
            # Lista para almacenar todos los datos extraídos
            all_issues_data = []
            seen_titles = set()  # Para evitar duplicados
            
            page_num = 1
            max_pages = 20  # Límite de seguridad
            
            while page_num <= max_pages:
                logger.info(f"Procesando página {page_num}...")
                
                # Obtener filas de la página actual
                rows = self.find_table_rows(highlight=False)
                
                if not rows:
                    logger.warning(f"No se encontraron filas en la página {page_num}")
                    break
                
                logger.info(f"Encontradas {len(rows)} filas en la página {page_num}")
                
                # Procesar filas en esta página
                page_issues_data = self._process_table_rows(rows, seen_titles)
                
                # Agregar los datos de esta página al resultado total
                all_issues_data.extend(page_issues_data)
                
                logger.info(f"Extraídos {len(page_issues_data)} issues de la página {page_num}")
                logger.info(f"Total de issues extraídos hasta ahora: {len(all_issues_data)}")
                
                # Si no hay paginación o ya procesamos todos los datos, terminar
                if not has_pagination or len(page_issues_data) == 0:
                    break
                
                # Intentar pasar a la siguiente página
                if pagination_elements:
                    if not self.click_pagination_next(pagination_elements):
                        logger.info("No se pudo pasar a la siguiente página, terminando extracción")
                        break
                        
                    # Esperar a que cargue la nueva página
                    time.sleep(3)
                    
                    # Actualizar elementos de paginación (pueden cambiar entre páginas)
                    pagination_elements = self.check_for_pagination()
                    
                    page_num += 1
                else:
                    break
            
            logger.info(f"Extracción completada. Total de issues extraídos: {len(all_issues_data)}")
            
            return all_issues_data
        
        except Exception as e:
            logger.error(f"Error en la extracción de datos: {e}")
            return []
    
    def _process_table_rows(self, rows, seen_titles):
        """Procesa las filas de la tabla y extrae los datos de cada issue"""
        issues_data = []
        processed_count = 0
        batch_size = 10  # Procesar en lotes para actualizar progreso
        
        for index, row in enumerate(rows):
            try:
                # Extraer título
                title = self._extract_title(row)
                
                if not title:
                    title = f"Issue sin título #{index+1}"
                
                # Verificar duplicados
                title_lower = title.lower()
                if title_lower in seen_titles:
                    continue
                
                seen_titles.add(title_lower)
                
                # Extraer resto de datos
                type_text = self._extract_type(row, title)
                priority = self._extract_priority(row)
                status = self._extract_status(row)
                deadline = self._extract_deadline(row)
                due_date = self._extract_due_date(row)
                created_by = self._extract_created_by(row)
                created_on = self._extract_created_on(row)
                
                # Datos del issue completos
                issue_data = {
                    'Title': title,
                    'Type': type_text,
                    'Priority': priority,
                    'Status': status,
                    'Deadline': deadline,
                    'Due Date': due_date,
                    'Created By': created_by,
                    'Created On': created_on
                }
                
                issues_data.append(issue_data)
                processed_count += 1
                
                if processed_count % batch_size == 0:
                    logger.info(f"Procesados {processed_count} issues hasta ahora")
                
            except Exception as e:
                logger.error(f"Error al procesar la fila {index}: {e}")
        
        logger.info(f"Procesamiento de filas completado. Total procesado: {processed_count} issues")
        return issues_data
    
    def _extract_title(self, row):
        """Extrae el título de una fila"""
        try:
            # Intentar múltiples métodos para extraer título
            title_extractors = [
                lambda r: r.find_element(By.XPATH, ".//a").text,
                lambda r: r.find_element(By.XPATH, ".//span[contains(@class, 'title')]").text,
                lambda r: r.find_element(By.XPATH, ".//div[contains(@class, 'title')]").text,
                lambda r: r.find_elements(By.XPATH, ".//div[@role='gridcell']")[0].text if r.find_elements(By.XPATH, ".//div[@role='gridcell']") else None,
                lambda r: r.find_elements(By.XPATH, ".//td")[0].text if r.find_elements(By.XPATH, ".//td") else None,
                lambda r: r.find_element(By.XPATH, ".//*[contains(@id, 'title')]").text,
                lambda r: r.find_element(By.XPATH, ".//div[@title]").get_attribute("title"),
                lambda r: r.find_element(By.XPATH, ".//span[@title]").get_attribute("title")
            ]

            for extractor in title_extractors:
                try:
                    title_text = extractor(row)
                    if title_text and title_text.strip():
                        return title_text.strip()
                except:
                    continue
            
            # Si no encontramos un título específico, usar el texto completo
            try:
                full_text = row.text.strip()
                if full_text:
                    lines = full_text.split('\n')
                    if lines:
                        title = lines[0].strip()
                        if len(title) > 100:  # Si es muy largo, recortar
                            title = title[:100] + "..."
                        return title
            except:
                pass
                
            return None
        except Exception as e:
            logger.debug(f"Error al extraer título: {e}")
            return None
    
    def _extract_type(self, row, title):
        """Extrae el tipo de issue"""
        try:
            # Buscar en la segunda columna
            type_elements = row.find_elements(By.XPATH, 
                ".//div[@role='gridcell'][2]//span | "
                ".//td[2]//span | "
                ".//div[contains(@class, 'type')] | "
                ".//span[contains(@class, 'type')]")
            
            if type_elements:
                type_text = type_elements[0].text.strip()
                if type_text:
                    return type_text
            
            # Si no se encontró, buscar en el título
            if title:
                potential_types = [
                    "Recommendation", "Implementation", "Question", 
                    "Problem", "Incident", "Request", "Task"
                ]
                for potential_type in potential_types:
                    if potential_type.lower() in title.lower():
                        return potential_type
            
            return "Issue"  # Valor por defecto
        except Exception as e:
            logger.debug(f"Error al extraer tipo: {e}")
            return "Issue"
    
    def _extract_priority(self, row):
        """Extrae la prioridad del issue"""
        try:
            # Buscar indicadores de prioridad
            priority_indicators = [
                (By.XPATH, ".//span[contains(@class, 'sapMGaugeNegativeColor')]", "Very High"),
                (By.XPATH, ".//span[contains(@class, 'sapMGaugeCriticalColor')]", "High"),
                (By.XPATH, ".//span[contains(@class, 'sapMGaugeNeutralColor')]", "Medium"),
                (By.XPATH, ".//span[contains(@class, 'sapMGaugePositiveColor')]", "Low"),
                (By.XPATH, ".//span[contains(text(), 'Very High')]", "Very High"),
                (By.XPATH, ".//span[contains(text(), 'High')]", "High"),
                (By.XPATH, ".//span[contains(text(), 'Medium')]", "Medium"),
                (By.XPATH, ".//span[contains(text(), 'Low')]", "Low"),
                (By.XPATH, ".//i[contains(@class, 'High')]", "High"),
                (By.XPATH, ".//div[@role='gridcell'][3]//span", "")
            ]
            
            for locator, indicator_text in priority_indicators:
                elements = row.find_elements(locator)
                if elements:
                    if indicator_text:
                        return indicator_text
                    elif elements[0].text:
                        cell_text = elements[0].text.lower()
                        if "very high" in cell_text:
                            return "Very High"
                        elif "high" in cell_text:
                            return "High"
                        elif "medium" in cell_text:
                            return "Medium"
                        elif "low" in cell_text:
                            return "Low"
            
            # Buscar por clases de color
            color_classes = [
                (By.XPATH, ".//*[contains(@class, 'red') or contains(@style, 'red')]", "High"),
                (By.XPATH, ".//*[contains(@class, 'yellow') or contains(@style, 'yellow')]", "Medium"),
                (By.XPATH, ".//*[contains(@class, 'green') or contains(@style, 'green')]", "Low"),
                (By.XPATH, ".//*[contains(@class, 'orange') or contains(@style, 'orange')]", "High")
            ]
            
            for locator, indicator_text in color_classes:
                elements = row.find_elements(locator)
                if elements:
                    return indicator_text
            
            return "N/A"
        except Exception as e:
            logger.debug(f"Error al extraer prioridad: {e}")
            return "N/A"
    
    def _extract_status(self, row):
        """Extrae el estado del issue"""
        try:
            # Intentar obtener estado de la columna correspondiente
            cells = self._get_row_cells(row)
            
            if cells and len(cells) >= 4:
                status_text = cells[3].text.strip()
                if status_text:
                    status_lines = status_text.split("\n")
                    if status_lines:
                        for line in status_lines:
                            if any(keyword in line.upper() for keyword in ["OPEN", "DONE", "READY", "DRAFT", "IN PROGRESS"]):
                                return self._normalize_status(line.strip())
                        return self._normalize_status(status_lines[0].strip())
                
                # Buscar por clase específica
                status_elements = cells[3].find_elements(By.XPATH, 
                    ".//div[contains(@class, 'status')] | .//span[contains(@class, 'status')]")
                if status_elements:
                    return self._normalize_status(status_elements[0].text.strip())
            
            # Buscar elementos con texto específico de estado
            status_texts = [
                "OPEN", "DONE", "READY FOR PUBLISHING", "IN PROGRESS", "CLOSED", "ACCEPTED", "DRAFT"
            ]
            
            for status_text in status_texts:
                status_elements = row.find_elements(
                    By.XPATH, 
                    f".//div[contains(text(), '{status_text}')] | .//span[contains(text(), '{status_text}')]"
                )
                
                if status_elements:
                    return self._normalize_status(status_text)
            
            # Buscar por clases relacionadas con estado
            status_class_elements = row.find_elements(
                By.XPATH,
                ".//*[contains(@class, 'status') or contains(@class, 'state')]"
            )
            
            if status_class_elements:
                for element in status_class_elements:
                    if element.text.strip():
                        return self._normalize_status(element.text.strip())
            
            return "N/A"
        except Exception as e:
            logger.debug(f"Error al extraer estado: {e}")
            return "N/A"







    def _normalize_status(self, status_text):
        """Normaliza el texto de estado a valores estándar"""
        if not status_text:
            return "N/A"
            
        status_upper = status_text.upper()
        
        if "DONE" in status_upper or "COMPLETED" in status_upper:
            return "DONE"
        elif "OPEN" in status_upper:
            return "OPEN"
        elif "IN PROGRESS" in status_upper or "PROCESSING" in status_upper:
            return "IN PROGRESS"
        elif "READY" in status_upper and "PUBLISHING" in status_upper:
            return "READY FOR PUBLISHING"
        elif "READY" in status_upper:
            return "READY"
        elif "DRAFT" in status_upper:
            return "DRAFT"
        elif "CLOSED" in status_upper:
            return "CLOSED"
        elif "ACCEPTED" in status_upper:
            return "ACCEPTED"
            
        return status_text
    
    def _extract_deadline(self, row):
        """Extrae la fecha límite del issue"""
        try:
            cells = self._get_row_cells(row)
            
            if cells and len(cells) >= 5:
                deadline_text = cells[4].text.strip()
                if deadline_text:
                    return deadline_text
            
            return ""
        except Exception as e:
            logger.debug(f"Error al extraer deadline: {e}")
            return ""
    
    def _extract_due_date(self, row):
        """Extrae la fecha de vencimiento del issue"""
        try:
            cells = self._get_row_cells(row)
            
            if cells and len(cells) >= 6:
                due_date_text = cells[5].text.strip()
                if due_date_text:
                    return due_date_text
            
            return "N/A"
        except Exception as e:
            logger.debug(f"Error al extraer due date: {e}")
            return "N/A"
    
    def _extract_created_by(self, row):
        """Extrae quién creó el issue"""
        try:
            cells = self._get_row_cells(row)
            
            if cells and len(cells) >= 7:
                created_by_text = cells[6].text.strip()
                if created_by_text:
                    # Limpiar el texto si contiene múltiples líneas
                    if "\n" in created_by_text:
                        created_by_lines = created_by_text.split("\n")
                        for line in created_by_lines:
                            if line.startswith("I") and len(line) <= 10:
                                return line  # Probablemente un ID de usuario
                    return created_by_text
            
            return "N/A"
        except Exception as e:
            logger.debug(f"Error al extraer creador: {e}")
            return "N/A"
    
    def _extract_created_on(self, row):
        """Extrae la fecha de creación del issue"""
        try:
            cells = self._get_row_cells(row)
            
            if cells and len(cells) >= 8:
                created_on_text = cells[7].text.strip()
                if created_on_text:
                    # Limpiar la fecha si tiene formato largo
                    if "," in created_on_text:
                        date_parts = created_on_text.split(",")
                        if len(date_parts) > 1:
                            return ",".join(date_parts[-2:]).strip()
                    return created_on_text
            
            return "N/A"
        except Exception as e:
            logger.debug(f"Error al extraer fecha de creación: {e}")
            return "N/A"
    
    def _get_row_cells(self, row):
        """Obtiene todas las celdas de una fila usando diferentes métodos"""
        cells = []
        
        try:
            # Intentar diferentes métodos para obtener celdas
            cell_extractors = [
                lambda r: r.find_elements(By.XPATH, ".//div[@role='gridcell']"),
                lambda r: r.find_elements(By.XPATH, ".//td"),
                lambda r: r.find_elements(By.XPATH, "./div")
            ]
            
            for extractor in cell_extractors:
                try:
                    extracted_cells = extractor(row)
                    if extracted_cells and len(extracted_cells) > 1:
                        return extracted_cells
                except:
                    continue
        except Exception as e:
            logger.debug(f"Error al extraer celdas: {e}")
            
        return cells
    
    def find_ui5_elements(self, control_type, properties=None):
        """Encuentra elementos UI5 específicos usando JavaScript"""
        script = """
        function findUI5Controls(controlType, properties) {
            if (!window.sap || !window.sap.ui) return [];
            
            var controls = sap.ui.getCore().byFieldGroupId().filter(function(control) {
                return control.getMetadata().getName() === controlType;
            });
            
            if (properties) {
                controls = controls.filter(function(control) {
                    for (var prop in properties) {
                        if (control.getProperty(prop) !== properties[prop]) {
                            return false;
                        }
                    }
                    return true;
                });
            }
            
            return controls.map(function(control) {
                return control.getId();
            });
        }
        return findUI5Controls(arguments[0], arguments[1]);
        """
        
        try:
            control_ids = self.driver.execute_script(script, control_type, properties)
            elements = []
            
            for control_id in control_ids:
                try:
                    element = self.driver.find_element(By.ID, control_id)
                    elements.append(element)
                except:
                    pass
                    
            return elements
        except:
            return []
    
    def close(self):
        """Cierra el navegador"""
        if self.driver:
            try:
                self.driver.quit()
                logger.info("Navegador cerrado correctamente")
                return True
            except Exception as e:
                logger.error(f"Error al cerrar el navegador: {e}")
                return False
        return True





class IssuesExtractor:
    """Clase principal para extraer issues de SAP con interfaz gráfica y base de datos"""

    def __init__(self):
        """Inicializa la clase"""
        self.excel_file_path = None
        self.driver = None
        
        # Variables para la GUI
        self.root = None
        self.status_var = None
        self.client_var = None
        self.project_var = None
        self.project_combo = None
        self.log_text = None
        self.excel_filename_var = None
        self.processing = False
        self.left_panel = None
        self.header_frame = None
        self.client_combo = None
        self.image_cache = {}
        
        # Componentes
        self.db_manager = DatabaseManager()
        self.excel_manager = ExcelManager()
        self.browser = SAPBrowser()
        
    def choose_excel_file(self):
        """Permite al usuario elegir un archivo Excel existente o crear uno nuevo"""
        file_path = self.excel_manager.select_file()
        self.excel_file_path = file_path
        self.excel_manager.file_path = file_path
        
        # Actualizar la interfaz si existe
        if hasattr(self, 'status_var') and self.status_var:
            self.status_var.set(f"Archivo Excel seleccionado: {os.path.basename(file_path)}")
        
        # Actualizar el nombre del archivo en la etiqueta
        if hasattr(self, 'excel_filename_var') and self.excel_filename_var:
            self.excel_filename_var.set(f"Archivo: {os.path.basename(file_path)}")
            
        return file_path
        
    def connect_to_browser(self):
        """Conecta con el navegador"""
        result = self.browser.connect()
        self.driver = self.browser.driver
        return result
        
    def update_excel(self, issues_data):
        """Actualiza el archivo Excel con los datos extraídos"""
        success, new_items, updated_items = self.excel_manager.update_with_issues(issues_data)
        
        # Actualizar la interfaz si existe
        if hasattr(self, 'status_var') and self.status_var:
            if success:
                self.status_var.set(f"Excel actualizado: {new_items} nuevos, {updated_items} actualizados")
            else:
                self.status_var.set("Error al actualizar Excel")
                
        # Mostrar mensaje de éxito
        if success and self.root:
            messagebox.showinfo(
                "Proceso Completado", 
                f"El archivo Excel ha sido actualizado correctamente.\n\n"
                f"Se han agregado {new_items} nuevos issues y actualizado {updated_items} issues existentes."
            )
            
        return success
        
    def run_extraction(self):
        """Ejecuta el proceso completo de extracción"""
        try:
            if not self.connect_to_browser():
                logger.error("Error al conectar con el navegador")
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set("Error al conectar con el navegador")
                    
                return False
                
            # Navegar automáticamente a la URL
            logger.info("Navegando automáticamente a la URL de SAP...")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Navegando a la URL de SAP...")
                
            if not self.browser.navigate_to_sap():
                logger.error("Error al navegar a la URL de SAP")
                return False
            
            # Si estamos en la GUI, el usuario debe seguir las instrucciones en pantalla
            if self.root:
                instructions = """
                Por favor, realice los siguientes pasos:
                
                1. Inicie sesión si es necesario
                2. Haga clic en 'Project Overview'
                3. Seleccione el cliente con ERP Number: {}
                4. Seleccione el proyecto con ID: {}
                5. Navegue a la pestaña 'Issues'
                
                Una vez en la tabla de issues, haga clic en 'Iniciar Extracción'
                """.format(self.client_var.get() if self.client_var else "1025541", 
                        self.project_var.get() if self.project_var else "20096444")
                
                messagebox.showinfo("Instrucciones de Navegación", instructions)
                
                # Actualizar la interfaz si existe
                if hasattr(self, 'status_var') and self.status_var:
                    self.status_var.set("Esperando a que el usuario navegue a la tabla de issues...")
                    
                # En la GUI, no continuamos automáticamente
                return True
            else:
                # En modo consola, mostrar instrucciones y esperar ENTER
                print("\n=== INSTRUCCIONES PARA NAVEGACIÓN ASISTIDA ===")
                print("Se ha navegado automáticamente a la URL de SAP.")
                print("Por favor, siga estos pasos:")
                print("   1. Inicie sesión si es necesario")
                print("   2. Haga clic en 'Project Overview'")
                print("   3. Seleccione el cliente con ERP Number: 1025541")
                print("   4. Seleccione el proyecto con ID: 20096444")
                print("   5. Navegue a la pestaña 'Issues'")
                
                print("\nUna vez que esté viendo la lista de issues,")
                input("presione ENTER para comenzar la extracción automática...\n")
                
                # Continuar con la extracción
                return self.perform_extraction()
                
        except Exception as e:
            logger.error(f"Error en el proceso de extracción: {e}")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error: {e}")
                
            return False
            
    def perform_extraction(self):
        """Método principal para ejecutar el proceso de extracción"""
        try:
            # Marcar como procesando
            self.processing = True
            
            logger.info("Comenzando proceso de extracción...")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Comenzando proceso de extracción...")
                if self.root:
                    self.root.update()
            
            # Verificación de si estamos en la página correcta
            in_issues_page = False
            
            # Estrategia 1: Buscar el texto "Issues (número)"
            try:
                issues_title_elements = self.driver.find_elements(
                    By.XPATH, 
                    "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                )
                if issues_title_elements:
                    logger.info(f"Página de Issues detectada por título: {issues_title_elements[0].text}")
                    in_issues_page = True
            except Exception as e:
                logger.debug(f"No se pudo detectar título de Issues: {e}")
            
            # Estrategia 2: Verificar si hay filas de datos visibles
            if not in_issues_page:
                issue_rows = self.browser.find_table_rows(highlight=False)
                if len(issue_rows) > 0:
                    logger.info(f"Se detectaron {len(issue_rows)} filas de datos que parecen issues")
                    in_issues_page = True
            
            # Estrategia 3: Verificar encabezados de columna típicos
            if not in_issues_page:
                try:
                    column_headers = self.driver.find_elements(
                        By.XPATH,
                        "//div[text()='Title'] | //div[text()='Type'] | //div[text()='Priority'] | //div[text()='Status']"
                    )
                    if len(column_headers) >= 3:
                        logger.info(f"Se detectaron encabezados de columna típicos de issues: {len(column_headers)}")
                        in_issues_page = True
                except Exception as e:
                    logger.debug(f"No se pudieron detectar encabezados de columna: {e}")
            
            # Si aún no estamos seguros, intentar hacer clic en la pestaña Issues
            if not in_issues_page:
                logger.warning("No se detectó la página de Issues. Intentando hacer clic en la pestaña...")
                
                tab_selectors = [
                    "//div[@role='tab' and contains(text(), 'Issues')]",
                    "//a[contains(text(), 'Issues')]",
                    "//li[contains(@class, 'tab') and contains(., 'Issues')]",
                    "//div[contains(@class, 'sapMITBItem') and contains(., 'Issues')]",
                    "//div[contains(@class, 'sapMITBItem')]//span[contains(text(), 'Issues')]/..",
                    "//*[contains(text(), 'Issues') and not(contains(text(), '('))]"
                ]
                
                issue_tab_found = False
                for selector in tab_selectors:
                    try:
                        issue_tabs = self.driver.find_elements(By.XPATH, selector)
                        if issue_tabs:
                            for tab in issue_tabs:
                                try:
                                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tab)
                                    time.sleep(1)
                                    
                                    try:
                                        self.driver.execute_script("arguments[0].click();", tab)
                                        logger.info(f"Clic en pestaña Issues realizado con JavaScript: {tab.text}")
                                        time.sleep(3)
                                        issue_tab_found = True
                                        break
                                    except:
                                        tab.click()
                                        logger.info(f"Clic en pestaña Issues realizado: {tab.text}")
                                        time.sleep(3)
                                        issue_tab_found = True
                                        break
                                except Exception as click_e:
                                    logger.debug(f"Error al hacer clic en pestaña: {click_e}")
                                    continue
                            
                            if issue_tab_found:
                                break
                    except Exception as e:
                        logger.debug(f"Error con selector {selector}: {e}")
                
                if issue_tab_found:
                    try:
                        issues_title_elements = self.driver.find_elements(
                            By.XPATH, "//div[contains(text(), 'Issues') and contains(text(), '(')]"
                        )
                        if issues_title_elements:
                            logger.info(f"Página de Issues detectada después de clic: {issues_title_elements[0].text}")
                            in_issues_page = True
                    except:
                        pass
            
            if not in_issues_page:
                logger.warning("No se pudo confirmar que estamos en la página de Issues, pero intentaremos extraer datos de todos modos.")
            
            # Intentar extracción con reintentos
            max_attempts = 3
            for attempt in range(max_attempts):
                try:
                    logger.info(f"Intento de extracción {attempt+1}/{max_attempts}")
                    
                    # Actualizar la interfaz si existe
                    if hasattr(self, 'status_var') and self.status_var:
                        self.status_var.set(f"Intento de extracción {attempt+1}/{max_attempts}...")
                    
                    issues_data = self.browser.extract_issues_data()
                    
                    if issues_data:
                        logger.info(f"Extracción exitosa: {len(issues_data)} issues encontrados")
                        
                        # Actualizar Excel con los datos extraídos
                        if self.excel_file_path:
                            self.update_excel(issues_data)
                        else:
                            logger.warning("No se ha seleccionado archivo Excel para guardar los datos")
                            
                            if hasattr(self, 'status_var') and self.status_var:
                                self.status_var.set("Advertencia: No se ha seleccionado archivo Excel")
                            
                            if self.root:
                                excel_path = self.choose_excel_file()
                                if excel_path:
                                    self.update_excel(issues_data)
                        
                        self.processing = False
                        return True
                    else:
                        logger.warning(f"No se encontraron issues en el intento {attempt+1}")
                        
                        # Si no es el último intento, esperar y reintentar
                        if attempt < max_attempts - 1:
                            logger.info("Esperando antes de reintentar...")
                            time.sleep(5)
                        else:
                            logger.error("Todos los intentos de extracción fallaron")
                            
                            if hasattr(self, 'status_var') and self.status_var:
                                self.status_var.set("Error: No se encontraron issues después de varios intentos")
                            
                            if self.root:
                                messagebox.showerror(
                                    "Error de Extracción", 
                                    "No se pudieron encontrar issues después de varios intentos. Verifique que está en la página correcta y que existen issues para extraer."
                                )
                            
                            self.processing = False
                            return False
                except Exception as e:
                    logger.error(f"Error en el intento {attempt+1}: {e}")
                    
                    # Si no es el último intento, esperar y reintentar
                    if attempt < max_attempts - 1:
                        logger.info("Esperando antes de reintentar...")
                        time.sleep(5)
                    else:
                        logger.error(f"Todos los intentos de extracción fallaron: {e}")
                        
                        if hasattr(self, 'status_var') and self.status_var:
                            self.status_var.set(f"Error de extracción: {e}")
                        
                        if self.root:
                            messagebox.showerror(
                                "Error de Extracción", 
                                f"Se produjo un error durante la extracción: {e}"
                            )
                        
                        self.processing = False
                        return False
            
            # Si llegamos aquí, todos los intentos fallaron
            logger.error("Extracción fallida después de varios intentos")
            
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set("Error: Extracción fallida")
            
            self.processing = False
            return False
            
        except Exception as e:
            logger.error(f"Error general en el proceso de extracción: {e}")
            
            # Actualizar la interfaz si existe
            if hasattr(self, 'status_var') and self.status_var:
                self.status_var.set(f"Error: {e}")
            
            self.processing = False
            return False









    def create_gui(self):
        """Crea una interfaz gráfica mejorada para la aplicación"""
        self.root = tk.Tk()
        self.root.title("SAP Recommendations Extractor")
        self.root.geometry("650x800")
        self.root.resizable(True, True)
        self.root.configure(bg=SAP_COLORS["light"])
        
        # Configurar icono de la aplicación
        logo_photo = None
        if 'PIL_AVAILABLE' in globals() and PIL_AVAILABLE:
            try:
                # Logo SAP como icono base64
                sap_logo_base64 = """
                iVBORw0KGgoAAAANSUhEUgAAADAAAAAwCAYAAABXAvmHAAAACXBIWXMAAAsTAAALEwEAmpwYAAAB
                g0lEQVR4nO2ZQU7DMBBFv1mcJbAhB4ADcIucgAsAYslpWKQSKnCHZoPEgnuAaFasqESrKE3qicdJ
                +0kjVXXi+T8ex3ZSoCiKotTDEMAUwBxAUrJmXjsCaJR94IeSnbXCVgvAPYAvQ/LWWnjtDUDXdjJr
                A5jvwFTZoWnVBrAImFhZFgBOygxx6VkOytzIYRpZ5jEGOA0sbwy70NTGQHcHkltfN7IxUBTYKPPY
                3kKriMmN8gxMYqLjPY6JdrjEk/SaiGqigYGixGSiYUNI3bQnZZ8npiamCiCDiZaeBbQ/LWBMfAJ4
                5LU7AF9oo0yyiW/eN36QxQvXL4t5YlkxDZwBuAbwxsOI18a89sXrFwDOY05i7J3iCUCnYK6T7x1j
                9A0MPPoGlhLb1jHPdLgmXw7uOU++tg88Lrn+w/UrwvVKcjuxLaEYA5L1VmcQs4UoO9TGLDPJeiuZ
                qDKAxNTaQMwea227kDQT
                """
                
                # Cargar icono desde base64
                logo_data = base64.b64decode(sap_logo_base64)
                logo_image = Image.open(BytesIO(logo_data))
                
                # Convertir a formato para tkinter
                logo_photo = ImageTk.PhotoImage(logo_image)
                
                # Establecer como icono de la ventana
                self.root.iconphoto(True, logo_photo)
            except Exception as e:
                logger.error(f"Error al cargar el icono: {e}")
            
        # Configurar estilo
        style = ttk.Style()
        style.configure('TCombobox', arrowsize=15)
        
        # Estilos para widgets
        style.configure(".", foreground=SAP_COLORS["text"])
        style.configure("TLabel", background=SAP_COLORS["light"], foreground=SAP_COLORS["text"], font=("Arial", 10, "bold"))
        style.configure("Header.TLabel", background=SAP_COLORS["light"], foreground=SAP_COLORS["secondary"], font=("Arial", 16, "bold"))
        style.configure("TLabelframe.Label", background=SAP_COLORS["light"], foreground=SAP_COLORS["text"], font=("Arial", 11, "bold"))
        style.configure("Primary.TButton", background=SAP_COLORS["primary"], foreground=SAP_COLORS["white"], font=("Arial", 10, "bold"))
        style.configure("Success.TButton", background=SAP_COLORS["success"], foreground=SAP_COLORS["white"], font=("Arial", 10, "bold"))
        style.configure("Danger.TButton", background=SAP_COLORS["danger"], foreground=SAP_COLORS["white"], font=("Arial", 10, "bold"))
        style.configure("TCombobox", selectbackground=SAP_COLORS["primary"], selectforeground=SAP_COLORS["white"], 
                        fieldbackground="white", background="white", foreground=SAP_COLORS["text"])
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Cabecera con logo
        self.header_frame = ttk.Frame(main_frame)
        self.header_frame.pack(fill=tk.X, pady=(0, 15))

        try:
            # Usar el logo si ya está cargado y PIL está disponible
            if 'PIL_AVAILABLE' in globals() and PIL_AVAILABLE and logo_photo:
                logo_label = tk.Label(self.header_frame, image=logo_photo, bg=SAP_COLORS["light"])
                logo_label.image = logo_photo  # Mantener referencia
                logo_label.pack(side=tk.LEFT, padx=(0, 10))
        except:
            pass
        
        # Título con fondo de alto contraste
        title_background = "#0A3D6E"  # Azul oscuro
        title_foreground = "#FFFFFF"  # Texto blanco
        
        title_label = tk.Label(
            self.header_frame, 
            text="Extractor de Recomendaciones SAP",
            font=("Arial", 18, "bold"),
            foreground=title_foreground,
            background=title_background,
            padx=8,
            pady=4
        )
        title_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Contenedor principal dividido
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Panel izquierdo para configuración
        self.left_panel = ttk.Frame(content_frame, padding=10, width=435)
        self.left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        self.left_panel.pack_propagate(False)  # Mantener ancho fijo
        
        # Sección de cliente
        client_frame = tk.LabelFrame(self.left_panel, 
                                text="Cliente", 
                                bg=SAP_COLORS["light"],
                                fg="#000000",
                                font=("Arial", 11, "bold"),
                                padx=10, pady=10)
        client_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta ERP
        tk.Label(client_frame, 
            text="ERP Number:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 9)).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Entry ERP
        self.client_var = tk.StringVar(value="1025541")
        client_entry = tk.Entry(client_frame, 
                            textvariable=self.client_var,
                            width=15,
                            font=("Arial", 10),
                            bg="white",
                            fg="black",
                            highlightbackground=SAP_COLORS["primary"],
                            highlightcolor=SAP_COLORS["primary"])
        client_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        # Etiqueta de clientes guardados
        tk.Label(client_frame, 
            text="Clientes guardados:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 9)).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Lista desplegable de clientes guardados
        client_list = self.db_manager.get_clients()
        self.client_combo = ttk.Combobox(client_frame, values=client_list, width=30)
        self.client_combo.config(state='readonly')
        self.client_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W+tk.E)
        self.client_combo.bind("<<ComboboxSelected>>", lambda e: self.select_client(self.client_combo.get()))

        # Sección de proyecto
        project_frame = tk.LabelFrame(self.left_panel, 
                                    text="Proyecto", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
        project_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta ID
        tk.Label(project_frame, 
            text="ID Proyecto:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10)).grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        










# Entry Proyecto
        self.project_var = tk.StringVar(value="20096444")
        project_entry = tk.Entry(project_frame, 
                            textvariable=self.project_var,
                            width=15,
                            font=("Arial", 10),
                            bg="white",
                            fg="black",
                            highlightbackground=SAP_COLORS["primary"],
                            highlightcolor=SAP_COLORS["primary"])
        project_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        # Etiqueta de proyectos
        tk.Label(project_frame, 
            text="Proyectos:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10)).grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        
        # Lista desplegable de proyectos guardados
        project_list = self.db_manager.get_projects("1025541")  # Proyectos para el cliente predeterminado
        self.project_combo = ttk.Combobox(project_frame, values=project_list, width=30)
        self.project_combo.config(state='readonly')
        self.project_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky=tk.W+tk.E)
        self.project_combo.bind("<<ComboboxSelected>>", lambda e: self.select_project(self.project_combo.get()))

        # Sección de navegador
        browser_frame = tk.LabelFrame(self.left_panel, 
                                    text="Navegador", 
                                    bg=SAP_COLORS["light"],
                                    fg="#000000",
                                    font=("Arial", 11, "bold"),
                                    padx=10, pady=10)
        browser_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta de navegador
        browser_label = tk.Label(
            browser_frame, 
            text="Iniciar un navegador con perfil dedicado:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        browser_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón de navegador
        browser_button = tk.Button(
            browser_frame, 
            text="Iniciar Navegador",
            command=self.start_browser,
            bg=SAP_COLORS["primary"],
            fg="#FFFFFF",
            activebackground="#0A3D6E",
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        browser_button.pack(fill=tk.X, pady=5)
        
        # Sección de archivo Excel
        excel_frame = tk.LabelFrame(self.left_panel, 
                                text="Archivo Excel", 
                                bg=SAP_COLORS["light"],
                                fg="#000000",
                                font=("Arial", 11, "bold"),
                                padx=10, pady=10)
        excel_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta de Excel
        excel_label = tk.Label(
            excel_frame, 
            text="Seleccione un archivo existente o cree uno nuevo:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        excel_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón de Excel
        excel_button = tk.Button(
            excel_frame, 
            text="Seleccionar o Crear Excel",
            command=self.choose_excel_file,
            bg=SAP_COLORS["success"],
            fg="#FFFFFF",
            activebackground="#085E2E",
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        excel_button.pack(fill=tk.X, pady=5)
        
        # Mostrar el nombre del archivo seleccionado
        self.excel_filename_var = tk.StringVar(value="No seleccionado")
        excel_file_label = tk.Label(
            excel_frame, 
            textvariable=self.excel_filename_var,
            bg=SAP_COLORS["light"],
            fg="#0A3D6E",
            font=("Arial", 9, "bold"),
            wraplength=200,
            anchor="w",
            justify="left"
        )
        excel_file_label.pack(fill=tk.X, pady=5)
        
        # Sección de acción
        action_frame = tk.LabelFrame(self.left_panel, 
                                text="Acciones", 
                                bg=SAP_COLORS["light"],
                                fg="#000000",
                                font=("Arial", 11, "bold"),
                                padx=10, pady=10)
        action_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Etiqueta de acción
        action_label = tk.Label(
            action_frame, 
            text="Extraer datos de issues desde SAP:",
            bg=SAP_COLORS["light"],
            fg="#000000",
            font=("Arial", 10),
            anchor="w",
            justify="left"
        )
        action_label.pack(fill=tk.X, pady=(0, 5))
        
        # Botón de extracción
        extract_button = tk.Button(
            action_frame, 
            text="Iniciar Extracción de Issues",
            command=self.start_extraction,
            bg=SAP_COLORS["warning"],
            fg="#FFFFFF",
            activebackground="#C25A00",
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        extract_button.pack(fill=tk.X, pady=5)
        
        # Separador visual
        separator = tk.Frame(action_frame, height=2, bg=SAP_COLORS["gray"])
        separator.pack(fill=tk.X, pady=10)
        
        # Botón de salir
        exit_button = tk.Button(
            action_frame, 
            text="Salir de la Aplicación",
            command=self.exit_app,
            bg=SAP_COLORS["danger"],
            fg="#FFFFFF",
            activebackground="#990000",
            activeforeground="#FFFFFF",
            font=("Arial", 10, "bold"),
            padx=10, pady=5
        )
        exit_button.pack(fill=tk.X, pady=5)
        
        # Panel derecho para logs
        right_panel = ttk.Frame(content_frame, padding=10, width=300)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        right_panel.pack_propagate(False)
        
        # Log frame
        log_frame = tk.LabelFrame(right_panel, 
                            text="Registro de Actividad", 
                            bg=SAP_COLORS["light"],
                            fg="#000000",
                            font=("Arial", 11, "bold"))
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        # Text widget para logs
        self.log_text = tk.Text(
            log_frame, 
            height=20, 
            wrap=tk.WORD, 
            bg="white",
            fg="black",
            font=("Consolas", 9),
            padx=5,
            pady=5,
            borderwidth=2,
            relief=tk.SUNKEN
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Colores para los logs
        self.log_text.tag_configure("INFO", foreground="black")
        self.log_text.tag_configure("WARNING", foreground="#CC6600")
        self.log_text.tag_configure("ERROR", foreground="#990000")
        self.log_text.tag_configure("DEBUG", foreground="#555555")
        
        # Scrollbar para el log
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Status bar
        self.status_var = tk.StringVar(value="Listo para iniciar")
        status_bar = tk.Label(
            self.root, 
            textvariable=self.status_var,
            fg="#000000",
            bg="#F0F0F0",
            relief=tk.SUNKEN, 
            anchor=tk.W, 
            padx=5,
            pady=2,
            font=("Arial", 10)
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Configurar logger para que también escriba en la GUI
        self.setup_gui_logger()
        
        # Manejar cierre de ventana
        self.root.protocol("WM_DELETE_WINDOW", self.exit_app)
        
        # Centrar la ventana en la pantalla
        self.root.update_idletasks()
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        size = tuple(int(_) for _ in self.root.geometry().split('+')[0].split('x'))
        x = screen_width/2 - size[0]/2
        y = screen_height/2 - size[1]/2
        self.root.geometry("%dx%d+%d+%d" % (size + (x, y)))
        
        # Cargar configuración guardada
        self.load_config()
        





    def setup_gui_logger(self):
        """Configura el logger para que también escriba en la GUI"""
        class TextHandler(logging.Handler):
            def __init__(self, text_widget):
                logging.Handler.__init__(self)
                self.text_widget = text_widget
            
            def emit(self, record):
                msg = self.format(record)
                def append():
                    self.text_widget.configure(state='normal')
                    
                    # Agregar marca de tiempo y nivel con color
                    time_str = msg.split(' - ')[0] + ' - '
                    level_str = record.levelname + ' - '
                    msg_content = msg.split(' - ', 2)[2] if len(msg.split(' - ')) > 2 else ""
                    
                    self.text_widget.insert(tk.END, time_str, "INFO")
                    self.text_widget.insert(tk.END, level_str, record.levelname)
                    self.text_widget.insert(tk.END, msg_content + '\n', record.levelname)
                    
                    self.text_widget.configure(state='disabled')
                    self.text_widget.yview(tk.END)
                    
                    # Limitar tamaño del log
                    self.limit_log_length()
                    
                # Llamar a append desde el hilo principal
                self.text_widget.after(0, append)
                
            def limit_log_length(self):
                """Limita la longitud del log para evitar consumo excesivo de memoria"""
                if float(self.text_widget.index('end-1c').split('.')[0]) > 1000:
                    self.text_widget.configure(state='normal')
                    self.text_widget.delete('1.0', '500.0')
                    self.text_widget.configure(state='disabled')
        
        # Crear handler para el widget Text
        text_handler = TextHandler(self.log_text)
        text_handler.setLevel(logging.INFO)
        text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        
        # Añadir el handler al logger
        logger.addHandler(text_handler)
        
        # Deshabilitar el widget
        self.log_text.configure(state='disabled')
    
    def select_client(self, client_string):
        """Maneja la selección de un cliente desde el combobox"""
        try:
            # Extraer el ERP number del string "1025541 - Nombre del cliente"
            erp_number = client_string.split(" - ")[0]
            self.client_var.set(erp_number)
            
            # Actualizar la lista de proyectos para este cliente
            projects = self.db_manager.get_projects(erp_number)
            self.project_combo['values'] = projects
            
            if projects:
                self.project_combo.current(0)
                self.select_project(projects[0])
                
            # Actualizar el uso de este cliente
            self.db_manager.update_client_usage(erp_number)
            
            logger.info(f"Cliente seleccionado: {client_string}")
            self.save_config()
        except Exception as e:
            logger.error(f"Error al seleccionar cliente: {e}")
    
    def select_project(self, project_string):
        """Maneja la selección de un proyecto desde el combobox"""
        try:
            # Extraer el ID del proyecto del string "20096444 - Nombre del proyecto"
            project_id = project_string.split(" - ")[0]
            self.project_var.set(project_id)
            
            # Actualizar el uso de este proyecto
            self.db_manager.update_project_usage(project_id)
            
            logger.info(f"Proyecto seleccionado: {project_string}")
            self.save_config()
        except Exception as e:
            logger.error(f"Error al seleccionar proyecto: {e}")
    
    def start_browser(self):
        """Inicia el navegador desde la interfaz gráfica"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                messagebox.showwarning("Proceso en curso", "Hay un proceso de extracción en curso.")
                return
                
            # Asegurarse de que no haya un navegador ya abierto
            if self.driver:
                messagebox.showinfo("Navegador ya iniciado", "El navegador ya está iniciado.")
                return
                
            # Actualizar la interfaz para mostrar que se está iniciando el navegador
            self.status_var.set("Iniciando navegador...")
            if self.root:
                self.root.update()
            
            # Iniciar el navegador en un hilo separado
            threading.Thread(target=self._start_browser_thread, daemon=True).start()
            
        except Exception as e:
            logger.error(f"Error al iniciar el navegador: {e}")
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", f"Error al iniciar el navegador: {e}")
    
    def _start_browser_thread(self):
        """Método para ejecutar la inicialización del navegador en un hilo separado"""
        try:
            if self.connect_to_browser():
                logger.info("Navegador iniciado")
                
                # Actualizar la interfaz en el hilo principal
                if self.root:
                    self.root.after(0, lambda: self.status_var.set("Navegador iniciado. Inicie la extracción cuando esté listo."))
                
                # Navegar a la URL de SAP
                self.browser.navigate_to_sap()
                
                # Mostrar instrucciones en el hilo principal
                if self.root:
                    self.root.after(0, self._show_navigation_instructions)
            else:
                if self.root:
                    self.root.after(0, lambda: self.status_var.set("Error al iniciar el navegador"))
                    self.root.after(0, lambda: messagebox.showerror("Error", "No se pudo iniciar el navegador. Revise el log para más detalles."))
        except Exception as e:
            logger.error(f"Error en hilo de navegador: {e}")
            if self.root:
                self.root.after(0, lambda: self.status_var.set(f"Error: {e}"))
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error al iniciar el navegador: {e}"))
    
    def _show_navigation_instructions(self):
        """Muestra instrucciones de navegación"""
        instructions = """
        Por favor, realice los siguientes pasos:
        
        1. Inicie sesión si es necesario
        2. Haga clic en 'Project Overview'
        3. Seleccione el cliente con ERP Number: {}
        4. Seleccione el proyecto con ID: {}
        5. Navegue a la pestaña 'Issues'
        
        Una vez en la tabla de issues, haga clic en 'Iniciar Extracción'
        """.format(self.client_var.get(), self.project_var.get())
        
        messagebox.showinfo("Instrucciones de Navegación", instructions)
    
    def start_extraction(self):
        """Inicia el proceso de extracción desde la interfaz gráfica"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                messagebox.showwarning("Proceso en curso", "Hay un proceso de extracción en curso.")
                return
                
            # Verificar que existe un archivo Excel seleccionado
            if not self.excel_file_path:
                messagebox.showwarning("Archivo Excel no seleccionado", "Debe seleccionar o crear un archivo Excel primero.")
                return
                
            # Verificar que el navegador está abierto
            if not self.driver:
                messagebox.showwarning("Navegador no iniciado", "Debe iniciar el navegador primero.")
                return
                
            # Iniciar extracción en un hilo separado para no bloquear la GUI
            threading.Thread(target=self.perform_extraction, daemon=True).start()
            
        except Exception as e:
            logger.error(f"Error al iniciar extracción: {e}")
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", f"Error al iniciar extracción: {e}")
    
    def exit_app(self):
        """Cierra la aplicación"""
        try:
            # Verificar si hay un proceso en curso
            if self.processing:
                confirm_exit = messagebox.askyesno(
                    "Proceso en curso", 
                    "Hay un proceso de extracción en curso. ¿Realmente desea salir?",
                    icon='warning'
                )
                if not confirm_exit:
                    return
                    
            if self.driver:
                try:
                    close_browser = messagebox.askyesno(
                        "Cerrar navegador", 
                        "¿Desea cerrar también el navegador?",
                        icon='question'
                    )
                    if close_browser:
                        self.browser.close()
                        logger.info("Navegador cerrado correctamente")
                except:
                    logger.warning("No se pudo cerrar el navegador correctamente")
            
            # Guardar configuración antes de salir
            self.save_config()
            
            self.root.destroy()
        except Exception as e:
            logger.error(f"Error al cerrar la aplicación: {e}")
            # En caso de error, forzar cierre
            self.root.destroy()
    
    def save_config(self):
        """Guarda la configuración actual"""
        try:
            config = {
                'client': self.client_var.get(),
                'project': self.project_var.get(),
                'excel_path': self.excel_file_path
            }
            
            config_dir = "config"
            if not os.path.exists(config_dir):
                os.makedirs(config_dir)
                
            config_path = os.path.join(config_dir, 'config.json')
            
            with open(config_path, 'w') as f:
                json.dump(config, f)
        except Exception as e:
            logger.error(f"Error al guardar configuración: {e}")
    
    def load_config(self):
        """Carga la configuración guardada"""
        try:
            config_path = os.path.join('config', 'config.json')
            
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
                    
                    if 'client' in config:
                        self.client_var.set(config['client'])
                        
                    if 'project' in config:
                        self.project_var.set(config['project'])
                        
                    if 'excel_path' in config and os.path.exists(config['excel_path']):
                        self.excel_file_path = config['excel_path']
                        self.excel_manager.file_path = config['excel_path']
                        self.excel_filename_var.set(f"Archivo: {os.path.basename(config['excel_path'])}")
                        
                    logger.info("Configuración cargada correctamente")
        except Exception as e:
            logger.error(f"Error al cargar configuración: {e}")
    
    def main_gui(self):
        """Punto de entrada principal con interfaz gráfica"""
        self.create_gui()
        self.root.mainloop()
        
        
        
        
        
        
        
def check_required_packages():
    """Verifica que estén instalados los paquetes requeridos"""
    required_packages = {
        "selenium": "Para automatización web",
        "pandas": "Para procesamiento de datos",
        "openpyxl": "Para manejo de archivos Excel"
    }
    
    missing = []
    for package, description in required_packages.items():
        try:
            __import__(package)
        except ImportError:
            missing.append(f"{package} ({description})")
    
    return missing

def create_shortcut(target_path, shortcut_path=None, icon_path=None):
    """Crea un acceso directo para la aplicación"""
    try:
        if not shortcut_path:
            desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')
            shortcut_path = os.path.join(desktop_path, "SAP Issues Extractor.lnk")
            
        if os.path.exists(shortcut_path):
            return shortcut_path
            
        import winshell
        from win32com.client import Dispatch
        
        shell = Dispatch('WScript.Shell')
        shortcut = shell.CreateShortCut(shortcut_path)
        shortcut.Targetpath = target_path
        shortcut.WorkingDirectory = os.path.dirname(target_path)
        if icon_path:
            shortcut.IconLocation = icon_path
        shortcut.save()
        
        return shortcut_path
    except Exception as e:
        logger.error(f"Error al crear acceso directo: {e}")
        return None

def main():
    """Función principal que ejecuta la aplicación"""
    extractor = None  # Inicializar la variable para evitar errores
    try:
        # Verificar paquetes requeridos
        missing_packages = check_required_packages()

        if missing_packages:
            print("Faltan las siguientes bibliotecas necesarias:")
            for package in missing_packages:
                print(f"  - {package}")
            print("\nPor favor, instálalas usando:")
            print(f"pip install {' '.join([p.split()[0] for p in missing_packages])}")
            input("\nPresiona ENTER para salir...")
            sys.exit(1)
        
        # Notificar sobre Pillow, pero no detener la ejecución
        try:
            __import__("PIL")
        except ImportError:
            print("Nota: La biblioteca Pillow no está disponible. Algunas características visuales estarán limitadas.")
            print("Si deseas instalarla, ejecuta: pip install Pillow")
        
        # Crear instancia del extractor
        extractor = IssuesExtractor()
        
        # Verificar si se desea interfaz gráfica o consola
        if len(sys.argv) > 1 and sys.argv[1] == "--console":
            # Modo consola
            extractor.choose_excel_file()
            extractor.run_extraction()
        else:
            # Modo interfaz gráfica (predeterminado)
            extractor.main_gui()
            
        logger.info("=== Proceso de extracción finalizado ===")
        
    except Exception as e:
        logger.critical(f"Error crítico en la ejecución: {e}")
        print(f"\n¡ERROR! Se ha producido un error crítico: {e}")
        print(f"Por favor, revisa el archivo de log para más detalles: {log_file}")
    finally:
        if extractor is not None and (not hasattr(extractor, 'root') or not extractor.root):
            # Solo mostrar mensaje final si estamos en modo consola
            input("\nPresiona ENTER para cerrar...")


if __name__ == "__main__":
    main()