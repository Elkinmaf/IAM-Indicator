# -*- coding: utf-8 -*-
"""
Script Mejorado para Extracción de Issues SAP con Mejor Manejo de Scroll
---
Versión con estrategia de scroll agresiva para capturar todas las filas
"""

import time
import pandas as pd
import os.path
import sys
import logging
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

# Configurar logging
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

log_file = os.path.join(log_dir, f"extraccion_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class IssuesExtractor:
    """Clase para extraer issues de SAP"""
    
    def __init__(self):
        """Inicializa la clase"""
        self.excel_file_path = None
        self.driver = None
    
    def choose_excel_file(self):
        """Permite al usuario elegir un archivo Excel existente o crear uno nuevo"""
        import tkinter as tk
        from tkinter import filedialog, messagebox
        
        logger.info("Solicitando al usuario opciones para el archivo Excel...")
        
        root = tk.Tk()
        root.withdraw()  # Ocultar la ventana principal de tkinter
        
        # Preguntar si desea usar un archivo existente o crear uno nuevo
        choice = messagebox.askquestion("Archivo Excel", 
                                        "¿Desea usar un archivo Excel existente?\n\n" +
                                        "Seleccione 'Sí' para elegir un archivo existente.\n" +
                                        "Seleccione 'No' para crear un nuevo archivo.")
        
        if choice == 'yes':
            # Permitir al usuario seleccionar un archivo Excel existente
            file_path = filedialog.askopenfilename(
                title="Seleccione el archivo Excel de seguimiento",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
            )
            
            if not file_path:  # El usuario canceló la selección
                logger.info("Usuario canceló la selección de archivo. Se creará uno nuevo.")
                # Crear un nombre de archivo por defecto con fecha y hora
                default_filename = f"Seguimiento_Recomendaciones_EPM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                # Permitir al usuario guardar con un nombre específico
                file_path = filedialog.asksaveasfilename(
                    title="Guardar nuevo archivo Excel",
                    defaultextension=".xlsx",
                    initialfile=default_filename,
                    filetypes=[("Archivos Excel", "*.xlsx")]
                )
                
                if not file_path:  # Si cancela de nuevo, usar el nombre por defecto
                    file_path = default_filename
                    logger.info(f"Se usará el nombre por defecto: {file_path}")
                
                # Crear un archivo Excel vacío con las columnas necesarias
                self._create_new_excel(file_path)
                logger.info(f"Creado nuevo archivo Excel: {file_path}")
        else:
            # Crear un nombre de archivo por defecto con fecha y hora
            default_filename = f"Seguimiento_Recomendaciones_EPM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            # Permitir al usuario guardar con un nombre específico
            file_path = filedialog.asksaveasfilename(
                title="Guardar nuevo archivo Excel",
                defaultextension=".xlsx",
                initialfile=default_filename,
                filetypes=[("Archivos Excel", "*.xlsx")]
            )
            
            if not file_path:  # Si cancela, usar el nombre por defecto
                file_path = default_filename
                logger.info(f"Se usará el nombre por defecto: {file_path}")
            
            # Crear un archivo Excel vacío con las columnas necesarias
            self._create_new_excel(file_path)
            logger.info(f"Creado nuevo archivo Excel: {file_path}")
        
        self.excel_file_path = file_path
        return file_path
    
    def _create_new_excel(self, file_path):
        """Crea un nuevo archivo Excel con las columnas necesarias"""
        try:
            # Crear un DataFrame vacío con las columnas necesarias
            columns = [
                'Title', 'Type', 'Priority', 'Status', 
                'Deadline', 'Due Date', 'Created By', 'Created On',
                'Last Updated', 'Comments'
            ]
            
            df = pd.DataFrame(columns=columns)
            
            # Guardar el DataFrame vacío como un archivo Excel
            df.to_excel(file_path, index=False)
            logger.info(f"Archivo Excel creado exitosamente: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error al crear nuevo archivo Excel: {e}")
            return False
    
    def connect_to_browser(self):
        """Conecta con el navegador abierto"""
        logger.info("Iniciando Chrome...")
        
        # Opciones mínimas, sin ocultar automatización para evitar problemas
        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 30)
            return True
        except Exception as e:
            logger.error(f"Error al iniciar Chrome: {e}")
            return False
    
    def scroll_to_load_all_items(self, total_expected=100, max_attempts=30):
        """
        Realiza scroll repetidamente para cargar todos los elementos
        Utiliza una estrategia más agresiva con diferentes técnicas de scroll
        """
        logger.info(f"Iniciando proceso de scroll para cargar aproximadamente {total_expected} elementos...")
        
        # Lista para almacenar las filas procesadas y evitar duplicados
        processed_titles = set()
        
        # Técnicas de scroll para probar
        scroll_techniques = [
            # Técnica 1: Usar scroll by
            lambda: self.driver.execute_script("window.scrollBy(0, 500);"),
            
            # Técnica 2: Usar scroll al final de la página
            lambda: self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);"),
            
            # Técnica 3: Usar PageDown
            lambda: self.driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN),
            
            # Técnica 4: Scroll a un elemento específico al final
            lambda: self.try_scroll_to_last_element(),
            
            # Técnica 5: Scroll con Arrow Down repetido
            lambda: self.driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ARROW_DOWN * 10)
        ]
        
        # Contadores y flags
        attempt = 0
        previous_rows_count = 0
        no_change_count = 0
        
        # Intentar diferentes técnicas de scroll
        while attempt < max_attempts:
            # Cambiar entre diferentes técnicas de scroll
            scroll_technique = scroll_techniques[attempt % len(scroll_techniques)]
            
            try:
                # Aplicar técnica de scroll
                scroll_technique()
                logger.info(f"Scroll intento #{attempt+1} usando técnica #{(attempt % len(scroll_techniques))+1}")
                
                # Importante: esperar a que carguen nuevos elementos
                time.sleep(2)
                
                # Cada 5 intentos, tomar una captura de pantalla
                if attempt % 5 == 0:
                    self.driver.save_screenshot(f"scroll_attempt_{attempt}.png")
                    logger.info(f"Captura de pantalla guardada: scroll_attempt_{attempt}.png")
                
                # Contar filas actuales
                rows = self.find_table_rows()
                current_rows_count = len(rows)
                
                logger.info(f"Filas visibles después del scroll: {current_rows_count}")
                
                # Si no hay cambio en el número de filas después de varios intentos, posiblemente hemos llegado al final
                if current_rows_count == previous_rows_count:
                    no_change_count += 1
                    if no_change_count >= 5:  # Si no hay cambios después de 5 intentos con diferentes técnicas
                        logger.info("No se detectaron más filas después de múltiples intentos. Posiblemente hemos cargado todas las filas disponibles.")
                        break
                else:
                    no_change_count = 0  # Resetear contador si hubo cambios
                
                previous_rows_count = current_rows_count
                attempt += 1
                
                # Si hemos cargado suficientes filas, terminar
                if current_rows_count >= total_expected:
                    logger.info(f"Se han cargado al menos {total_expected} filas. Finalizando scroll.")
                    break
            except Exception as e:
                logger.warning(f"Error durante el scroll en intento {attempt+1}: {e}")
                attempt += 1
        
        logger.info(f"Proceso de scroll completado después de {attempt} intentos. Filas visibles: {previous_rows_count}")
        return previous_rows_count
    
    def try_scroll_to_last_element(self):
        """Intenta hacer scroll al último elemento visible de la tabla"""
        try:
            rows = self.find_table_rows()
            if rows:
                last_row = rows[-1]
                self.driver.execute_script("arguments[0].scrollIntoView(false);", last_row)
                return True
        except Exception as e:
            logger.warning(f"Error al intentar scroll al último elemento: {e}")
            return False
    
    def find_table_rows(self):
        """Encuentra todas las filas de la tabla usando múltiples selectores"""
        # Lista para almacenar las filas encontradas
        all_rows = []
        
        # Intentar diferentes selectores
        selectors = [
            "//div[contains(@class, 'sapMListItems')]/div[contains(@class, 'sapMListItem')]",
            "//table/tbody/tr[not(contains(@class, 'sapMListHeaderSubTitleItems'))]",
            "//div[@role='row'][not(contains(@class, 'sapMListHeaderSubTitleItems'))]",
            "//div[contains(@class, 'sapMListItem')]",
            "//div[contains(@class, 'sapMLIB')]"
        ]
        
        for selector in selectors:
            try:
                rows = self.driver.find_elements(By.XPATH, selector)
                if len(rows) > 0:
                    all_rows = rows
                    logger.info(f"Se encontraron {len(rows)} filas con selector: {selector}")
                    break
            except Exception as e:
                logger.debug(f"No se encontraron filas con selector {selector}: {e}")
        
        return all_rows
    
    def extract_issues_data(self):
        """Extrae datos de issues desde la tabla con mejor manejo de scroll"""
        try:
            logger.info("Esperando a que cargue la tabla de issues...")
            
            # Esperar a que cargue la página inicial
            time.sleep(5)
            
            # Toma capturas de pantalla para verificar qué se está viendo
            self.driver.save_screenshot("pantalla_inicial.png")
            logger.info("Captura de pantalla guardada como 'pantalla_inicial.png'")
            
            # Intentar obtener el número total de issues desde el texto
            total_issues = 0
            try:
                issues_header_text = self.driver.find_element(By.XPATH, 
                    "//div[contains(text(), 'Issues') and contains(text(), '(')]").text
                logger.info(f"Texto encontrado para issues: {issues_header_text}")
                
                import re
                match = re.search(r'\((\d+)\)', issues_header_text)
                if match:
                    total_issues = int(match.group(1))
                    logger.info(f"Total de issues a procesar: {total_issues}")
                else:
                    logger.warning(f"No se pudo extraer el número de issues del texto: {issues_header_text}")
                    total_issues = 100  # valor predeterminado
            except Exception as e:
                logger.error(f"Error al obtener el número total de issues: {e}")
                total_issues = 100  # valor predeterminado
            
            # Hacer scroll para cargar todos los elementos
            self.scroll_to_load_all_items(total_issues)
            
            # Obtener todas las filas después del scroll
            logger.info("Extrayendo datos de todas las filas después del scroll...")
            rows = self.find_table_rows()
            
            if not rows:
                logger.error("No se pudieron encontrar filas en la tabla")
                self.driver.save_screenshot("error_no_rows.png")
                return []
            
            # Lista para almacenar los datos
            issues_data = []
            processed_count = 0
            seen_titles = set()  # Para evitar duplicados
            
            # Procesar cada fila
            logger.info(f"Procesando {len(rows)} filas...")
            for index, row in enumerate(rows):
                try:
                    # Para debugging: sacar screenshot de algunas filas
                    if index % 20 == 0:  # Cada 20 filas
                        self.driver.execute_script("arguments[0].scrollIntoView(true);", row)
                        time.sleep(0.5)
                        self.driver.save_screenshot(f"row_{index}.png")
                    
                    # Estructura específica para la tabla observada
                    # El título está en un enlace o en la primera celda
                    title = None
                    
                    # Intentar múltiples formas de extraer el título
                    title_extractors = [
                        lambda r: r.find_element(By.XPATH, ".//a").text,
                        lambda r: r.find_element(By.XPATH, ".//span[contains(@class, 'title')]").text,
                        lambda r: r.find_element(By.XPATH, ".//div[contains(@class, 'title')]").text,
                        lambda r: r.find_elements(By.XPATH, ".//div[@role='gridcell']")[0].text if r.find_elements(By.XPATH, ".//div[@role='gridcell']") else None,
                        lambda r: r.find_elements(By.XPATH, ".//td")[0].text if r.find_elements(By.XPATH, ".//td") else None,
                        lambda r: r.find_element(By.XPATH, ".//*[contains(@id, 'title')]").text,
                        lambda r: r.find_element(By.XPATH, ".//div[@title]").get_attribute("title"),
                        lambda r: r.find_element(By.XPATH, ".//span[@title]").get_attribute("title")
                    ]
                    
                    for extractor in title_extractors:
                        try:
                            title_text = extractor(row)
                            if title_text and title_text.strip():
                                title = title_text.strip()
                                break
                        except:
                            continue
                    
                    if not title:
                        logger.warning(f"No se pudo extraer título para la fila {index}, saltando...")
                        continue
                    
                    # Si ya procesamos este título, saltarlo
                    if title in seen_titles:
                        logger.debug(f"Título duplicado: '{title}', saltando...")
                        continue
                    
                    seen_titles.add(title)
                    
                    # Extraer otras columnas
                    type_text = "N/A"
                    priority = "N/A"
                    status = "N/A"
                    deadline = "N/A"
                    due_date = "N/A"
                    created_by = "N/A"
                    created_on = "N/A"
                    
                    # Intentar extraer celdas por diferentes métodos
                    cells = []
                    cell_extractors = [
                        lambda r: r.find_elements(By.XPATH, ".//div[@role='gridcell']"),
                        lambda r: r.find_elements(By.XPATH, ".//td"),
                        lambda r: r.find_elements(By.XPATH, "./div")
                    ]
                    
                    for extractor in cell_extractors:
                        try:
                            extracted_cells = extractor(row)
                            if extracted_cells and len(extracted_cells) > 1:
                                cells = extracted_cells
                                break
                        except:
                            continue
                    
                    # Extraer información de las celdas
                    if cells:
                        if len(cells) >= 2:
                            type_text = cells[1].text.strip() if cells[1].text.strip() else "N/A"
                        if len(cells) >= 3:
                            priority = cells[2].text.strip() if cells[2].text.strip() else "N/A"
                        if len(cells) >= 4:
                            status = cells[3].text.strip() if cells[3].text.strip() else "N/A"
                        if len(cells) >= 5:
                            deadline = cells[4].text.strip() if cells[4].text.strip() else "N/A"
                        if len(cells) >= 6:
                            due_date = cells[5].text.strip() if cells[5].text.strip() else "N/A"
                        if len(cells) >= 7:
                            created_by = cells[6].text.strip() if cells[6].text.strip() else "N/A"
                        if len(cells) >= 8:
                            created_on = cells[7].text.strip() if cells[7].text.strip() else "N/A"
                    
                    # Extraer estado desde una clase específica o atributo si no se pudo obtener de las celdas
                    if status == "N/A":
                        try:
                            status_elements = row.find_elements(By.XPATH, 
                                ".//div[contains(@class, 'status')] | .//span[contains(@class, 'status')]")
                            if status_elements:
                                status = status_elements[0].text.strip()
                        except:
                            pass
                    
                    # Datos del issue
                    issue_data = {
                        'Title': title,
                        'Type': type_text,
                        'Priority': priority,
                        'Status': status,
                        'Deadline': deadline,
                        'Due Date': due_date,
                        'Created By': created_by,
                        'Created On': created_on
                    }
                    
                    issues_data.append(issue_data)
                    processed_count += 1
                    
                    if processed_count % 10 == 0:
                        logger.info(f"Procesados {processed_count} issues hasta ahora")
                except Exception as e:
                    logger.error(f"Error al procesar la fila {index}: {e}")
            
            logger.info(f"Extracción completada. Total de issues procesados: {len(issues_data)}")
            
            # Guardar los datos crudos para depuración
            with open("issues_data_raw.txt", "w", encoding="utf-8") as f:
                for item in issues_data:
                    f.write(str(item) + "\n")
            
            return issues_data
        except Exception as e:
            logger.error(f"Error en la extracción de datos: {e}")
            self.driver.save_screenshot("error_extract.png")
            return []
    
    def update_excel(self, issues_data):
        """Actualiza el archivo Excel con los datos extraídos"""
        try:
            logger.info(f"Actualizando archivo Excel: {self.excel_file_path}...")
            
            if not issues_data:
                logger.warning("No hay datos para actualizar en Excel")
                return False
            
            # Leer el archivo Excel existente o crear uno nuevo
            if os.path.exists(self.excel_file_path):
                existing_df = pd.read_excel(self.excel_file_path)
                logger.info(f"Archivo Excel existente cargado con {len(existing_df)} registros.")
            else:
                existing_df = pd.DataFrame(columns=[
                    'Title', 'Type', 'Priority', 'Status', 
                    'Deadline', 'Due Date', 'Created By', 'Created On',
                    'Last Updated', 'Comments'
                ])
                logger.info("Creando nuevo archivo Excel.")
            
            # Convertir los datos extraídos a DataFrame
            new_df = pd.DataFrame(issues_data)
            
            # Inicializar contador de cambios
            new_items = 0
            updated_items = 0
            
            # Verificar Issues existentes y actualizarlas si cambiaron de estado
            updated_df = existing_df.copy()
            
            for _, new_row in new_df.iterrows():
                # Buscar si el issue ya existe en el DataFrame existente (por título)
                title_exists = False
                if 'Title' in existing_df.columns:
                    existing_row = existing_df[existing_df['Title'] == new_row['Title']]
                    title_exists = len(existing_row) > 0
                
                if not title_exists:
                    # Si no existe, agregar el nuevo issue
                    new_row_df = pd.DataFrame([new_row])
                    updated_df = pd.concat([updated_df, new_row_df], ignore_index=True)
                    new_items += 1
                    logger.info(f"Nuevo issue añadido: '{new_row['Title']}'")
                else:
                    # Si existe, verificar si el estado ha cambiado
                    if 'Status' in existing_row.columns and existing_row['Status'].values[0] != new_row['Status']:
                        # Actualizar la fila con el nuevo estado
                        mask = updated_df['Title'] == new_row['Title']
                        updated_df.loc[mask, 'Status'] = new_row['Status']
                        # También actualizar la fecha de última modificación
                        updated_df.loc[mask, 'Last Updated'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        updated_items += 1
                        logger.info(f"Actualizado estado de '{new_row['Title']}': '{existing_row['Status'].values[0]}' → '{new_row['Status']}'")
            
            # Guardar el DataFrame actualizado en el archivo Excel
            updated_df.to_excel(self.excel_file_path, index=False)
            logger.info(f"Archivo Excel actualizado correctamente: {self.excel_file_path}")
            logger.info(f"Resumen: {new_items} nuevos items, {updated_items} items actualizados")
            
            # Aplicar formato al archivo Excel para mejor presentación
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Cargar el workbook
                wb = load_workbook(self.excel_file_path)
                ws = wb.active
                
                # Aplicar formato a los encabezados
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Aplicar bordes a todas las celdas
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                
                # Formatear encabezados
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Formatear datos
                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                        
                        # Colorear celdas según el Status
                        if col == 4:  # Columna Status
                            status = cell.value.upper() if cell.value else ""
                            
                            if "DONE" in status:
                                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                            elif "OPEN" in status:
                                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                            elif "READY" in status:
                                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                
                # Ajustar anchos de columna
                for col in range(1, ws.max_column + 1):
                    # Obtener la longitud máxima del contenido en esa columna
                    max_length = 0
                    for row in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    
                    # Aplicar ancho de columna (con un mínimo y un máximo razonable)
                    adjusted_width = max(10, min(50, max_length + 2))
                    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = adjusted_width
                
                # Guardar con formato
                wb.save(self.excel_file_path)
                logger.info("Formato aplicado al archivo Excel correctamente")
            except Exception as format_e:
                logger.warning(f"No se pudo aplicar formato al Excel: {format_e}")
            
            return True
        except Exception as e:
            logger.error(f"Error al actualizar el archivo Excel: {e}")
            return False
    
    def run_extraction(self):
        """Ejecuta el proceso completo de extracción"""
        try:
            # Conectar con el navegador
            if not self.connect_to_browser():
                logger.error("Error al conectar con el navegador")
                return False
            
            print("\n=== INSTRUCCIONES PARA NAVEGACIÓN MANUAL ===")
            print("1. Se abrirá una ventana de Chrome.")
            print("2. Por favor, realice los siguientes pasos:")
            print("   a. Navegue a: https://xalm-prod.x.eu20.alm.cloud.sap/launchpad#sdwork-center&/")
            print("   b. Inicie sesión si es necesario y acepte cualquier certificado")
            print("   c. Haga clic en 'Project Overview'")
            print("   d. Seleccione el cliente con ERP Number: 1025541")
            print("   e. Seleccione el proyecto con ID: 20096444")
            print("   f. Navegue a la pestaña 'Issues'")
            print("\nUna vez que esté viendo la lista de issues,")
            input("presione ENTER para comenzar la extracción automática...\n")
            
            # Paso 1: Extraer datos
            logger.info("Comenzando extracción de datos...")
            issues_data = self.extract_issues_data()
            
            if not issues_data:
                logger.warning("No se obtuvieron datos para procesar")
                return False
            
            logger.info(f"Se extrajeron {len(issues_data)} issues")
            
            # Paso 2: Actualizar Excel
            if self.update_excel(issues_data):
                logger.info("Archivo Excel actualizado correctamente")
                return True
            else:
                logger.error("Error al actualizar el archivo Excel")
                return False
        except Exception as e:
            logger.error(f"Error en el proceso de extracción: {e}")
            return False
        finally:
            # Preguntar si cerrar el navegador
            if self.driver:
                close_browser = input("¿Desea cerrar el navegador? (S/N): ").strip().upper() == 'S'
                if close_browser:
                    try:
                        self.driver.quit()
                        logger.info("Navegador cerrado correctamente")
                    except:
                        logger.warning("No se pudo cerrar el navegador correctamente")
                else:
                    logger.info("El navegador permanece abierto")

def main():
    try:
        # Crear instancia del extractor
        extractor = IssuesExtractor()
        
        # Seleccionar archivo Excel
        extractor.choose_excel_file()
        
        # Ejecutar el proceso de extracción
        extractor.run_extraction()
        
        logger.info("=== Proceso de extracción finalizado ===")
        print("\nProceso completado. Consulta el archivo de log para detalles:", log_file)
    except Exception as e:
        logger.critical(f"Error crítico en la ejecución: {e}")
        print(f"\n¡ERROR! Se ha producido un error crítico: {e}")
        print(f"Por favor, revisa el archivo de log para más detalles: {log_file}")
    finally:
        input("\nPresiona ENTER para cerrar...")

if __name__ == "__main__":
    # Verificar instalación de dependencias
    required_packages = ["selenium", "pandas", "openpyxl", "tkinter"]

    missing_packages = []
    for package in required_packages:
        try:
            if package == "tkinter":
                import tkinter
            else:
                __import__(package)
        except ImportError:
            missing_packages.append(package)
    
    if missing_packages:
        print("Faltan las siguientes bibliotecas necesarias:")
        for package in missing_packages:
            print(f"  - {package}")
        print("\nPor favor, instálalas usando:")
        print(f"pip install {' '.join(missing_packages)}")
        sys.exit(1)
    
    main()